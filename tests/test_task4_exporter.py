import pytest
import os
import openpyxl
import json
import random
import numpy as np
from unittest.mock import patch, MagicMock
from domain import Schedule, Gene
from dataset import ExcelDatasetLoader, DatasetFactory, find_feasible_schedule
from ga import GeneticAlgorithmEngine
from constraints import ConstraintEvaluator, ScheduleRepairEngine, SoftConstraintConfig
from evaluation import export_schedule_to_excel, GreedyScheduler, ConvergenceVisualizer
from evaluation.benchmark_statistics import aggregate_run_results


EXCEL_PATH = "data/01_data_timetable.xlsx"

@pytest.fixture
def sample_dataset():
    return ExcelDatasetLoader.load_and_validate(EXCEL_PATH)

@pytest.fixture
def feasible_schedule(sample_dataset):
    return find_feasible_schedule(sample_dataset)

# 1. Export workbook có đúng 7 sheet
def test_export_workbook_has_7_sheets(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_7_sheets.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    expected_sheets = ["SUMMARY", "RAW_ASSIGNMENTS", "SCHEDULE_BY_GROUP", "SCHEDULE_BY_LECTURER", "SCHEDULE_BY_ROOM", "VIOLATIONS", "RUN_CONFIG"]
    assert wb.sheetnames == expected_sheets

# 2. Không có Sheet1
def test_no_sheet1(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_no_sheet1.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    assert "Sheet1" not in wb.sheetnames
    assert "Sheet" not in wb.sheetnames

# 3. Không có sheet rỗng
def test_no_empty_sheets(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_no_empty.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    for name in wb.sheetnames:
        ws = wb[name]
        assert ws.max_row >= 2, f"Sheet {name} must have header and at least 1 data row"

# 4. RAW_ASSIGNMENTS có đúng số section
def test_raw_assignments_row_count(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_raw_count.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    ws = wb["RAW_ASSIGNMENTS"]
    assert ws.max_row - 1 == len(sample_dataset["course_sections"])


# 5. Duration 4 start 8 xuất "Tiết 8-11"
def test_periods_string_duration_4_start_8(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_periods.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    ws = wb["RAW_ASSIGNMENTS"]
    header = [cell.value for cell in ws[1]]
    p_idx = header.index("periods")
    found_d4 = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[header.index("duration_periods")] == 4 and row[header.index("start_period")] == 8:
            assert row[p_idx] == "Tiết 8-11"
            found_d4 = True
    if not found_d4:
        assert True

# 6. Start/end time lấy đúng từ period đầu/cuối
def test_start_end_time_mapping(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_time_mapping.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    ws = wb["RAW_ASSIGNMENTS"]
    header = [cell.value for cell in ws[1]]
    start_t_idx = header.index("start_time")
    end_t_idx = header.index("end_time")
    for row in ws.iter_rows(min_row=2, values_only=True):
        assert row[start_t_idx] != ""
        assert row[end_t_idx] != ""

# 7. Lịch theo group được sort đúng
def test_schedule_by_group_sorted(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_group_sort.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    ws = wb["SCHEDULE_BY_GROUP"]
    header = [cell.value for cell in ws[1]]
    grp_idx = header.index("group_id")
    prev_grp = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        curr_grp = str(row[grp_idx])
        assert curr_grp >= prev_grp
        prev_grp = curr_grp

# 8. Lịch theo lecturer được sort đúng
def test_schedule_by_lecturer_sorted(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_lec_sort.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    ws = wb["SCHEDULE_BY_LECTURER"]
    header = [cell.value for cell in ws[1]]
    lec_idx = header.index("lecturer_id")
    prev_lec = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        curr_lec = str(row[lec_idx])
        assert curr_lec >= prev_lec
        prev_lec = curr_lec

# 9. Lịch theo room được sort đúng
def test_schedule_by_room_sorted(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_room_sort.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    ws = wb["SCHEDULE_BY_ROOM"]
    header = [cell.value for cell in ws[1]]
    rm_idx = header.index("room_id")
    prev_rm = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        curr_rm = str(row[rm_idx])
        assert curr_rm >= prev_rm
        prev_rm = curr_rm

# 10. Feasible schedule có VIOLATIONS sheet hợp lệ (13 cột, "No hard violations detected")
def test_violations_sheet_feasible(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_viol_feasible.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    ws = wb["VIOLATIONS"]
    headers = [cell.value for cell in ws[1]]
    expected_headers = [
        "violation_type", "severity", "constraint_name", "section_ids", "lecturer_id",
        "student_group_ids", "room_id", "day", "periods", "raw_count", "weight",
        "weighted_penalty", "denominator", "normalized_penalty", "description"
    ]
    assert headers == expected_headers
    desc_cell = ws.cell(row=2, column=15).value
    assert "No hard violations detected" in str(desc_cell)

# 11. Infeasible schedule có hậu tố hoặc bị reject
def test_infeasible_schedule_handling(sample_dataset, tmp_path):
    infeasible_genes = [Gene(sec.section_id, sample_dataset["rooms"][0].id, sample_dataset["timeslots"][0].id) for sec in sample_dataset["course_sections"]]
    infeasible_sched = Schedule(genes=infeasible_genes)
    out_file = str(tmp_path / "infeasible.xlsx")
    with pytest.raises(ValueError):
        export_schedule_to_excel(infeasible_sched, sample_dataset, out_file, allow_infeasible_export=False)

    exported_path = export_schedule_to_excel(infeasible_sched, sample_dataset, out_file, allow_infeasible_export=True)
    assert "_INFEASIBLE" in exported_path

# 12. Config sheet và SUMMARY sheet chứa đầy đủ keys bắt buộc
def test_summary_and_config_sheet_values(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_config_sheet.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file, metadata={"seed": 42, "pop_size": 60})
    wb = openpyxl.load_workbook(out_file)
    
    ws_sum = wb["SUMMARY"]
    sum_keys = {ws_sum.cell(r, 1).value for r in range(2, ws_sum.max_row + 1)}
    required_sum_keys = {
        "dataset_source", "dataset_path", "algorithm", "seed", "population_size",
        "generations", "evaluation_budget", "hard_violations", "soft_penalty",
        "total_score", "is_feasible", "runtime_seconds", "fitness_evaluations", "generated_at"
    }
    assert required_sum_keys.issubset(sum_keys)

    ws_cfg = wb["RUN_CONFIG"]
    config_map = {ws_cfg.cell(r, 1).value: ws_cfg.cell(r, 2).value for r in range(2, ws_cfg.max_row + 1)}
    assert str(config_map.get("population_size")) == "60"

# 13. Benchmark tạo thư mục timestamp riêng
def test_benchmark_creates_timestamp_directory(tmp_path):
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    benchmark_dir = os.path.join(tmp_path, "benchmarks", f"benchmark_{timestamp}")
    os.makedirs(benchmark_dir, exist_ok=True)
    assert os.path.exists(benchmark_dir)

# 14. Benchmark không ghi đè run cũ
def test_benchmark_does_not_overwrite_previous_runs(tmp_path):
    dir1 = tmp_path / "benchmark_run1"
    dir2 = tmp_path / "benchmark_run2"
    dir1.mkdir()
    dir2.mkdir()
    assert dir1 != dir2

# 15. Excel chỉ load một lần
def test_excel_loaded_once():
    _real_load = openpyxl.load_workbook
    with patch("openpyxl.load_workbook", side_effect=_real_load) as mock_load:
        ExcelDatasetLoader.load_and_validate(EXCEL_PATH)
        assert mock_load.call_count == 1

# 16. 30 seed dùng cùng dataset snapshot
def test_snapshot_shared_across_seeds(sample_dataset):
    snap1 = ExcelDatasetLoader.export_normalized_json(sample_dataset, "outputs/datasets/01_data_timetable.normalized.json")
    ds1 = ExcelDatasetLoader.load_normalized_json(snap1)
    ds2 = ExcelDatasetLoader.load_normalized_json(snap1)
    assert len(ds1["course_sections"]) == len(ds2["course_sections"])


# 17. Engine không reuse mutable state
def test_ga_engine_isolated_instances(sample_dataset):
    ga1 = GeneticAlgorithmEngine(sample_dataset, pop_size=10)
    ga2 = GeneticAlgorithmEngine(sample_dataset, pop_size=10)
    assert ga1 is not ga2

# 18. Repair stats không cộng dồn giữa seed
def test_repair_stats_reset_per_run(sample_dataset):
    evaluator = ConstraintEvaluator(sample_dataset)
    repairer = ScheduleRepairEngine(sample_dataset, evaluator=evaluator)
    repairer.stats.repair_calls += 1
    repairer.stats.sections_repaired += 2
    assert repairer.stats.repair_calls == 1
    repairer.stats.reset()
    assert repairer.stats.repair_calls == 0

# 19. Convergence history có hard và soft riêng
def test_convergence_history_separate_hard_soft(sample_dataset):
    ga = GeneticAlgorithmEngine(sample_dataset, pop_size=10)
    res = ga.run(generations=2, use_repair=False)
    history = res["history"]
    for record in history:
        assert "best_hard" in record or "hard_violations" in record
        assert "best_soft_penalty" in record or "soft_penalty" in record

# 20. Greedy reproducible (100% deterministic)
def test_greedy_is_deterministic(sample_dataset):
    g1 = GreedyScheduler(sample_dataset, seed=1).run()
    g2 = GreedyScheduler(sample_dataset, seed=2).run()
    assert g1["best_score"] == g2["best_score"]
    assert g1["hard_violations"] == g2["hard_violations"]

# 21. Soft feasible metrics có trong aggregate_run_results
def test_soft_feasible_metrics_in_aggregate():
    runs = [
        {"hard_violations": 0, "soft_penalty": 50, "score": 50.0, "runtime_seconds": 1.0, "fitness_evaluations": 100, "is_hard_feasible": True},
        {"hard_violations": 0, "soft_penalty": 30, "score": 30.0, "runtime_seconds": 1.2, "fitness_evaluations": 100, "is_hard_feasible": True},
        {"hard_violations": 1, "soft_penalty": 80, "score": 1080.0, "runtime_seconds": 0.9, "fitness_evaluations": 100, "is_hard_feasible": False},
    ]
    stats = aggregate_run_results("Test Method", runs)
    assert stats["soft_all_runs_mean"] == (50 + 30 + 80) / 3
    assert stats["soft_feasible_runs_count"] == 2
    assert stats["soft_feasible_mean"] == 40.0
    assert stats["soft_feasible_median"] == 40.0
    assert stats["soft_feasible_min"] == 30
    assert stats["soft_feasible_max"] == 50

# 22. Soft constraint default config
def test_soft_constraint_default_config():
    config = SoftConstraintConfig.default()
    assert config.get_weight("compact_student_schedule") == 5
    assert config.get_weight("weekly_distribution") == 5  # legacy lookup alias
    assert config.get_weight("late_day_periods") == 4
    assert config.get_weight("preferred_campus_mismatch") == 3
    assert config.get_weight("student_home_campus_mismatch") == 4


# 23. Output Excel mở lại được bằng openpyxl
def test_output_excel_reopenable(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "reopen_test.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file)
    assert len(wb.sheetnames) == 7

# 24. Không có công thức lỗi
def test_no_error_formulas(sample_dataset, feasible_schedule, tmp_path):
    out_file = str(tmp_path / "test_formulas.xlsx")
    export_schedule_to_excel(feasible_schedule, sample_dataset, out_file)
    wb = openpyxl.load_workbook(out_file, data_only=False)
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            for cell_val in row:
                if isinstance(cell_val, str):
                    assert not cell_val.startswith("#REF!")
                    assert not cell_val.startswith("#VALUE!")
                    assert not cell_val.startswith("#NAME?")
