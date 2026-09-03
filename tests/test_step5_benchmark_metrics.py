"""Tests for Step 5 — Benchmark Metrics Accuracy, Fairness & Reproducibility."""

import io
import pytest
import openpyxl

from domain import Room, CourseSection, Timeslot, Schedule, Gene, StudentGroup, Lecturer
from dataset import DatasetFactory, create_theory_timeslots
from constraints import ConstraintEvaluator, ScheduleRepairEngine, SoftConstraintConfig
from evaluation.run_metrics import EvaluationCounters, RunMetrics, RepairStatus
from evaluation.benchmark_statistics import aggregate_run_results
from evaluation.schedule_exporter import export_schedule_to_excel
from ga import GeneticAlgorithmEngine
from evaluation.baselines import GreedyScheduler, RandomSearchScheduler


# ============================================================
# 15.1. Kiểm thử đơn vị EvaluationCounters
# ============================================================

@pytest.mark.unit
def test_evaluation_counters_initialization_and_reset():
    counters = EvaluationCounters()
    assert counters.search_fitness_evaluations == 0
    assert counters.hard_constraint_evaluations == 0
    assert counters.soft_constraint_evaluations == 0
    assert counters.candidate_checks == 0
    assert counters.total_constraint_evaluations == 0

    counters.search_fitness_evaluations += 5
    counters.hard_constraint_evaluations += 3
    counters.soft_constraint_evaluations += 4
    counters.candidate_checks += 10

    assert counters.total_constraint_evaluations == 7

    snap = counters.snapshot()
    assert snap.search_fitness_evaluations == 5
    assert snap.hard_constraint_evaluations == 3
    assert snap.soft_constraint_evaluations == 4
    assert snap.candidate_checks == 10

    counters.reset()
    assert counters.search_fitness_evaluations == 0
    assert counters.hard_constraint_evaluations == 0
    assert counters.total_constraint_evaluations == 0
    # Bản chụp không thay đổi
    assert snap.search_fitness_evaluations == 5


@pytest.mark.unit
def test_two_counters_independent():
    c1 = EvaluationCounters()
    c2 = EvaluationCounters()
    c1.search_fitness_evaluations += 10
    c2.search_fitness_evaluations += 20
    assert c1.search_fitness_evaluations == 10
    assert c2.search_fitness_evaluations == 20


# ============================================================
# 15.2. Đo lường ConstraintEvaluator (không đếm hai lần)
# ============================================================

@pytest.mark.unit
def test_evaluator_no_double_count(small_dataset):
    evaluator = ConstraintEvaluator(small_dataset)
    sections = small_dataset["course_sections"]
    rooms = small_dataset["rooms"]
    genes = [Gene(s.section_id, rooms[0].id, 0) for s in sections]
    sched = Schedule(genes=genes)

    # 1. evaluate_unified -> +1 hard, +1 soft
    evaluator.evaluate_unified(sched)
    assert evaluator.counters.hard_constraint_evaluations == 1
    assert evaluator.counters.soft_constraint_evaluations == 1
    assert evaluator.counters.total_constraint_evaluations == 2
    assert evaluator.counters.search_fitness_evaluations == 0

    # 2. calculate_fitness(is_search_eval=True) -> +1 tìm kiếm, +1 hard, +1 soft
    evaluator.calculate_fitness(sched, is_search_eval=True)
    assert evaluator.counters.search_fitness_evaluations == 1
    assert evaluator.counters.hard_constraint_evaluations == 2
    assert evaluator.counters.soft_constraint_evaluations == 2
    assert evaluator.counters.total_constraint_evaluations == 4


# ============================================================
# 15.3. Kiểm thử thời gian đến lời giải khả thi đầu tiên
# ============================================================

@pytest.mark.unit
def test_ga_engine_records_time_to_first_feasible(small_dataset):
    engine = GeneticAlgorithmEngine(small_dataset, pop_size=10, seed=42)
    res = engine.run(generations=10, use_repair=True, seed=42)
    metrics: RunMetrics = res["run_metrics"]

    if metrics.feasible:
        assert metrics.time_to_first_feasible_seconds is not None
        assert metrics.time_to_first_feasible_seconds >= 0.0
        assert metrics.first_feasible_generation is not None
        assert metrics.first_feasible_search_evaluation is not None
        assert metrics.first_feasible_total_constraint_evaluation is not None
    else:
        assert metrics.time_to_first_feasible_seconds is None
        assert metrics.first_feasible_generation is None
        assert metrics.first_feasible_search_evaluation is None


# ============================================================
# 15.4. Kiểm thử trạng thái và thống kê Repair
# ============================================================

@pytest.mark.unit
def test_repair_status_classification(small_dataset):
    repairer = ScheduleRepairEngine(small_dataset)
    sections = small_dataset["course_sections"]
    rooms = small_dataset["rooms"]
    timeslots = small_dataset["timeslots"]

    # Tạo lịch
    genes = [Gene(s.section_id, rooms[0].id, timeslots[0].id) for s in sections]
    sched = Schedule(genes=genes)

    res = repairer.repair(sched)
    assert res.status in (RepairStatus.IMPROVED, RepairStatus.UNCHANGED, RepairStatus.FAILED)
    assert repairer.stats.repair_calls == (
        repairer.stats.repair_improved + repairer.stats.repair_unchanged + repairer.stats.repair_failed
    )
    assert res.success == (res.status != RepairStatus.FAILED)


# ============================================================
# 15.5. Kiểm thử số lần kiểm tra ứng viên (Greedy và Repair)
# ============================================================

@pytest.mark.unit
def test_greedy_candidate_checks(small_dataset):
    greedy = GreedyScheduler(small_dataset, seed=0)
    res = greedy.run(seed=0)
    metrics: RunMetrics = res["run_metrics"]

    assert metrics.candidate_checks > 0
    assert metrics.search_fitness_evaluations == 1
    assert metrics.total_constraint_evaluations > 0


@pytest.mark.unit
def test_repair_engine_candidate_checks(small_dataset):
    repairer = ScheduleRepairEngine(small_dataset)
    sections = small_dataset["course_sections"]
    rooms = small_dataset["rooms"]
    timeslots = small_dataset["timeslots"]

    # Gán mọi lớp vào cùng phòng/khung giờ để buộc phát sinh xung đột và tìm kiếm sửa lỗi
    genes = [Gene(s.section_id, rooms[0].id, timeslots[0].id) for s in sections]
    sched = Schedule(genes=genes)

    repairer.repair(sched)
    assert repairer.stats.candidate_checks > 0


# ============================================================
# 15.6. Kiểm thử khả năng tái lập theo seed
# ============================================================

@pytest.mark.unit
def test_ga_engine_reproducibility(small_dataset):
    engine1 = GeneticAlgorithmEngine(small_dataset, pop_size=10, seed=123)
    res1 = engine1.run(generations=5, use_repair=True, seed=123)

    engine2 = GeneticAlgorithmEngine(small_dataset, pop_size=10, seed=123)
    res2 = engine2.run(generations=5, use_repair=True, seed=123)

    m1: RunMetrics = res1["run_metrics"]
    m2: RunMetrics = res2["run_metrics"]

    assert m1.final_hard_violations == m2.final_hard_violations
    assert m1.final_soft_penalty == m2.final_soft_penalty
    assert m1.search_fitness_evaluations == m2.search_fitness_evaluations
    assert m1.total_constraint_evaluations == m2.total_constraint_evaluations
    assert m1.candidate_checks == m2.candidate_checks
    assert m1.repair_calls == m2.repair_calls


@pytest.mark.unit
def test_random_search_reproducibility(small_dataset):
    s1 = RandomSearchScheduler(small_dataset, seed=99)
    res1 = s1.run(iterations=50, seed=99)

    s2 = RandomSearchScheduler(small_dataset, seed=99)
    res2 = s2.run(iterations=50, seed=99)

    assert res1["hard_violations"] == res2["hard_violations"]
    assert res1["soft_penalty"] == res2["soft_penalty"]
    assert res1["fitness_evaluations"] == res2["fitness_evaluations"]


# ============================================================
# 15.7. Kiểm thử bộ tổng hợp (xử lý TTFF bằng None)
# ============================================================

@pytest.mark.unit
def test_aggregate_run_results_handles_none_ttff():
    runs = [
        {
            "score": 100, "hard_violations": 0, "soft_penalty": 100, "runtime_seconds": 1.5,
            "fitness_evaluations": 500, "search_fitness_evaluations": 500,
            "hard_constraint_evaluations": 500, "soft_constraint_evaluations": 500,
            "total_constraint_evaluations": 1000, "candidate_checks": 20,
            "repair_calls": 5, "repair_improved": 3, "repair_unchanged": 1, "repair_failed": 1,
            "time_to_first_feasible_seconds": 0.5, "first_feasible_generation": 2,
            "first_feasible_search_evaluation": 100, "is_hard_feasible": True, "is_perfect": False
        },
        {
            "score": 200, "hard_violations": 2, "soft_penalty": 200, "runtime_seconds": 2.0,
            "fitness_evaluations": 500, "search_fitness_evaluations": 500,
            "hard_constraint_evaluations": 500, "soft_constraint_evaluations": 500,
            "total_constraint_evaluations": 1000, "candidate_checks": 15,
            "repair_calls": 5, "repair_improved": 0, "repair_unchanged": 2, "repair_failed": 3,
            "time_to_first_feasible_seconds": None, "first_feasible_generation": None,
            "first_feasible_search_evaluation": None, "is_hard_feasible": False, "is_perfect": False
        },
    ]

    agg = aggregate_run_results("Test Method", runs)
    assert agg["feasible_count"] == 1
    assert agg["feasible_rate"] == 0.5
    # Trung vị TTFF phải là 0,5 (từ lần chạy khả thi duy nhất), KHÔNG phải 0,0 hay trung bình với 0
    assert agg["time_to_first_feasible_median"] == 0.5
    assert agg["repair_improved_total"] == 3
    assert agg["repair_unchanged_total"] == 3
    assert agg["repair_failed_total"] == 4


# ============================================================
# 15.8. Kiểm thử lượt xuất-nạp Excel (trang tính RUN_METRICS)
# ============================================================

@pytest.mark.unit
def test_excel_export_run_metrics_sheet(tmp_path, small_dataset):
    from dataset import find_feasible_schedule

    sched = find_feasible_schedule(small_dataset)
    assert sched is not None

    m = RunMetrics(
        method="Hybrid GA + Repair", seed=0, runtime_seconds=1.234,
        time_to_first_feasible_seconds=0.456, search_fitness_evaluations=500,
        hard_constraint_evaluations=600, soft_constraint_evaluations=600,
        total_constraint_evaluations=1200, candidate_checks=45,
        repair_calls=10, repair_improved=7, repair_unchanged=2, repair_failed=1,
        first_feasible_search_evaluation=100, first_feasible_total_constraint_evaluation=240,
        first_feasible_generation=3, final_hard_violations=0, final_soft_penalty=85,
        feasible=True, score=85.0
    )

    out_file = str(tmp_path / "metrics_export_test.xlsx")
    meta = {
        "evaluation_budget": 1000,
        "all_runs_flat": [m.to_dict()]
    }
    export_schedule_to_excel(sched, small_dataset, out_file, metadata=meta)

    wb = openpyxl.load_workbook(out_file)
    assert "RUN_METRICS" in wb.sheetnames

    ws_m = wb["RUN_METRICS"]
    rows = list(ws_m.iter_rows(values_only=True))
    assert len(rows) == 2  # tiêu đề + 1 hàng
    header = rows[0]
    assert "method" in header
    assert "candidate_checks" in header
    assert "repair_improved" in header

    run_row = rows[1]
    assert run_row[header.index("method")] == "Hybrid GA + Repair"
    assert run_row[header.index("search_fitness_evaluations")] == 500
    assert run_row[header.index("repair_improved")] == 7

    ws_cfg = wb["RUN_CONFIG"]
    cfg_rows = dict(ws_cfg.iter_rows(values_only=True))
    assert "search_evaluation_budget" in cfg_rows
