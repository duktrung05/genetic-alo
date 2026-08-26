"""Tests for normalized Excel-driven Soft Constraints S1–S7."""

import io
import copy
import pytest
import openpyxl

from domain import (
    Room,
    StudentGroup,
    CourseSection,
    Timeslot,
    Lecturer,
    Gene,
    Schedule,
    ConstraintDefinition,
)
from dataset import (
    DatasetFactory,
    DatasetValidator,
    ExcelDatasetLoader,
    create_theory_timeslots,
)
from dataset.excel_loader import ExcelValidationError
from constraints import (
    ConstraintEvaluator,
    SoftConstraintChecker,
    SoftConstraintConfig,
    SoftConstraintDefinition,
    SOFT_CONSTRAINT_KEY_BY_ID,
    SOFT_CONSTRAINT_KEYS,
)


# ============================================================
# Helper: Build minimal in-memory Excel workbook with CONSTRAINTS
# ============================================================

def _build_test_xlsx_with_constraints(
    constraints_rows=None,
    sec_preferred_shift="afternoon",
    room_campus="CS1",
    sec_preferred_campus="CS1",
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_campus = wb.create_sheet("CAMPUSES")
    ws_campus.append(["campus_id", "campus_name"])
    for campus_id in sorted({"CS1", room_campus, sec_preferred_campus}):
        ws_campus.append([campus_id, campus_id])

    # TIMESLOTS
    ws_ts = wb.create_sheet("TIMESLOTS")
    ws_ts.append(["timeslot_id", "day_name", "period_no", "shift", "start_time", "end_time"])
    ws_ts.append(["TS-M1", "Thứ 2", 1, "Sáng", "07:00", "07:50"])
    ws_ts.append(["TS-M2", "Thứ 2", 2, "Sáng", "07:50", "08:40"])
    ws_ts.append(["TS-E1", "Thứ 2", 13, "Tối", "18:00", "18:50"])

    # ROOMS
    ws_rm = wb.create_sheet("ROOMS")
    ws_rm.append(["room_id", "campus_id", "building", "room_number", "capacity", "room_type"])
    ws_rm.append(["R1", room_campus, "A", "101", 50, "NORMAL"])

    # LECTURERS
    ws_lec = wb.create_sheet("LECTURERS")
    ws_lec.append(["lecturer_id", "lecturer_name"])
    ws_lec.append(["GV01", "Giảng viên 1"])

    ws_avail = wb.create_sheet("LECTURER_AVAILABILITY")
    ws_avail.append(["lecturer_id", "lecturer_name", "TS-M1", "TS-M2", "TS-E1"])
    ws_avail.append(["GV01", "Giảng viên 1", True, True, True])

    # STUDENT_GROUPS
    ws_grp = wb.create_sheet("STUDENT_GROUPS")
    ws_grp.append(["group_id", "group_name", "size", "home_campus_id"])
    ws_grp.append(["G1", "Group 1", 30, room_campus])

    # COURSES
    ws_crs = wb.create_sheet("COURSES")
    ws_crs.append(["course_id", "course_code", "course_name", "difficulty"])
    ws_crs.append(["C01", "IT001", "Course 1", "MEDIUM"])

    # COURSE_SECTIONS
    ws_sec = wb.create_sheet("COURSE_SECTIONS")
    ws_sec.append([
        "section_id", "class_code", "course_id", "course_code", "course_name", "lecturer_id",
        "student_group_id", "student_count", "required_room_type",
        "duration_periods", "preferred_campus_id", "preferred_shift",
        "meetings_per_week",
    ])
    ws_sec.append([
        "SEC-001", "CLASS-001", "C01", "IT001", "Course 1", "GV01",
        "G1", 30, "NORMAL", 1,
        sec_preferred_campus, sec_preferred_shift, 1,
    ])

    # CONSTRAINTS
    ws_c = wb.create_sheet("CONSTRAINTS")
    ws_c.append(["constraint_id", "constraint_type", "constraint_name", "weight", "enabled"])
    if constraints_rows is None:
        constraints_rows = [
            ("H1", "HARD", "Mỗi lớp học phần xếp 1 lần", 1000000, True),
            ("S1", "SOFT", "Phân bố môn học đều", 10, True),
            ("S2", "SOFT", "Hạn chế tiết cuối ngày", 5, True),
            ("S3", "SOFT", "Ưu tiên ca học mong muốn", 4, True),
            ("S4", "SOFT", "Giảm số ghế trống", 2, True),
            ("S5", "SOFT", "Hạn chế chuyển cơ sở liên tiếp", 8, True),
        ]
    for row in constraints_rows:
        ws_c.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# 14.1. Loader tests cho CONSTRAINTS sheet
# ============================================================

@pytest.mark.unit
def test_loader_reads_constraints_sheet(tmp_path):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx_with_constraints())
    ds = ExcelDatasetLoader.load(str(p))
    assert "constraints" in ds
    c_defs = ds["constraints"]
    assert len(c_defs) == 6
    s_defs = [c for c in c_defs if c.constraint_type == "SOFT"]
    assert len(s_defs) == 5

    s1 = next(c for c in s_defs if c.constraint_id == "S1")
    assert s1.weight == 10
    assert s1.enabled is True

@pytest.mark.unit
@pytest.mark.parametrize("raw_w, expected_w", [
    (10, 10),
    (10.0, 10),
    ("10", 10),
    ("10.0", 10),
    (0, 0),
])
def test_loader_parses_various_weight_formats(tmp_path, raw_w, expected_w):
    rows = [
        ("S1", "SOFT", "Phân bố môn học", raw_w, True),
    ]
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx_with_constraints(constraints_rows=rows))
    ds = ExcelDatasetLoader.load(str(p))
    s1 = ds["constraints"][0]
    assert s1.weight == expected_w

@pytest.mark.unit
@pytest.mark.parametrize("raw_en, expected_en", [
    (True, True),
    (False, False),
    ("true", True),
    ("TRUE", True),
    ("false", False),
    ("FALSE", False),
    (1, True),
    (0, False),
])
def test_loader_parses_enabled_formats(tmp_path, raw_en, expected_en):
    rows = [
        ("S1", "SOFT", "Constraint S1", 10, raw_en),
    ]
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx_with_constraints(constraints_rows=rows))
    ds = ExcelDatasetLoader.load(str(p))
    s1 = ds["constraints"][0]
    assert s1.enabled is expected_en

@pytest.mark.unit
def test_loader_rejects_duplicate_constraint_id(tmp_path):
    rows = [
        ("S1", "SOFT", "Descr 1", 10, True),
        ("S1", "SOFT", "Descr 2", 5, True),
    ]
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx_with_constraints(constraints_rows=rows))
    with pytest.raises(ExcelValidationError, match="Duplicate constraint_id"):
        ExcelDatasetLoader.load(str(p))

@pytest.mark.unit
def test_loader_rejects_missing_header_in_constraints(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONSTRAINTS"
    ws.append(["constraint_id", "weight"])  # missing required columns
    p = tmp_path / "bad_header.xlsx"
    wb.save(p)
    with pytest.raises(ExcelValidationError, match="Required header column is missing"):
        ExcelDatasetLoader.load(str(p))

@pytest.mark.unit
def test_loader_rejects_negative_weight(tmp_path):
    rows = [("S1", "SOFT", "Name", -5, True)]
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx_with_constraints(constraints_rows=rows))
    with pytest.raises(ExcelValidationError, match="cannot be negative"):
        ExcelDatasetLoader.load(str(p))

@pytest.mark.unit
def test_loader_rejects_non_integral_weight(tmp_path):
    rows = [("S1", "SOFT", "Name", 2.5, True)]
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx_with_constraints(constraints_rows=rows))
    with pytest.raises(ExcelValidationError, match="Non-integer weight not allowed"):
        ExcelDatasetLoader.load(str(p))

@pytest.mark.unit
def test_loader_rejects_unsupported_soft_id(tmp_path):
    rows = [("S9", "SOFT", "Unknown constraint", 10, True)]
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx_with_constraints(constraints_rows=rows))
    with pytest.raises(ExcelValidationError, match="Unsupported soft constraint_id='S9'"):
        ExcelDatasetLoader.load(str(p))


# ============================================================
# 14.2. S1 — compact_student_schedule tests
# ============================================================

@pytest.mark.unit
def test_s1_more_active_days_has_maximum_compactness_penalty():
    """One group active on all five available days has normalized S1=1."""
    timeslots = create_theory_timeslots(days=["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6"], max_period=5)
    rooms = [Room(id="R1", name="R1", capacity=50)]
    group = StudentGroup(id="G1", name="G1", student_count=30)
    # 5 sections duration=2 scheduled on 5 different days
    sections = [
        CourseSection(f"SEC{i}", f"C{i}", f"Course {i}", "GV1", "G1", 30, duration_periods=2)
        for i in range(5)
    ]
    # TS ids for period 1 on each day
    # timeslots contains 5 days * 5 periods = 25 timeslots
    ts_by_day = {t.day: t for t in timeslots if t.period == 1}
    genes = [
        Gene(section_id=sections[i].section_id, room_id="R1", timeslot_id=ts_by_day[f"Thứ {i+2}"].id)
        for i in range(5)
    ]
    sched = Schedule(genes=genes)
    ds = {
        "timeslots": timeslots, "rooms": rooms, "lecturers": [],
        "student_groups": [group], "course_sections": sections
    }
    checker = SoftConstraintChecker(
        {s.section_id: s for s in sections},
        {r.id: r for r in rooms},
        {t.id: t for t in timeslots},
    )
    _, details, metrics, _, _ = checker.evaluate_metrics(sched)
    assert details["compact_student_schedule"] == 4
    assert metrics["compact_student_schedule"].normalized == pytest.approx(1.0)

@pytest.mark.unit
def test_s1_compact_two_day_schedule_has_lower_penalty():
    """One group active on two of five days has normalized S1=1/4."""
    days = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6"]
    timeslots = create_theory_timeslots(days=days, max_period=10)
    rooms = [Room(id="R1", name="R1", capacity=50)]
    group = StudentGroup(id="G1", name="G1", student_count=30)
    # 5 sections duration=2: 3 on Mon (periods 1, 3, 5 = 6 periods), 2 on Tue (periods 1, 3 = 4 periods)
    sections = [
        CourseSection(f"SEC{i}", "C1", "Course", "GV1", "G1", 30, duration_periods=2)
        for i in range(5)
    ]
    ts_map = {(t.day, t.period): t.id for t in timeslots}

    genes = [
        Gene(section_id=sections[0].section_id, room_id="R1", timeslot_id=ts_map[("Thứ 2", 1)]),
        Gene(section_id=sections[1].section_id, room_id="R1", timeslot_id=ts_map[("Thứ 2", 3)]),
        Gene(section_id=sections[2].section_id, room_id="R1", timeslot_id=ts_map[("Thứ 2", 5)]),
        Gene(section_id=sections[3].section_id, room_id="R1", timeslot_id=ts_map[("Thứ 3", 1)]),
        Gene(section_id=sections[4].section_id, room_id="R1", timeslot_id=ts_map[("Thứ 3", 3)]),
    ]
    sched = Schedule(genes=genes)
    checker = SoftConstraintChecker(
        {s.section_id: s for s in sections},
        {r.id: r for r in rooms},
        {t.id: t for t in timeslots},
    )
    _, details, metrics, _, _ = checker.evaluate_metrics(sched)
    assert details["compact_student_schedule"] == 1
    assert metrics["compact_student_schedule"].normalized == pytest.approx(0.25)


# ============================================================
# 14.3. S2 — late_day_periods tests
# ============================================================

@pytest.mark.unit
def test_s2_morning_afternoon_no_penalty():
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=12)
    rooms = [Room(id="R1", name="R1", capacity=50)]
    sec = CourseSection("SEC1", "C1", "Course", "GV1", "G1", 30, duration_periods=3)
    # Morning slot period 1
    ts_morning = next(t for t in timeslots if t.session == "morning")
    sched = Schedule(genes=[Gene("SEC1", "R1", ts_morning.id)])
    checker = SoftConstraintChecker({"SEC1": sec}, {"R1": rooms[0]}, {t.id: t for t in timeslots})
    _, details = checker.evaluate(sched)
    assert details["late_day_periods"] == 0

@pytest.mark.unit
def test_s2_evening_periods_penalty():
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=16)
    rooms = [Room(id="R1", name="R1", capacity=50)]
    sec = CourseSection("SEC1", "C1", "Course", "GV1", "G1", 30, duration_periods=3)
    # Evening slot (period >= 13)
    ts_evening = next(t for t in timeslots if t.session == "evening" and t.period == 13)
    sched = Schedule(genes=[Gene("SEC1", "R1", ts_evening.id)])
    checker = SoftConstraintChecker({"SEC1": sec}, {"R1": rooms[0]}, {t.id: t for t in timeslots})
    _, details = checker.evaluate(sched)
    assert details["late_day_periods"] == 3


# ============================================================
# 14.4. S3 — preferred_shift_mismatch tests
# ============================================================

@pytest.mark.unit
def test_s3_matching_preferred_shift():
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=12)
    rooms = [Room(id="R1", name="R1", capacity=50)]
    sec = CourseSection("SEC1", "C1", "Course", "GV1", "G1", 30, preferred_shift="morning")
    ts_morning = next(t for t in timeslots if t.session == "morning")
    sched = Schedule(genes=[Gene("SEC1", "R1", ts_morning.id)])
    checker = SoftConstraintChecker({"SEC1": sec}, {"R1": rooms[0]}, {t.id: t for t in timeslots})
    _, details = checker.evaluate(sched)
    assert details["preferred_shift_mismatch"] == 0

@pytest.mark.unit
def test_s3_mismatched_preferred_shift():
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=12)
    rooms = [Room(id="R1", name="R1", capacity=50)]
    sec = CourseSection("SEC1", "C1", "Course", "GV1", "G1", 30, preferred_shift="afternoon")
    ts_morning = next(t for t in timeslots if t.session == "morning")
    sched = Schedule(genes=[Gene("SEC1", "R1", ts_morning.id)])
    checker = SoftConstraintChecker({"SEC1": sec}, {"R1": rooms[0]}, {t.id: t for t in timeslots})
    _, details = checker.evaluate(sched)
    assert details["preferred_shift_mismatch"] == 1

@pytest.mark.unit
def test_s3_none_preferred_shift():
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=12)
    rooms = [Room(id="R1", name="R1", capacity=50)]
    sec = CourseSection("SEC1", "C1", "Course", "GV1", "G1", 30, preferred_shift=None)
    ts_morning = next(t for t in timeslots if t.session == "morning")
    sched = Schedule(genes=[Gene("SEC1", "R1", ts_morning.id)])
    checker = SoftConstraintChecker({"SEC1": sec}, {"R1": rooms[0]}, {t.id: t for t in timeslots})
    _, details = checker.evaluate(sched)
    assert details["preferred_shift_mismatch"] == 0


# ============================================================
# 14.5. S4 — room_seat_waste tests
# ============================================================

@pytest.mark.unit
@pytest.mark.parametrize("capacity, student_count, expected_waste", [
    (100, 100, 0.0),
    (100, 50, 0.5),
    (200, 50, 0.75),
    (30, 40, 0.0),  # insufficient capacity is hard-only; S4 ignores it
])
def test_s4_room_seat_waste(capacity, student_count, expected_waste):
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=5)
    room = Room(id="R1", name="R1", capacity=capacity)
    sec = CourseSection("SEC1", "C1", "Course", "GV1", "G1", student_count)
    sched = Schedule(genes=[Gene("SEC1", "R1", timeslots[0].id)])
    checker = SoftConstraintChecker({"SEC1": sec}, {"R1": room}, {t.id: t for t in timeslots})
    _, details = checker.evaluate(sched)
    assert details["room_seat_waste"] == pytest.approx(expected_waste)


# ============================================================
# 14.6. S5 — consecutive_cross_campus tests
# ============================================================

@pytest.mark.unit
def test_s5_consecutive_cross_campus_penalty():
    """GV dạy tiết 1-3 tại CS1 và tiết 4-6 tại CS2 cùng ngày → 1 raw violation."""
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=12)
    rooms = {
        "R_CS1": Room(id="R_CS1", name="R_CS1", capacity=50, campus_id="CS1"),
        "R_CS2": Room(id="R_CS2", name="R_CS2", capacity=50, campus_id="CS2"),
    }
    sec1 = CourseSection("SEC1", "C1", "Course 1", "GV1", "G1", 30, duration_periods=3)
    sec2 = CourseSection("SEC2", "C2", "Course 2", "GV1", "G2", 30, duration_periods=3)

    ts1 = next(t for t in timeslots if t.period == 1)
    ts4 = next(t for t in timeslots if t.period == 4)  # consecutive block start

    sched = Schedule(genes=[
        Gene("SEC1", "R_CS1", ts1.id),
        Gene("SEC2", "R_CS2", ts4.id),
    ])
    checker = SoftConstraintChecker(
        {"SEC1": sec1, "SEC2": sec2},
        rooms,
        {t.id: t for t in timeslots},
    )
    _, details = checker.evaluate(sched)
    assert details["consecutive_cross_campus"] == 1

@pytest.mark.unit
def test_s5_consecutive_same_campus_no_penalty():
    """GV dạy tiết 1-3 tại CS1 và tiết 4-6 tại CS1 cùng ngày → 0 raw violation."""
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=12)
    rooms = {
        "R_CS1_A": Room(id="R_CS1_A", name="R_CS1_A", capacity=50, campus_id="CS1"),
        "R_CS1_B": Room(id="R_CS1_B", name="R_CS1_B", capacity=50, campus_id="CS1"),
    }
    sec1 = CourseSection("SEC1", "C1", "Course 1", "GV1", "G1", 30, duration_periods=3)
    sec2 = CourseSection("SEC2", "C2", "Course 2", "GV1", "G2", 30, duration_periods=3)

    ts1 = next(t for t in timeslots if t.period == 1)
    ts4 = next(t for t in timeslots if t.period == 4)

    sched = Schedule(genes=[
        Gene("SEC1", "R_CS1_A", ts1.id),
        Gene("SEC2", "R_CS1_B", ts4.id),
    ])
    checker = SoftConstraintChecker(
        {"SEC1": sec1, "SEC2": sec2},
        rooms,
        {t.id: t for t in timeslots},
    )
    _, details = checker.evaluate(sched)
    assert details["consecutive_cross_campus"] == 0

@pytest.mark.unit
def test_s5_gap_between_cross_campus_no_penalty():
    """GV dạy tiết 1-3 tại CS1 và tiết 5-6 tại CS2 (tiết 4 nghỉ) → 0 raw violation."""
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=12)
    rooms = {
        "R_CS1": Room(id="R_CS1", name="R_CS1", capacity=50, campus_id="CS1"),
        "R_CS2": Room(id="R_CS2", name="R_CS2", capacity=50, campus_id="CS2"),
    }
    sec1 = CourseSection("SEC1", "C1", "Course 1", "GV1", "G1", 30, duration_periods=3)
    sec2 = CourseSection("SEC2", "C2", "Course 2", "GV1", "G2", 30, duration_periods=2)

    ts1 = next(t for t in timeslots if t.period == 1)
    ts5 = next(t for t in timeslots if t.period == 5)  # gap of period 4

    sched = Schedule(genes=[
        Gene("SEC1", "R_CS1", ts1.id),
        Gene("SEC2", "R_CS2", ts5.id),
    ])
    checker = SoftConstraintChecker(
        {"SEC1": sec1, "SEC2": sec2},
        rooms,
        {t.id: t for t in timeslots},
    )
    _, details = checker.evaluate(sched)
    assert details["consecutive_cross_campus"] == 0


# ============================================================
# 14.7. Weight test (normalized metrics × stakeholder weights)
# ============================================================

@pytest.mark.unit
def test_weight_calculation_formula():
    """All supplied metrics are normalized and weighted without raw-unit bias."""
    config = SoftConstraintConfig.default()
    details = {
        "compact_student_schedule": 0.2,
        "late_day_periods": 0.3,
        "preferred_shift_mismatch": 0.4,
        "room_seat_waste": 0.5,
        "consecutive_cross_campus": 0.6,
        "preferred_campus_mismatch": 0.7,
        "student_home_campus_mismatch": 0.8,
    }
    checker = SoftConstraintChecker({}, {}, {}, config=config)
    tot = checker.calculate_weighted_penalty(details)
    expected = 0.2 * 5 + 0.3 * 4 + 0.4 * 4 + 0.5 * 4 + 0.6 * 4 + 0.7 * 3 + 0.8 * 4
    assert tot == pytest.approx(expected)


# ============================================================
# 14.8. Enabled test
# ============================================================

@pytest.mark.unit
def test_disabled_constraint_contributes_zero():
    """Tắt S4 (enabled=False): dù có room_seat_waste, penalty=0."""
    c_defs = [
        ConstraintDefinition("S1", "SOFT", "Weekly", 10, True),
        ConstraintDefinition("S2", "SOFT", "Late", 5, True),
        ConstraintDefinition("S3", "SOFT", "Shift", 4, True),
        ConstraintDefinition("S4", "SOFT", "Waste", 2, False),  # DISABLED
        ConstraintDefinition("S5", "SOFT", "Campus", 8, True),
    ]
    config = SoftConstraintConfig.from_constraint_definitions(c_defs)
    assert not config.is_enabled("room_seat_waste")

    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=5)
    room = Room(id="R1", name="R1", capacity=100)  # 60 unused seats
    sec = CourseSection("SEC1", "C1", "Course", "GV1", "G1", 40)
    sched = Schedule(genes=[Gene("SEC1", "R1", timeslots[0].id)])

    checker = SoftConstraintChecker({"SEC1": sec}, {"R1": room}, {t.id: t for t in timeslots}, config=config)
    raw, details, items = checker.evaluate_detailed(sched)
    # S4 is in details but item weighted_penalty is 0
    s4_items = [i for i in items if i.get("constraint_key") == "room_seat_waste"]
    assert len(s4_items) == 0  # disabled constraint creates no violation items
    _, _, metrics, _, _ = checker.evaluate_metrics(sched)
    assert metrics["room_seat_waste"].raw == 0
    assert metrics["room_seat_waste"].weighted == 0


# ============================================================
# 14.9. Unified evaluator test
# ============================================================

@pytest.mark.unit
def test_unified_evaluator_consistency(small_dataset):
    evaluator = ConstraintEvaluator(small_dataset)
    # Build a valid dummy schedule
    sections = small_dataset["course_sections"]
    rooms = small_dataset["rooms"]
    genes = [Gene(s.section_id, rooms[0].id, 0) for s in sections]
    sched = Schedule(genes=genes)

    res = evaluator.evaluate_unified(sched)
    calculated_soft = sum(item.weighted_penalty for item in res.soft_breakdown)
    assert res.soft_penalty == calculated_soft
    assert len(res.soft_breakdown) == 7

    # Check key names in breakdown
    keys_in_breakdown = [item.constraint_key for item in res.soft_breakdown]
    assert keys_in_breakdown == SOFT_CONSTRAINT_KEYS


# ============================================================
# 14.10. Exporter test — RUN_CONFIG & VIOLATIONS sheets
# ============================================================

@pytest.mark.unit
def test_exporter_uses_s1_s5_not_legacy(tmp_path):
    from evaluation import export_schedule_to_excel
    from dataset import find_feasible_schedule

    c_defs = [
        ConstraintDefinition("S1", "SOFT", "Weekly", 10, True),
        ConstraintDefinition("S2", "SOFT", "Late", 5, True),
        ConstraintDefinition("S3", "SOFT", "Shift", 4, True),
        ConstraintDefinition("S4", "SOFT", "Waste", 2, True),
        ConstraintDefinition("S5", "SOFT", "Campus", 8, True),
    ]
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=16)
    rooms = [Room(id="R1", name="R1", capacity=50, campus_id="CS1")]
    lecturers = [Lecturer(id="GV1", name="GV1")]
    groups = [StudentGroup(id="G1", name="G1", student_count=30, home_campus_id="CS1")]
    sections = [CourseSection("SEC1", "C1", "Course", "GV1", "G1", 30)]

    ds = {
        "timeslots": timeslots, "rooms": rooms, "lecturers": lecturers,
        "student_groups": groups, "course_sections": sections,
        "constraints": c_defs,
    }
    sched = find_feasible_schedule(ds)
    assert sched is not None

    out = str(tmp_path / "s1_s5_test.xlsx")
    export_schedule_to_excel(sched, ds, out)

    wb = openpyxl.load_workbook(out)

    # Check RUN_CONFIG
    ws_cfg = wb["RUN_CONFIG"]
    cfg_params = {row[0]: row[1] for row in ws_cfg.iter_rows(min_row=2, values_only=True)}
    assert "S1_compact_student_schedule_weight" in cfg_params
    assert "S2_late_day_periods_weight" in cfg_params
    assert "S3_preferred_shift_mismatch_weight" in cfg_params
    assert "S4_room_seat_waste_weight" in cfg_params
    assert "S5_consecutive_cross_campus_weight" in cfg_params
    assert "S6_preferred_campus_mismatch_weight" in cfg_params
    assert "S7_student_home_campus_mismatch_weight" in cfg_params

    # Legacy weights MUST NOT exist
    assert "student_gaps_weight" not in cfg_params
    assert "consecutive_teaching_weight" not in cfg_params
    assert "difficult_afternoon_weight" not in cfg_params
    assert "daily_imbalance_weight" not in cfg_params


# ============================================================
# 14.11. Pipeline integration test
# ============================================================

@pytest.mark.unit
def test_ga_engine_runs_with_excel_constraints(tmp_path):
    """GeneticAlgorithmEngine runs clean with a legacy S1-S5 workbook config."""
    from ga import GeneticAlgorithmEngine

    c_defs = [
        ConstraintDefinition("H1", "HARD", "Hard 1", 1000000, True),
        ConstraintDefinition("S1", "SOFT", "Weekly", 10, True),
        ConstraintDefinition("S2", "SOFT", "Late", 5, True),
        ConstraintDefinition("S3", "SOFT", "Shift", 4, True),
        ConstraintDefinition("S4", "SOFT", "Waste", 2, True),
        ConstraintDefinition("S5", "SOFT", "Campus", 8, True),
    ]
    ds = DatasetFactory.create_small_dataset()
    ds["constraints"] = c_defs

    engine = GeneticAlgorithmEngine(ds, pop_size=10)
    res = engine.run(generations=5, use_repair=True)
    assert "best_schedule" in res
    assert "soft_penalty" in res
    assert res["soft_penalty"] >= 0
