"""Tests for Step 6 — Workflow Architecture, Production Entry Point, and Dynamic CLI Benchmarking."""

import pytest
import openpyxl
import json
import os
import csv
from pathlib import Path
from unittest.mock import MagicMock, patch

from evaluation import parse_methods, SUPPORTED_METHODS, METHOD_DISPLAY_NAMES, METHOD_RUNNERS
from evaluation.run_metrics import RunMetrics
from domain import Schedule, Gene
from ga import GeneticAlgorithmEngine
from evaluation.baselines import RandomSearchScheduler, GreedyScheduler


# ============================================================
# 1. Kiểm thử đơn vị việc đọc phương pháp (Yêu cầu 19.1)
# ============================================================

@pytest.mark.unit
def test_parse_methods_single():
    assert parse_methods("ga_repair_sls") == ["ga_repair_sls"]
    assert parse_methods("hybrid") == ["ga_repair"]


@pytest.mark.unit
def test_parse_methods_multiple():
    assert parse_methods("ga_repair_sls,ga_repair,ga") == [
        "ga_repair_sls", "ga_repair", "ga"
    ]
    assert parse_methods("repair_only, greedy, random") == [
        "repair_only", "greedy", "random"
    ]


@pytest.mark.unit
def test_parse_methods_case_insensitive_and_whitespace():
    assert parse_methods("  GA_Repair_SLS ,  GA  ") == ["ga_repair_sls", "ga"]


@pytest.mark.unit
def test_parse_methods_deduplication_preserves_first_occurrence():
    assert parse_methods("hybrid, greedy, ga_repair, ga, greedy") == [
        "ga_repair", "greedy", "ga"
    ]


@pytest.mark.unit
def test_parse_methods_unsupported_raises():
    with pytest.raises(ValueError, match="Unsupported method 'gready'"):
        parse_methods("ga_repair,gready")


@pytest.mark.unit
def test_parse_methods_empty_raises():
    with pytest.raises(ValueError, match="Methods list cannot be empty"):
        parse_methods("")
    with pytest.raises(ValueError, match="Methods list cannot be empty"):
        parse_methods("   ")


# ============================================================
# 2. Kiểm thử tính cô lập của trình chạy đã chọn (Yêu cầu 19.2)
# ============================================================

@pytest.mark.unit
def test_selected_runner_isolation(small_dataset):
    ga_config = {"pop_size": 10, "generations": 5, "crossover_rate": 0.8, "mutation_rate": 0.2}

    # Chỉ thực thi trình chạy GA + Repair và GA thuần.
    selected = parse_methods("ga_repair,ga")

    executed_methods = []
    for m in selected:
        res = METHOD_RUNNERS[m](small_dataset, ga_config, budget=10, seed=0)
        executed_methods.append(m)
        assert res["run_metrics"].search_fitness_evaluations == 10

    assert executed_methods == ["ga_repair", "ga"]
    assert "repair_only" not in executed_methods
    assert "ga_repair_sls" not in executed_methods
    assert "greedy" not in executed_methods
    assert "random" not in executed_methods


# ============================================================
# 3. Kiểm thử benchmark chỉ dùng Hybrid (Yêu cầu 19.3)
# ============================================================

@pytest.mark.integration
def test_ga_repair_sls_matches_production_flow(small_dataset):
    ga_config = {"pop_size": 10, "generations": 5, "crossover_rate": 0.8, "mutation_rate": 0.2}
    selected = parse_methods("ga_repair_sls")

    runs = [METHOD_RUNNERS["ga_repair_sls"](small_dataset, ga_config, budget=20, seed=0)]

    assert selected == ["ga_repair_sls"]
    for r in runs:
        assert r["run_metrics"].method == "GA + Repair + SLS (Production)"
        assert r["run_metrics"].search_fitness_evaluations == 20
        assert r["use_repair"] is True
        assert r["use_soft_local_search"] is True


# ============================================================
# 4. Kiểm thử benchmark Hybrid + GA (Yêu cầu 19.4)
# ============================================================

@pytest.mark.integration
def test_ga_repair_and_plain_ga_benchmark_execution(small_dataset):
    ga_config = {"pop_size": 10, "generations": 5, "crossover_rate": 0.8, "mutation_rate": 0.2}
    selected = parse_methods("ga_repair,ga")

    repair_runs = [METHOD_RUNNERS["ga_repair"](small_dataset, ga_config, budget=10, seed=0)]
    ga_runs = [METHOD_RUNNERS["ga"](small_dataset, ga_config, budget=10, seed=0)]

    assert selected == ["ga_repair", "ga"]
    assert repair_runs[0]["run_metrics"].method == "GA + Repair"
    assert ga_runs[0]["run_metrics"].method == "GA without Repair"


@pytest.mark.unit
def test_repair_only_baseline_uses_budget_without_ga_operators(small_dataset):
    ga_config = {"hard_weight": 1000, "soft_weight": 1}

    result = METHOD_RUNNERS["repair_only"](
        small_dataset, ga_config, budget=5, seed=0
    )

    metrics = result["run_metrics"]
    assert metrics.method == "Repair-only Random Restart"
    assert metrics.search_fitness_evaluations == 5
    assert metrics.repair_calls == 5
    assert result["uses_ga_operators"] is False


@pytest.mark.unit
def test_best_timetable_selection_rechecks_feasibility(small_dataset):
    from dataset import find_feasible_schedule
    from main_benchmark import select_best_feasible_run

    feasible_schedule = find_feasible_schedule(small_dataset)
    assert feasible_schedule is not None

    room_id = small_dataset["rooms"][0].id
    timeslot_id = small_dataset["timeslots"][0].id
    infeasible_schedule = Schedule(genes=[
        Gene(section.section_id, room_id, timeslot_id)
        for section in small_dataset["course_sections"]
    ])

    selected = select_best_feasible_run(
        [
            {
                "method": "GA + Repair + SLS (Production)",
                "best_schedule": infeasible_schedule,
                "hard_violations": 0,  # Chỉ số cũ được cố ý giữ lại.
                "soft_penalty": 0,
            },
            {
                "method": "GA without Repair",
                "best_schedule": feasible_schedule,
                "hard_violations": 99,  # Phải được thay bằng kết quả đánh giá lại.
                "soft_penalty": 9999,
            },
        ],
        small_dataset,
    )

    assert selected is not None
    assert selected["method"] == "GA without Repair"
    assert selected["hard_violations"] == 0


@pytest.mark.unit
def test_best_timetable_selection_returns_none_when_all_runs_are_infeasible(small_dataset):
    from main_benchmark import select_best_feasible_run

    room_id = small_dataset["rooms"][0].id
    timeslot_id = small_dataset["timeslots"][0].id
    infeasible_schedule = Schedule(genes=[
        Gene(section.section_id, room_id, timeslot_id)
        for section in small_dataset["course_sections"]
    ])

    selected = select_best_feasible_run(
        [{
            "method": "GA + Repair + SLS (Production)",
            "best_schedule": infeasible_schedule,
            "hard_violations": 0,
            "soft_penalty": 0,
        }],
        small_dataset,
    )

    assert selected is None


@pytest.mark.integration
def test_benchmark_completes_without_export_when_no_run_is_feasible(
    tmp_path, small_dataset, capsys
):
    from main_benchmark import main as benchmark_main

    infeasible_result = RandomSearchScheduler(small_dataset, seed=0).run(
        evaluation_budget=1,
        seed=0,
    )
    assert infeasible_result["hard_violations"] > 0

    def return_infeasible_result(dataset, ga_config, budget, seed):
        return infeasible_result

    argv = [
        "main_benchmark.py",
        "--mode", "quick",
        "--methods", "random",
        "--seeds", "0",
        "--search-evaluation-budget", "1",
        "--data-source", "mock",
        "--preset", "small",
        "--output-dir", str(tmp_path),
    ]

    with patch("sys.argv", argv), \
         patch("main_benchmark.DatasetFactory.create_small_dataset", return_value=small_dataset), \
         patch.dict("main_benchmark.METHOD_RUNNERS", {"random": return_infeasible_result}), \
         patch("main_benchmark.ConvergenceVisualizer.plot_convergence"), \
         patch("main_benchmark.export_schedule_to_excel") as export_mock:
        benchmark_main()

    export_mock.assert_not_called()
    assert (tmp_path / "raw_runs.json").exists()
    assert (tmp_path / "summary.json").exists()
    assert not (tmp_path / "best_timetable.xlsx").exists()
    assert "CHƯA CÓ LỊCH HARD-FEASIBLE" in capsys.readouterr().out


# ============================================================
# 5. Kiểm thử Greedy tùy chọn (Yêu cầu 19.5)
# ============================================================

@pytest.mark.unit
def test_greedy_optional_single_run(small_dataset):
    ga_config = {}
    greedy_res = METHOD_RUNNERS["greedy"](small_dataset, ga_config, budget=1000, seed=0)

    metrics = greedy_res["run_metrics"]
    assert metrics.method == "Greedy Search"
    assert metrics.search_fitness_evaluations == 1


# ============================================================
# 6. Kiểm thử tìm kiếm ngẫu nhiên tùy chọn (Yêu cầu 19.6)
# ============================================================

@pytest.mark.unit
def test_random_search_optional_execution(small_dataset):
    ga_config = {}
    random_res = METHOD_RUNNERS["random"](small_dataset, ga_config, budget=15, seed=0)

    metrics = random_res["run_metrics"]
    assert metrics.method == "Random Search"
    assert metrics.search_fitness_evaluations == 15


# ============================================================
# 7. Kiểm thử hàm main production (Yêu cầu 19.7)
# ============================================================

@pytest.mark.unit
def test_production_method_identity_matches_sls_flag():
    from main import get_production_method_identity

    assert get_production_method_identity(False) == (
        "ga_repair",
        "GA + Repair",
    )
    assert get_production_method_identity(True) == (
        "ga_repair_sls",
        "GA + Repair + SLS (Production)",
    )

@pytest.mark.integration
def test_production_main_execution(tmp_path, small_dataset):
    from dataset import ExcelDatasetLoader

    # Giả lập load_and_validate để trả về small_dataset
    out_file = tmp_path / "test_production_timetable.xlsx"

    with patch.object(ExcelDatasetLoader, "load_and_validate", return_value=small_dataset):
        with patch("sys.argv", ["main.py", "--input", "dummy.xlsx", "--output", str(out_file), "--seed", "42", "--search-evaluation-budget", "20", "--population-size", "10"]):
            from main import main as main_prod
            main_prod()


    assert out_file.exists()
    wb = openpyxl.load_workbook(out_file)
    assert "SUMMARY" in wb.sheetnames
    assert "RUN_CONFIG" in wb.sheetnames


    # Kiểm tra siêu dữ liệu RUN_CONFIG
    ws_cfg = wb["RUN_CONFIG"]
    rows = dict(list(ws_cfg.iter_rows(values_only=True))[1:])
    assert rows.get("primary_method") == "ga_repair"
    assert rows.get("selected_methods") == "ga_repair"

    ws_metrics = wb["RUN_METRICS"]
    metric_rows = list(ws_metrics.iter_rows(values_only=True))
    metric_header = metric_rows[0]
    metric_data = metric_rows[1]
    assert metric_data[metric_header.index("method")] == "GA + Repair"


# ============================================================
# 8. Xác minh đầu ra động (Yêu cầu 19.8)
# ============================================================

@pytest.mark.unit
def test_dynamic_export_verification(tmp_path, small_dataset):
    from evaluation import export_schedule_to_excel
    from dataset import find_feasible_schedule

    sched = find_feasible_schedule(small_dataset)
    m_hybrid = RunMetrics(
        method="Hybrid GA + Repair", seed=0, runtime_seconds=1.0,
        time_to_first_feasible_seconds=0.5, search_fitness_evaluations=10,
        hard_constraint_evaluations=10, soft_constraint_evaluations=10,
        total_constraint_evaluations=20, candidate_checks=50,
        repair_calls=2, repair_improved=2, repair_unchanged=0, repair_failed=0,
        first_feasible_search_evaluation=1, first_feasible_total_constraint_evaluation=2,
        first_feasible_generation=0, final_hard_violations=0, final_soft_penalty=10,
        feasible=True, score=10.0
    )

    out_path = tmp_path / "dynamic_export.xlsx"
    meta = {
        "primary_method": "hybrid",
        "selected_methods": "hybrid,ga",
        "all_runs_flat": [m_hybrid.to_dict()],
        "summary_list": [{"method": "Hybrid GA + Repair", "run_count": 1, "feasible_rate": 1.0}],
    }

    export_schedule_to_excel(sched, small_dataset, out_path, metadata=meta)

    wb = openpyxl.load_workbook(out_path)

    # Kiểm tra RUN_CONFIG
    ws_cfg = wb["RUN_CONFIG"]
    cfg_dict = dict(list(ws_cfg.iter_rows(values_only=True))[1:])
    assert cfg_dict.get("primary_method") == "hybrid"
    assert cfg_dict.get("selected_methods") == "hybrid,ga"

    # Kiểm tra BENCHMARK_SUMMARY chỉ chứa 1 hàng
    ws_sum = wb["BENCHMARK_SUMMARY"]
    rows = list(ws_sum.iter_rows(values_only=True))
    assert len(rows) == 2  # tiêu đề + 1 hàng dữ liệu
    assert rows[1][0] == "Hybrid GA + Repair"
