import pytest
import os
import openpyxl
from unittest.mock import patch, MagicMock
from dataset import ExcelDatasetLoader, DatasetValidator, DatasetFactory, find_feasible_schedule
from dataset.excel_loader import ExcelValidationError
from domain import Gene, Schedule
from constraints import ConstraintEvaluator
from ga import GeneticAlgorithmEngine

EXCEL_PATH = "data/01_data_timetable.xlsx"
_real_load_workbook = openpyxl.load_workbook

@pytest.mark.unit
def test_excel_loader_file_not_found():
    with pytest.raises(FileNotFoundError):
        ExcelDatasetLoader.load("non_existent_file.xlsx")

@pytest.mark.unit
def test_excel_loader_parses_entities():
    assert os.path.exists(EXCEL_PATH), f"Excel file {EXCEL_PATH} must exist for tests."
    ds = ExcelDatasetLoader.load(EXCEL_PATH)

    assert "timeslots" in ds
    assert "campuses" in ds
    assert "rooms" in ds
    assert "lecturers" in ds
    assert "student_groups" in ds
    assert "courses" in ds
    assert "course_sections" in ds

    assert len(ds["timeslots"]) == 96
    assert len(ds["campuses"]) == 2
    assert len(ds["rooms"]) == 11
    assert len(ds["lecturers"]) == 15
    assert len(ds["student_groups"]) == 12
    assert len(ds["courses"]) == 20
    assert len(ds["course_sections"]) == 62

@pytest.mark.unit
def test_excel_loader_validate_report():
    ds = ExcelDatasetLoader.load_and_validate(EXCEL_PATH)
    report = DatasetValidator.validate_report(ds)

    assert report["valid"] is True
    assert len(report["errors"]) == 0
    assert report["statistics"]["sections"] == 62
    assert report["statistics"]["total_periods"] == 96


@pytest.mark.unit
def test_json_snapshot_roundtrip_no_data_loss(tmp_path):
    ds_orig = ExcelDatasetLoader.load_and_validate(EXCEL_PATH)
    json_path = str(tmp_path / "normalized_test.json")

    # Export to JSON snapshot
    ExcelDatasetLoader.export_normalized_json(ds_orig, json_path)
    assert os.path.exists(json_path)

    # Load back from JSON snapshot
    ds_reconstructed = ExcelDatasetLoader.load_normalized_json(json_path)
    report = DatasetValidator.validate_report(ds_reconstructed)
    assert report["valid"] is True

    # Every scheduling-relevant entity and field must survive the round trip.
    entity_keys = (
        "campuses",
        "timeslots",
        "rooms",
        "lecturers",
        "student_groups",
        "courses",
        "course_sections",
        "constraints",
    )
    assert set(ds_reconstructed) == set(ds_orig) == set(entity_keys)
    for key in entity_keys:
        assert ds_reconstructed[key] == ds_orig[key], f"Round-trip data loss in '{key}'"

    # The same chromosome must represent and score the exact same problem after
    # loading the snapshot. It does not need to be feasible for this invariant.
    room_id = ds_orig["rooms"][0].id
    timeslot_id = ds_orig["timeslots"][0].id
    schedule = Schedule(
        genes=[
            Gene(section_id=section.section_id, room_id=room_id, timeslot_id=timeslot_id)
            for section in ds_orig["course_sections"]
        ]
    )
    original_result = ConstraintEvaluator(ds_orig).evaluate_unified(schedule)
    reconstructed_result = ConstraintEvaluator(ds_reconstructed).evaluate_unified(schedule)

    assert reconstructed_result.hard_violations == original_result.hard_violations
    assert reconstructed_result.hard_details == original_result.hard_details
    assert reconstructed_result.soft_penalty == original_result.soft_penalty
    assert reconstructed_result.soft_breakdown == original_result.soft_breakdown

@pytest.mark.unit
def test_single_load_workbook_call():
    with patch("openpyxl.load_workbook", side_effect=_real_load_workbook) as mock_load:
        ExcelDatasetLoader.load_and_validate(EXCEL_PATH)
        assert mock_load.call_count == 1

@pytest.mark.unit
def test_output_sheets_ignored():
    ds = ExcelDatasetLoader.load(EXCEL_PATH)
    allowed_keys = {"campuses", "timeslots", "rooms", "lecturers", "student_groups", "courses", "course_sections", "constraints"}
    assert set(ds.keys()) == allowed_keys


@pytest.mark.unit
def test_detailed_error_messages_missing_sheet():
    mock_wb = MagicMock()
    mock_wb.sheetnames = ["CAMPUSES", "ROOMS"]  # Missing TIMESLOTS
    campus_ws = MagicMock()
    campus_ws.iter_rows.return_value = [
        ("campus_id", "campus_name"),
        ("CS1", "Campus 1"),
    ]
    mock_wb.__getitem__.return_value = campus_ws
    with patch("os.path.exists", return_value=True):
        with patch("openpyxl.load_workbook", return_value=mock_wb):
            with pytest.raises(ExcelValidationError) as exc_info:
                ExcelDatasetLoader.load("dummy.xlsx")
            assert "Sheet 'TIMESLOTS'" in str(exc_info.value)
            assert "missing from workbook" in str(exc_info.value)

@pytest.mark.unit
def test_detailed_error_messages_missing_column():
    mock_wb = MagicMock()
    mock_wb.sheetnames = ["CAMPUSES", "TIMESLOTS"]
    campus_ws = MagicMock()
    campus_ws.iter_rows.return_value = [
        ("campus_id", "campus_name"),
        ("CS1", "Campus 1"),
    ]
    timeslot_ws = MagicMock()
    timeslot_ws.iter_rows.return_value = [
        ("timeslot_id", "day_name")
    ]  # Missing period_no, shift
    mock_wb.__getitem__.side_effect = lambda name: {
        "CAMPUSES": campus_ws,
        "TIMESLOTS": timeslot_ws,
    }[name]
    with patch("os.path.exists", return_value=True):
        with patch("openpyxl.load_workbook", return_value=mock_wb):
            with pytest.raises(ExcelValidationError) as exc_info:
                ExcelDatasetLoader.load("dummy.xlsx")
            assert "Sheet 'TIMESLOTS'" in str(exc_info.value)
            assert "Required header column is missing" in str(exc_info.value)

@pytest.mark.integration
def test_excel_dataset_feasibility_checker():
    ds = ExcelDatasetLoader.load_and_validate(EXCEL_PATH)
    feasible_sched = find_feasible_schedule(ds)
    assert feasible_sched is not None, "Excel dataset must possess at least one feasible reference schedule."

@pytest.mark.integration
def test_ga_engine_runs_on_excel_dataset():
    import random
    random.seed(42)
    ds = ExcelDatasetLoader.load_and_validate(EXCEL_PATH)
    ga = GeneticAlgorithmEngine(ds, pop_size=30)
    res = ga.run(generations=10, use_repair=True)

    assert res["best_schedule"] is not None
    assert res["hard_violations"] == 0
    assert res["use_repair"] is True
    assert "repair_stats" in res
