import os
import csv
import pytest
from domain import Schedule, Gene, CourseSection, Room, Lecturer, StudentGroup
from dataset import create_theory_timeslots
from evaluation import export_schedule_to_csv, export_metadata_to_json, ConvergenceVisualizer

@pytest.fixture
def exporter_dataset():
    rooms = [
        Room(id="P101", name="Phòng 101", capacity=100, room_type="NORMAL"),
        Room(id="LAB01", name="Phòng LAB 01", capacity=100, room_type="LAB"),
    ]
    timeslots = create_theory_timeslots(days=["Thứ 2"], max_period=6)
    lecturers = [Lecturer(id="GV01", name="Giảng viên 1"), Lecturer(id="GV02", name="Giảng viên 2")]
    groups = [StudentGroup(id="SV_CNTT1", name="CNTT 1", student_count=60)]
    sections = [
        CourseSection("LHP01", "CS101", "Nhập môn Lập trình", "GV01", "SV_CNTT1", 60, duration_periods=2, required_room_type="NORMAL"),
        CourseSection("LHP02", "CS102", "Thực hành Lập trình", "GV02", "SV_CNTT1", 60, duration_periods=3, required_room_type="LAB"),
    ]
    return {
        "rooms": rooms,
        "timeslots": timeslots,
        "lecturers": lecturers,
        "student_groups": groups,
        "course_sections": sections,
    }

@pytest.mark.integration
def test_csv_exporter_multi_period_fields(exporter_dataset, tmp_path):
    output_path = tmp_path / "test_timetable.csv"
    sched = Schedule(genes=[
        Gene(section_id="LHP01", timeslot_id=0, room_id="P101"),  # Thứ Hai, tiết 1..2 (07:00..08:40)
        Gene(section_id="LHP02", timeslot_id=2, room_id="LAB01"), # Thứ Hai, tiết 3..5 (08:45..11:25)
    ])

    path = export_schedule_to_csv(sched, exporter_dataset, output_path)
    assert os.path.exists(path)

    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    assert len(rows) == 2

    # Kiểm tra các cột
    expected_headers = [
        "activity_id", "section_id", "meeting_index", "meeting_count", "meeting",
        "class_code", "course_id", "course_code",
        "lecturer_id", "student_group_id", "room_id",
        "day", "start_period", "end_period", "start_time", "end_time",
        "duration_periods", "session", "room_type"
    ]
    assert reader.fieldnames == expected_headers

    # Hàng 1 (LHP01, thời lượng 2, tiết 1..2, 07:00..08:40)
    row1 = rows[0]
    assert row1["section_id"] == "LHP01"
    assert row1["start_period"] == "1"
    assert row1["end_period"] == "2"
    assert row1["start_time"] == "07:00"
    assert row1["end_time"] == "08:40"
    assert row1["duration_periods"] == "2"

    # Hàng 2 (LHP02, thời lượng 3, tiết 3..5, 08:45..11:25)
    row2 = rows[1]
    assert row2["section_id"] == "LHP02"
    assert row2["start_period"] == "3"
    assert row2["end_period"] == "5"
    assert row2["start_time"] == "08:45"
    assert row2["end_time"] == "11:25"
    assert row2["duration_periods"] == "3"

@pytest.mark.unit
def test_csv_exporter_rejects_hard_violations(exporter_dataset, tmp_path):
    output_path = tmp_path / "invalid_timetable.csv"
    # Lịch trùng nhau ở P101
    bad_sched = Schedule(genes=[
        Gene(section_id="LHP01", timeslot_id=0, room_id="P101"),
        Gene(section_id="LHP02", timeslot_id=0, room_id="P101"),
    ])

    with pytest.raises(ValueError, match="Cannot export schedule with hard violations"):
        export_schedule_to_csv(bad_sched, exporter_dataset, output_path)

@pytest.mark.integration
def test_metadata_json_exporter(tmp_path):
    json_path = tmp_path / "metadata.json"
    metadata = {"seed": 42, "best_hard_violations": 0, "best_soft_penalty": 15}

    path = export_metadata_to_json(metadata, json_path)
    assert os.path.exists(path)

@pytest.mark.integration
def test_excel_exporter_creates_valid_file(exporter_dataset, tmp_path):
    import openpyxl
    from evaluation import export_schedule_to_excel

    output_path = tmp_path / "test_timetable.xlsx"
    sched = Schedule(genes=[
        Gene(section_id="LHP01", timeslot_id=0, room_id="P101"),  # Thứ Hai, tiết 1..2 (07:00..08:40)
        Gene(section_id="LHP02", timeslot_id=2, room_id="LAB01"), # Thứ Hai, tiết 3..5 (08:45..11:25)
    ])

    path = export_schedule_to_excel(sched, exporter_dataset, output_path)
    assert os.path.exists(path)
    assert path.endswith(".xlsx")

    # Xác minh nội dung Excel (Yêu cầu #2, #4, #6, #7, #8, #11)
    wb = openpyxl.load_workbook(path)
    assert "SUMMARY" in wb.sheetnames
    assert "RAW_ASSIGNMENTS" in wb.sheetnames
    ws = wb["RAW_ASSIGNMENTS"]
    num_sections = len(exporter_dataset["course_sections"])
    assert ws.max_row == num_sections + 1
    assert ws["A1"].value == "activity_id"
    assert "section_id" in [cell.value for cell in ws[1]]
    assert "course_name" in [cell.value for cell in ws[1]]

    assert ws.freeze_panes == "A2"

    # Kiểm tra định dạng giá trị ô
    row2_vals = [cell.value for cell in ws[2]]
    assert "LHP01" in row2_vals


@pytest.mark.unit
def test_excel_exporter_rejects_hard_violations_and_deletes_old_file(exporter_dataset, tmp_path):
    from evaluation import export_schedule_to_excel
    output_path = tmp_path / "test_timetable.xlsx"

    # Tạo tệp cũ giả lập trước
    output_path.write_text("old content")
    assert output_path.exists()

    bad_sched = Schedule(genes=[
        Gene(section_id="LHP01", timeslot_id=0, room_id="P101"),
        Gene(section_id="LHP02", timeslot_id=0, room_id="P101"),
    ])

    with pytest.raises(ValueError, match="Cannot export Excel schedule with hard violations"):
        export_schedule_to_excel(bad_sched, exporter_dataset, output_path)

    # Xác minh tệp cũ đã bị xóa (Yêu cầu #14)
    assert not output_path.exists()

@pytest.mark.integration
def test_convergence_visualizer(tmp_path):
    history = [
        {"fitness_evaluations": 10, "best_hard": 5, "best_soft_penalty": 100},
        {"fitness_evaluations": 50, "best_hard": 0, "best_soft_penalty": 40},
    ]
    hard_img = str(tmp_path / "hard.png")
    soft_img = str(tmp_path / "soft.png")

    ConvergenceVisualizer.plot_convergence(
        ga_without_repair_history=history,
        hybrid_ga_history=history,
        random_history=history,
        hard_output_path=hard_img,
        soft_output_path=soft_img,
        evaluation_budget=50,
    )

    assert os.path.exists(hard_img)
    assert os.path.exists(soft_img)
