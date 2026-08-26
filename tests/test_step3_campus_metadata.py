"""Tests for Step 3 — Campus & Preference metadata preservation."""
import io
import copy
import json
import pytest
import openpyxl

from domain import Room, StudentGroup, CourseSection, Timeslot, Lecturer, Gene, Schedule
from dataset import DatasetFactory, DatasetValidator, ExcelDatasetLoader, create_theory_timeslots
from dataset.excel_loader import ExcelValidationError


# ============================================================
# 11.1  Domain model — field preservation
# ============================================================

@pytest.mark.unit
def test_room_campus_id_stored():
    room = Room(id="R1", name="Room 1", capacity=50, campus_id="CS2")
    assert room.campus_id == "CS2"

@pytest.mark.unit
def test_room_campus_id_default_none():
    room = Room(id="R1", name="Room 1", capacity=50)
    assert room.campus_id is None

@pytest.mark.unit
def test_student_group_home_campus_stored():
    grp = StudentGroup(id="G1", name="Group 1", student_count=40, home_campus_id="CS1")
    assert grp.home_campus_id == "CS1"

@pytest.mark.unit
def test_student_group_home_campus_default_none():
    grp = StudentGroup(id="G1", name="Group 1", student_count=40)
    assert grp.home_campus_id is None

@pytest.mark.unit
def test_course_section_preference_fields():
    sec = CourseSection(
        section_id="SEC001", course_id="C01", course_name="Test Course",
        lecturer_id="GV01", group_id="G01", student_count=30,
        preferred_campus_id="CS2", preferred_shift="afternoon", meetings_per_week=1,
    )
    assert sec.preferred_campus_id == "CS2"
    assert sec.preferred_shift == "afternoon"
    assert sec.meetings_per_week == 1

@pytest.mark.unit
def test_course_section_preference_defaults():
    sec = CourseSection("SEC001", "C01", "Course", "GV01", "G01", 30)
    assert sec.preferred_campus_id is None
    assert sec.preferred_shift is None
    assert sec.meetings_per_week == 1

@pytest.mark.unit
@pytest.mark.parametrize("shift", ["morning", "afternoon", "evening"])
def test_course_section_valid_shifts(shift):
    sec = CourseSection("S1", "C1", "C", "G1", "Grp1", 30, preferred_shift=shift)
    assert sec.preferred_shift == shift

@pytest.mark.unit
@pytest.mark.parametrize("bad_shift", ["night", "MORNING", "Chieu"])
def test_course_section_invalid_shift_raises(bad_shift):
    with pytest.raises(ValueError, match="Invalid preferred_shift"):
        CourseSection("S1", "C1", "C", "G1", "Grp1", 30, preferred_shift=bad_shift)

@pytest.mark.unit
def test_course_section_invalid_meetings_per_week_raises():
    with pytest.raises(ValueError, match="meetings_per_week must be >= 1"):
        CourseSection("S1", "C1", "C", "G1", "Grp1", 30, meetings_per_week=0)
    with pytest.raises(ValueError, match="meetings_per_week must be >= 1"):
        CourseSection("S1", "C1", "C", "G1", "Grp1", 30, meetings_per_week=-1)


# ============================================================
# Helper: build minimal in-memory xlsx
# ============================================================

def _build_test_xlsx(
    room_campus="CS2",
    group_home_campus="CS2",
    sec_preferred_campus="CS2",
    sec_preferred_shift="afternoon",
    sec_meetings_per_week=1,
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_campus = wb.create_sheet("CAMPUSES")
    ws_campus.append(["campus_id", "campus_name"])
    for campus_id in sorted({"CS1", room_campus, group_home_campus, sec_preferred_campus}):
        ws_campus.append([campus_id, campus_id])

    ws_ts = wb.create_sheet("TIMESLOTS")
    ws_ts.append(["timeslot_id", "day_name", "period_no", "shift", "start_time", "end_time"])
    ws_ts.append(["TS-M1", "Thu 2", 1, "Sáng", "07:00", "07:50"])
    ws_ts.append(["TS-M2", "Thu 2", 2, "Sáng", "07:50", "08:40"])

    ws_rm = wb.create_sheet("ROOMS")
    ws_rm.append(["room_id", "campus_id", "building", "room_number", "capacity", "room_type"])
    ws_rm.append(["R1", room_campus, "A", "101", 80, "NORMAL"])
    ws_rm.append(["R2", "CS1",      "B", "201", 60, "NORMAL"])

    ws_lec = wb.create_sheet("LECTURERS")
    ws_lec.append(["lecturer_id", "lecturer_name"])
    ws_lec.append(["GV01", "Giang vien 1"])

    ws_avail = wb.create_sheet("LECTURER_AVAILABILITY")
    ws_avail.append(["lecturer_id", "lecturer_name", "TS-M1", "TS-M2"])
    ws_avail.append(["GV01", "Giang vien 1", True, True])

    ws_grp = wb.create_sheet("STUDENT_GROUPS")
    ws_grp.append(["group_id", "group_name", "size", "home_campus_id"])
    ws_grp.append(["G1", "Group 1", 30, group_home_campus])

    ws_crs = wb.create_sheet("COURSES")
    ws_crs.append(["course_id", "course_code", "course_name", "difficulty"])
    ws_crs.append(["C01", "IT001", "Course 1", "MEDIUM"])

    ws_sec = wb.create_sheet("COURSE_SECTIONS")
    ws_sec.append([
        "section_id", "class_code", "course_id", "course_code", "course_name", "lecturer_id",
        "student_group_id", "student_count", "required_room_type",
        "duration_periods", "preferred_campus_id", "preferred_shift",
        "meetings_per_week",
    ])
    ws_sec.append([
        "SEC-001", "CLASS-001", "C01", "IT001", "Course 1", "GV01",
        "G1", 30, "NORMAL", 1,
        sec_preferred_campus, sec_preferred_shift, sec_meetings_per_week,
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# 11.2  Loader tests
# ============================================================

@pytest.mark.unit
def test_loader_reads_room_campus_id(tmp_path):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx(room_campus="CS2"))
    ds = ExcelDatasetLoader.load(str(p))
    rooms_by_id = {r.id: r for r in ds["rooms"]}
    assert rooms_by_id["R1"].campus_id == "CS2"
    assert rooms_by_id["R2"].campus_id == "CS1"

@pytest.mark.unit
def test_loader_reads_group_home_campus_id(tmp_path):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx(group_home_campus="CS2"))
    ds = ExcelDatasetLoader.load(str(p))
    groups_by_id = {g.id: g for g in ds["student_groups"]}
    assert groups_by_id["G1"].home_campus_id == "CS2"

@pytest.mark.unit
def test_loader_reads_section_preferred_campus(tmp_path):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx(sec_preferred_campus="CS2"))
    ds = ExcelDatasetLoader.load(str(p))
    secs_by_id = {s.section_id: s for s in ds["course_sections"]}
    assert secs_by_id["SEC-001"].preferred_campus_id == "CS2"

@pytest.mark.unit
@pytest.mark.parametrize("raw_shift, expected", [
    ("morning", "morning"),
    ("MORNING", "morning"),
    ("afternoon", "afternoon"),
    ("Afternoon", "afternoon"),
    ("evening", "evening"),
    ("EVENING", "evening"),
])
def test_loader_normalizes_preferred_shift(tmp_path, raw_shift, expected):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx(sec_preferred_shift=raw_shift))
    ds = ExcelDatasetLoader.load(str(p))
    secs_by_id = {s.section_id: s for s in ds["course_sections"]}
    assert secs_by_id["SEC-001"].preferred_shift == expected

@pytest.mark.unit
def test_loader_reads_meetings_per_week(tmp_path):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx(sec_meetings_per_week=1))
    ds = ExcelDatasetLoader.load(str(p))
    secs_by_id = {s.section_id: s for s in ds["course_sections"]}
    assert secs_by_id["SEC-001"].meetings_per_week == 1

@pytest.mark.unit
def test_loader_meetings_per_week_above_one_supported(tmp_path):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx(sec_meetings_per_week=2.0))
    dataset = ExcelDatasetLoader.load(str(p))
    assert dataset["course_sections"][0].meetings_per_week == 2

@pytest.mark.unit
def test_loader_meetings_per_week_non_integer_float_rejected(tmp_path):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx(sec_meetings_per_week=2.5))
    with pytest.raises(ExcelValidationError, match="Non-integer value not allowed"):
        ExcelDatasetLoader.load(str(p))

@pytest.mark.unit
def test_loader_meetings_per_week_zero_rejected(tmp_path):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx(sec_meetings_per_week=0))
    with pytest.raises(ExcelValidationError, match="Must be >= 1"):
        ExcelDatasetLoader.load(str(p))

@pytest.mark.unit
def test_loader_invalid_preferred_shift_rejected(tmp_path):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx(sec_preferred_shift="night"))
    with pytest.raises(ExcelValidationError, match="Invalid shift"):
        ExcelDatasetLoader.load(str(p))


# ============================================================
# 11.3  _normalize_optional_str
# ============================================================

@pytest.mark.unit
@pytest.mark.parametrize("raw, expected", [
    ("  CS2  ", "CS2"),
    ("CS1", "CS1"),
    ("", None),
    ("   ", None),
    ("nan", None),
    ("NaN", None),
    (None, None),
])
def test_normalize_optional_str(raw, expected):
    assert ExcelDatasetLoader._normalize_optional_str(raw) == expected


# ============================================================
# 11.4  Validator
# ============================================================

@pytest.mark.unit
def test_validator_rejects_invalid_preferred_campus(small_dataset):
    ds = copy.deepcopy(small_dataset)
    sec = ds["course_sections"][0]
    object.__setattr__(sec, "preferred_campus_id", "CS999")
    report = DatasetValidator.validate_report(ds)
    assert not report["valid"]
    assert any("CS999" in e and "preferred_campus_id" in e for e in report["errors"])

@pytest.mark.unit
def test_validator_rejects_invalid_home_campus(small_dataset):
    ds = copy.deepcopy(small_dataset)
    grp = ds["student_groups"][0]
    object.__setattr__(grp, "home_campus_id", "CS999")
    report = DatasetValidator.validate_report(ds)
    assert not report["valid"]
    assert any("CS999" in e and "home_campus_id" in e for e in report["errors"])

@pytest.mark.unit
def test_validator_rejects_invalid_preferred_shift(small_dataset):
    ds = copy.deepcopy(small_dataset)
    sec = ds["course_sections"][0]
    object.__setattr__(sec, "preferred_shift", "night")
    report = DatasetValidator.validate_report(ds)
    assert not report["valid"]
    assert any("preferred_shift" in e and "night" in e for e in report["errors"])

@pytest.mark.unit
def test_validator_meetings_per_week_gt1_is_supported(small_dataset):
    ds = copy.deepcopy(small_dataset)
    sec = ds["course_sections"][0]
    object.__setattr__(sec, "meetings_per_week", 2)
    report = DatasetValidator.validate_report(ds)
    assert report["valid"]


@pytest.mark.unit
def test_loader_meetings_per_week_boolean_rejected(tmp_path):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_build_test_xlsx(sec_meetings_per_week=True))
    with pytest.raises(ExcelValidationError, match="not a boolean"):
        ExcelDatasetLoader.load(str(p))


def _mutate_test_workbook(raw_bytes, mutation):
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
    mutation(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.unit
@pytest.mark.parametrize(
    "mutation, message",
    [
        (lambda wb: wb["LECTURER_AVAILABILITY"].cell(1, 5, "UNKNOWN"), "unknown timeslot columns"),
        (lambda wb: wb["LECTURER_AVAILABILITY"].cell(1, 4, "TS-M1"), "duplicate timeslot columns"),
        (lambda wb: wb["LECTURER_AVAILABILITY"].delete_cols(4), "missing timeslot columns"),
        (lambda wb: wb["LECTURER_AVAILABILITY"].cell(2, 3, 1), "must be boolean"),
        (lambda wb: wb["LECTURER_AVAILABILITY"].cell(2, 1, "GV404"), "unknown lecturer_id"),
        (lambda wb: wb["LECTURER_AVAILABILITY"].append(["GV01", "Duplicate", True, True]), "Duplicate lecturer_id"),
    ],
)
def test_loader_strict_lecturer_availability_validation(tmp_path, mutation, message):
    p = tmp_path / "test.xlsx"
    p.write_bytes(_mutate_test_workbook(_build_test_xlsx(), mutation))
    with pytest.raises(ExcelValidationError, match=message):
        ExcelDatasetLoader.load(str(p))


@pytest.mark.unit
def test_course_and_class_codes_round_trip_all_exports(tmp_path):
    from constraints import SoftConstraintConfig
    from evaluation import export_schedule_query_data, export_schedule_to_excel

    source = tmp_path / "source.xlsx"
    source.write_bytes(_build_test_xlsx())
    dataset = ExcelDatasetLoader.load_and_validate(str(source))
    schedule = Schedule([Gene("SEC-001", "R1", 0)])
    soft_config = SoftConstraintConfig.from_profile("student-centric")

    normalized = tmp_path / "normalized.json"
    ExcelDatasetLoader.export_normalized_json(dataset, str(normalized))
    restored = ExcelDatasetLoader.load_normalized_json(str(normalized))
    assert restored["courses"][0].course_code == "IT001"
    assert restored["course_sections"][0].class_code == "CLASS-001"

    query_path = tmp_path / "query.json"
    export_schedule_query_data(schedule, dataset, query_path, soft_config=soft_config)
    query_data = json.loads(query_path.read_text(encoding="utf-8"))
    assignment = query_data["assignments"][0]
    assert assignment["course_code"] == "IT001"
    assert assignment["class_code"] == "CLASS-001"
    assert assignment["campus_id"] == "CS2"
    assert query_data["meta"]["effective_soft_constraints"] == soft_config.to_metadata()

    dataset_without_campus = copy.deepcopy(dataset)
    dataset_without_campus["rooms"][0].campus_id = None
    no_campus_query = tmp_path / "query-no-campus.json"
    export_schedule_query_data(
        schedule,
        dataset_without_campus,
        no_campus_query,
        soft_config=soft_config,
    )
    no_campus_assignment = json.loads(
        no_campus_query.read_text(encoding="utf-8")
    )["assignments"][0]
    assert no_campus_assignment["campus_id"] is None

    xlsx_path = tmp_path / "schedule.xlsx"
    export_schedule_to_excel(schedule, dataset, xlsx_path, soft_config=soft_config)
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sheet_name in ("RAW_ASSIGNMENTS", "SCHEDULE_BY_LECTURER", "SCHEDULE_BY_ROOM"):
        header = [cell.value for cell in workbook[sheet_name][1]]
        row = [cell.value for cell in workbook[sheet_name][2]]
        assert row[header.index("course_code")] == "IT001"
        assert row[header.index("class_code")] == "CLASS-001"
    run_config = {
        row[0].value: row[1].value
        for row in workbook["RUN_CONFIG"].iter_rows(min_row=2)
    }
    assert json.loads(run_config["effective_soft_constraints"]) == soft_config.to_metadata()

@pytest.mark.unit
def test_validator_meetings_per_week_zero_is_error(small_dataset):
    ds = copy.deepcopy(small_dataset)
    sec = ds["course_sections"][0]
    object.__setattr__(sec, "meetings_per_week", 0)
    report = DatasetValidator.validate_report(ds)
    assert not report["valid"]
    assert any("meetings_per_week" in e for e in report["errors"])


# ============================================================
# 11.5  Exporter round-trip tests
# ============================================================

@pytest.mark.unit
def test_exporter_campus_from_domain_not_hardcoded(tmp_path):
    from evaluation import export_schedule_to_excel
    from dataset import find_feasible_schedule

    timeslots = create_theory_timeslots(days=["Thu 2", "Thu 3"], max_period=16)
    rooms = [Room(id="R_CS2", name="A-101", capacity=100, room_type="NORMAL", campus_id="CS2")]
    lecturers = [Lecturer(id="GV01", name="GV1")]
    groups = [StudentGroup(id="G1", name="G1", student_count=30, home_campus_id="CS2")]
    sections = [CourseSection(
        "SEC001", "C01", "Course 1", "GV01", "G1", 30,
        preferred_campus_id="CS2", preferred_shift="afternoon", meetings_per_week=1,
    )]
    ds = {"timeslots": timeslots, "rooms": rooms, "lecturers": lecturers,
          "student_groups": groups, "course_sections": sections}
    sched = find_feasible_schedule(ds)
    assert sched is not None

    out = str(tmp_path / "roundtrip.xlsx")
    export_schedule_to_excel(sched, ds, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["RAW_ASSIGNMENTS"]
    header = [cell.value for cell in ws[1]]
    campus_idx = header.index("campus_id")
    pref_campus_idx = header.index("preferred_campus_id")
    pref_shift_idx = header.index("preferred_shift")

    for row in ws.iter_rows(min_row=2, values_only=True):
        assert row[campus_idx] == "CS2", f"Expected 'CS2', got '{row[campus_idx]}'"
        assert row[pref_campus_idx] == "CS2", f"Expected 'CS2', got '{row[pref_campus_idx]}'"
        assert row[pref_shift_idx] == "afternoon", f"Expected 'afternoon', got '{row[pref_shift_idx]}'"

@pytest.mark.unit
def test_exporter_none_campus_exported_as_empty_not_cs1(tmp_path):
    from evaluation import export_schedule_to_excel
    from dataset import find_feasible_schedule

    timeslots = create_theory_timeslots(days=["Thu 2"], max_period=16)
    rooms = [Room(id="R1", name="Room1", capacity=100, room_type="NORMAL")]  # campus_id=None
    lecturers = [Lecturer(id="GV01", name="GV1")]
    groups = [StudentGroup(id="G1", name="G1", student_count=30)]
    sections = [CourseSection("SEC001", "C01", "Course 1", "GV01", "G1", 30)]
    ds = {"timeslots": timeslots, "rooms": rooms, "lecturers": lecturers,
          "student_groups": groups, "course_sections": sections}
    sched = find_feasible_schedule(ds)
    assert sched is not None

    out = str(tmp_path / "no_campus.xlsx")
    export_schedule_to_excel(sched, ds, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["RAW_ASSIGNMENTS"]
    header = [cell.value for cell in ws[1]]
    campus_idx = header.index("campus_id")
    pref_campus_idx = header.index("preferred_campus_id")
    pref_shift_idx = header.index("preferred_shift")

    for row in ws.iter_rows(min_row=2, values_only=True):
        # openpyxl returns None for empty cells, "" for cells written with empty string
        # Both are acceptable — the key assertion is NO hard-coded "CS1"/"morning"
        assert row[campus_idx] != "CS1", "Hard-coded 'CS1' found — should be empty/None"
        assert row[pref_campus_idx] != "CS1", "Hard-coded 'CS1' found in preferred_campus_id"
        assert row[pref_shift_idx] != "morning", "Hard-coded 'morning' found in preferred_shift"
        # Must be either None or empty string (not a hardcoded value)
        assert row[campus_idx] in (None, ""), f"Unexpected campus_id: '{row[campus_idx]}'"
        assert row[pref_campus_idx] in (None, ""), f"Unexpected preferred_campus_id: '{row[pref_campus_idx]}'"
        assert row[pref_shift_idx] in (None, ""), f"Unexpected preferred_shift: '{row[pref_shift_idx]}'"



# ============================================================
# 11.8  Mock dataset verification
# ============================================================

@pytest.mark.unit
def test_medium_dataset_rooms_have_campus_id(medium_dataset):
    for room in medium_dataset["rooms"]:
        assert room.campus_id is not None, f"Room {room.id} missing campus_id"
        assert room.campus_id in {"CS1", "CS2"}

@pytest.mark.unit
def test_medium_dataset_rooms_campus_diversity(medium_dataset):
    campuses = {r.campus_id for r in medium_dataset["rooms"]}
    assert "CS1" in campuses and "CS2" in campuses

@pytest.mark.unit
def test_medium_dataset_groups_have_home_campus(medium_dataset):
    for grp in medium_dataset["student_groups"]:
        assert grp.home_campus_id is not None, f"Group {grp.id} missing home_campus_id"

@pytest.mark.unit
def test_medium_dataset_sections_have_preferred_fields(medium_dataset):
    for sec in medium_dataset["course_sections"]:
        assert sec.preferred_campus_id is not None
        assert sec.preferred_shift is not None
        assert sec.preferred_shift in {"morning", "afternoon", "evening"}
        assert sec.meetings_per_week == 1

@pytest.mark.unit
def test_medium_dataset_preferred_shift_diversity(medium_dataset):
    shifts = {s.preferred_shift for s in medium_dataset["course_sections"]}
    assert len(shifts) >= 2, f"Expected >= 2 shift types, got {shifts}"
