"""Tests for Step 5.1 — Metrics Audit, Invariants, and Reporting Integrity."""

import pytest
import openpyxl
from dataclasses import FrozenInstanceError

from domain import Schedule, Gene, EvaluationCounters, RepairStatus
from constraints import ConstraintEvaluator, ScheduleRepairEngine, SoftConstraintConfig
from evaluation.run_metrics import RunMetrics, AggregateRunMetrics
from evaluation.benchmark_statistics import aggregate_run_results
from evaluation.schedule_exporter import export_schedule_to_excel
from ga import GeneticAlgorithmEngine
from evaluation.baselines import GreedyScheduler, RandomSearchScheduler


# ============================================================
# 1. Kiểm tra tính bất biến của chỉ số Repair theo từng lần chạy
# ============================================================

@pytest.mark.unit
def test_run_metrics_repair_invariant_validation():
    # Chỉ số hợp lệ
    valid_m = RunMetrics(
        method="Hybrid GA + Repair", seed=0, runtime_seconds=1.0,
        time_to_first_feasible_seconds=0.5, search_fitness_evaluations=100,
        hard_constraint_evaluations=100, soft_constraint_evaluations=100,
        total_constraint_evaluations=200, candidate_checks=50,
        repair_calls=10, repair_improved=7, repair_unchanged=2, repair_failed=1,
        first_feasible_search_evaluation=10, first_feasible_total_constraint_evaluation=20,
        first_feasible_generation=1, final_hard_violations=0, final_soft_penalty=50,
        feasible=True, score=50.0
    )
    valid_m.validate()
    assert valid_m.repair_calls == valid_m.repair_improved + valid_m.repair_unchanged + valid_m.repair_failed

    # Chỉ số không hợp lệ: repair_calls != tổng các trạng thái
    with pytest.raises(ValueError, match="Invalid repair metrics"):
        RunMetrics(
            method="Hybrid GA + Repair", seed=0, runtime_seconds=1.0,
            time_to_first_feasible_seconds=0.5, search_fitness_evaluations=100,
            hard_constraint_evaluations=100, soft_constraint_evaluations=100,
            total_constraint_evaluations=200, candidate_checks=50,
            repair_calls=10, repair_improved=7, repair_unchanged=2, repair_failed=0,  # tổng=9 != 10
            first_feasible_search_evaluation=10, first_feasible_total_constraint_evaluation=20,
            first_feasible_generation=1, final_hard_violations=0, final_soft_penalty=50,
            feasible=True, score=50.0
        )


@pytest.mark.unit
def test_run_metrics_ttff_invariants():
    # Lần chạy khả thi với TTFF=None phải thất bại
    with pytest.raises(ValueError, match="Feasible run .* must have time_to_first_feasible_seconds"):
        RunMetrics(
            method="Hybrid GA + Repair", seed=0, runtime_seconds=1.0,
            time_to_first_feasible_seconds=None, search_fitness_evaluations=100,
            hard_constraint_evaluations=100, soft_constraint_evaluations=100,
            total_constraint_evaluations=200, candidate_checks=50,
            repair_calls=0, repair_improved=0, repair_unchanged=0, repair_failed=0,
            first_feasible_search_evaluation=10, first_feasible_total_constraint_evaluation=20,
            first_feasible_generation=1, final_hard_violations=0, final_soft_penalty=50,
            feasible=True, score=50.0
        )

    # Lần chạy không khả thi với TTFF!=None phải thất bại
    with pytest.raises(ValueError, match="Infeasible run .* must have time_to_first_feasible_seconds=None"):
        RunMetrics(
            method="GA without Repair", seed=0, runtime_seconds=1.0,
            time_to_first_feasible_seconds=0.5, search_fitness_evaluations=100,
            hard_constraint_evaluations=100, soft_constraint_evaluations=100,
            total_constraint_evaluations=200, candidate_checks=0,
            repair_calls=0, repair_improved=0, repair_unchanged=0, repair_failed=0,
            first_feasible_search_evaluation=None, first_feasible_total_constraint_evaluation=None,
            first_feasible_generation=None, final_hard_violations=2, final_soft_penalty=50,
            feasible=False, score=2050.0
        )


# ============================================================
# 2. Kiểm tra tổng hợp và tính bất biến
# ============================================================

@pytest.mark.unit
def test_aggregate_totals_invariant():
    runs = [
        {
            "score": 100, "hard_violations": 0, "soft_penalty": 100, "runtime_seconds": 1.0,
            "search_fitness_evaluations": 100, "total_constraint_evaluations": 200, "candidate_checks": 50,
            "repair_calls": 10, "repair_improved": 8, "repair_unchanged": 2, "repair_failed": 0,
            "time_to_first_feasible_seconds": 0.5, "is_hard_feasible": True, "is_perfect": False,
        },
        {
            "score": 150, "hard_violations": 0, "soft_penalty": 150, "runtime_seconds": 1.2,
            "search_fitness_evaluations": 100, "total_constraint_evaluations": 200, "candidate_checks": 40,
            "repair_calls": 10, "repair_improved": 7, "repair_unchanged": 2, "repair_failed": 1,
            "time_to_first_feasible_seconds": 0.6, "is_hard_feasible": True, "is_perfect": False,
        },
        {
            "score": 120, "hard_violations": 0, "soft_penalty": 120, "runtime_seconds": 1.1,
            "search_fitness_evaluations": 100, "total_constraint_evaluations": 200, "candidate_checks": 45,
            "repair_calls": 10, "repair_improved": 9, "repair_unchanged": 1, "repair_failed": 0,
            "time_to_first_feasible_seconds": 0.4, "is_hard_feasible": True, "is_perfect": False,
        },
    ]

    agg = aggregate_run_results("Hybrid GA + Repair", runs)
    assert agg["total_repair_calls"] == 30
    assert agg["total_repair_improved"] == 24
    assert agg["total_repair_unchanged"] == 5
    assert agg["total_repair_failed"] == 1
    assert agg["total_repair_calls"] == (
        agg["total_repair_improved"] + agg["total_repair_unchanged"] + agg["total_repair_failed"]
    )
    assert pytest.approx(agg["improvement_rate"], 0.001) == 24 / 30
    assert pytest.approx(agg["non_failure_rate"], 0.001) == 29 / 30


@pytest.mark.unit
def test_aggregate_raises_on_invalid_run_metrics():
    bad_runs = [
        {
            "score": 100, "hard_violations": 0, "soft_penalty": 100, "runtime_seconds": 1.0,
            "search_fitness_evaluations": 100, "total_constraint_evaluations": 200, "candidate_checks": 50,
            "repair_calls": 10, "repair_improved": 5, "repair_unchanged": 2, "repair_failed": 0,  # tổng=7 != 10
            "time_to_first_feasible_seconds": 0.5, "is_hard_feasible": True, "is_perfect": False,
        }
    ]
    with pytest.raises(ValueError, match="Invalid repair metrics"):
        aggregate_run_results("Hybrid GA + Repair", bad_runs)


# ============================================================
# 3. Kiểm thử phân tích loại đánh giá ràng buộc
# ============================================================

@pytest.mark.unit
def test_constraint_evaluation_category_breakdown(small_dataset):
    evaluator = ConstraintEvaluator(small_dataset)
    sections = small_dataset["course_sections"]
    rooms = small_dataset["rooms"]
    genes = [Gene(s.section_id, rooms[0].id, 0) for s in sections]
    sched = Schedule(genes=genes)

    # 1. Đánh giá tìm kiếm
    evaluator.evaluate_hard(sched, category="search")
    evaluator.evaluate_soft(sched, category="search")
    assert evaluator.counters.search_hard_constraint_evaluations == 1
    assert evaluator.counters.search_soft_constraint_evaluations == 1
    assert evaluator.counters.search_constraint_evaluations == 2

    # 2. Đánh giá nội bộ (Repair)
    evaluator.evaluate_hard(sched, category="internal")
    evaluator.evaluate_soft(sched, category="internal")
    assert evaluator.counters.internal_hard_constraint_evaluations == 1
    assert evaluator.counters.internal_soft_constraint_evaluations == 1
    assert evaluator.counters.internal_constraint_evaluations == 2

    # 3. Đánh giá báo cáo
    evaluator.evaluate_unified(sched, category="reporting")
    assert evaluator.counters.reporting_hard_constraint_evaluations == 1
    assert evaluator.counters.reporting_soft_constraint_evaluations == 1
    assert evaluator.counters.reporting_constraint_evaluations == 2

    # Tổng số lần đánh giá ràng buộc được cộng chính xác
    assert evaluator.counters.total_constraint_evaluations == 6


# ============================================================
# 4. Kiểm thử tính bất biến của SoftConstraintConfig
# ============================================================

@pytest.mark.unit
def test_soft_constraint_config_immutability():
    config = SoftConstraintConfig.default()

    # Việc thử gán mục mới vào definitions sẽ phát sinh TypeError
    with pytest.raises(TypeError):
        config.definitions["S1"] = None  # type: ignore

    # Việc thử thay đổi thuộc tính dataclass bên trong sẽ phát sinh FrozenInstanceError
    with pytest.raises(FrozenInstanceError):
        config.definitions["compact_student_schedule"].weight = 99  # type: ignore


# ============================================================
# 5. Lượt xuất-nạp trang tính tổng kết và chỉ số Excel
# ============================================================

@pytest.mark.unit
def test_excel_summary_and_metrics_sheet_export(tmp_path, small_dataset):
    from dataset import find_feasible_schedule

    sched = find_feasible_schedule(small_dataset)
    assert sched is not None

    m = RunMetrics(
        method="Hybrid GA + Repair", seed=0, runtime_seconds=1.2,
        time_to_first_feasible_seconds=0.4, search_fitness_evaluations=100,
        hard_constraint_evaluations=120, soft_constraint_evaluations=120,
        total_constraint_evaluations=240, candidate_checks=30,
        repair_calls=10, repair_improved=8, repair_unchanged=2, repair_failed=0,
        first_feasible_search_evaluation=20, first_feasible_total_constraint_evaluation=48,
        first_feasible_generation=2, final_hard_violations=0, final_soft_penalty=100,
        feasible=True, score=100.0
    )

    summary_data = {
        "method": "Hybrid GA + Repair",
        "run_count": 1,
        "is_deterministic": False,
        "feasible_count": 1,
        "feasible_rate": 1.0,
        "median_final_hard": 0.0,
        "median_final_soft": 100.0,
        "mean_final_soft": 100.0,
        "median_runtime_seconds": 1.2,
        "median_time_to_first_feasible_seconds": 0.4,
        "median_search_fitness_evaluations": 100,
        "median_total_constraint_evaluations": 240,
        "median_candidate_checks": 30,
        "median_repair_calls": 10,
        "total_repair_calls": 10,
        "total_repair_improved": 8,
        "total_repair_unchanged": 2,
        "total_repair_failed": 0,
        "improvement_rate": 0.8,
        "non_failure_rate": 1.0,
    }

    out_file = str(tmp_path / "test_benchmark_export.xlsx")
    meta = {
        "all_runs_flat": [m.to_dict()],
        "summary_list": [summary_data],
    }
    export_schedule_to_excel(sched, small_dataset, out_file, metadata=meta)

    wb = openpyxl.load_workbook(out_file)
    assert "RUN_METRICS" in wb.sheetnames
    assert "BENCHMARK_SUMMARY" in wb.sheetnames

    # Kiểm tra các cột của trang tính BENCHMARK_SUMMARY
    ws_s = wb["BENCHMARK_SUMMARY"]
    rows = list(ws_s.iter_rows(values_only=True))
    assert len(rows) == 2  # tiêu đề + 1 hàng dữ liệu
    header = rows[0]
    data = rows[1]

    assert header[0] == "method"
    assert header[header.index("total_repair_calls")] == "total_repair_calls"
    assert data[header.index("total_repair_calls")] == 10
    assert data[header.index("total_repair_improved")] == 8
    assert data[header.index("total_repair_unchanged")] == 2
    assert data[header.index("total_repair_failed")] == 0
    assert data[header.index("improvement_rate")] == 0.8
