import random
import pytest
import openpyxl
from pathlib import Path

from domain import Schedule, Gene, CourseSection, Room, Timeslot, Lecturer
from constraints import ConstraintEvaluator, ScheduleRepairEngine
from constraints.evaluator import UnifiedEvaluationResult, SoftBreakdownItem
from ga import GeneticAlgorithmEngine
from evaluation.schedule_exporter import export_schedule_to_excel


@pytest.fixture
def sample_dataset():
    random.seed(42)
    sections = [
        CourseSection(section_id="SEC01", course_id="CRS01", course_name="Course 1", lecturer_id="LEC01", group_id="GRP01", student_count=30, duration_periods=2, is_difficult=True),
        CourseSection(section_id="SEC02", course_id="CRS02", course_name="Course 2", lecturer_id="LEC01", group_id="GRP01", student_count=30, duration_periods=2, is_difficult=False),
        CourseSection(section_id="SEC03", course_id="CRS03", course_name="Course 3", lecturer_id="LEC02", group_id="GRP02", student_count=20, duration_periods=1, is_difficult=True),
    ]
    rooms = [
        Room(id="R101", name="Room 101", capacity=40, room_type="NORMAL"),
        Room(id="R102", name="Room 102", capacity=50, room_type="NORMAL"),
    ]
    timeslots = [
        Timeslot(id=1, day="Thứ 2", period=1, start_time="07:00", end_time="07:50", session="morning"),
        Timeslot(id=2, day="Thứ 2", period=2, start_time="07:55", end_time="08:45", session="morning"),
        Timeslot(id=3, day="Thứ 2", period=3, start_time="08:50", end_time="09:40", session="morning"),
        Timeslot(id=4, day="Thứ 2", period=4, start_time="09:45", end_time="10:35", session="morning"),
        Timeslot(id=5, day="Thứ 2", period=5, start_time="10:40", end_time="11:30", session="morning"),
        Timeslot(id=7, day="Thứ 2", period=7, start_time="13:00", end_time="13:50", session="afternoon"),
        Timeslot(id=8, day="Thứ 2", period=8, start_time="13:55", end_time="14:45", session="afternoon"),
    ]
    lecturers = [
        Lecturer(id="LEC01", name="Lecturer 1"),
        Lecturer(id="LEC02", name="Lecturer 2"),
    ]
    return {
        "course_sections": sections,
        "rooms": rooms,
        "timeslots": timeslots,
        "lecturers": lecturers,
    }


def test_unified_evaluation_result_breakdown_identity(sample_dataset):
    random.seed(42)
    evaluator = ConstraintEvaluator(sample_dataset)
    sched = Schedule(genes=[
        Gene("SEC01", "R101", 7),  # Buổi chiều khó xếp
        Gene("SEC02", "R101", 1),
        Gene("SEC03", "R102", 2),
    ])
    unified = evaluator.evaluate_unified(sched)

    calculated_soft = sum(item.weighted_penalty for item in unified.soft_breakdown)
    assert unified.soft_penalty == calculated_soft, (
        f"Expected soft_penalty {unified.soft_penalty} == sum of weighted penalties {calculated_soft}"
    )


def test_workbook_violations_total_equals_summary(sample_dataset, tmp_path):
    random.seed(42)
    evaluator = ConstraintEvaluator(sample_dataset)
    sched = Schedule(genes=[
        Gene("SEC01", "R101", 7),
        Gene("SEC02", "R102", 1),
        Gene("SEC03", "R102", 3),
    ])
    unified = evaluator.evaluate_unified(sched)

    out_file = tmp_path / "test_timetable.xlsx"
    export_schedule_to_excel(
        schedule=sched,
        dataset=sample_dataset,
        output_path=out_file,
        metadata={"method": "Hybrid GA + Repair", "repair_calls": 5, "repair_successes": 5},
        allow_infeasible_export=True
    )

    wb = openpyxl.load_workbook(out_file)

    # Đọc soft_penalty trong SUMMARY
    ws_sum = wb["SUMMARY"]
    sum_soft_penalty = None
    for row in ws_sum.iter_rows(values_only=True):
        if row[0] == "soft_penalty":
            sum_soft_penalty = float(row[1])
            break

    assert sum_soft_penalty is not None
    assert sum_soft_penalty == unified.soft_penalty

    # Đọc TOTAL SOFT PENALTY trong VIOLATIONS
    ws_viol = wb["VIOLATIONS"]
    viol_total_penalty = None
    for row in ws_viol.iter_rows(values_only=True):
        if row[0] == "SUMMARY" and row[2] == "TOTAL_SOFT_PENALTY":
            viol_total_penalty = float(row[11])  # weighted_penalty vẫn ở cột 12
            break

    assert viol_total_penalty is not None
    assert viol_total_penalty == pytest.approx(sum_soft_penalty)
    assert sum_soft_penalty == pytest.approx(unified.soft_penalty)


def test_repair_engine_direct_hard_conflict(sample_dataset):
    """TEST A: Direct Repair Engine execution on guaranteed hard conflict."""
    random.seed(42)
    repairer = ScheduleRepairEngine(sample_dataset)
    evaluator = ConstraintEvaluator(sample_dataset)

    # Lịch xung đột: SEC01 và SEC02 cùng giảng viên (LEC01) tại khung giờ 1
    conflicting_sched = Schedule(genes=[
        Gene("SEC01", "R101", 1),
        Gene("SEC02", "R102", 1),
        Gene("SEC03", "R102", 3),
    ])

    h_before, _ = evaluator.evaluate_hard(conflicting_sched)
    soft_before, _ = evaluator.evaluate_soft(conflicting_sched)
    assert h_before > 0, "Initial schedule must have hard violations before repair"

    repairer.stats.reset()
    res = repairer.repair(conflicting_sched)

    h_after, _ = evaluator.evaluate_hard(res.schedule)
    soft_after, _ = evaluator.evaluate_soft(res.schedule)

    assert repairer.stats.repair_calls == 1
    assert repairer.stats.repair_attempts > 0
    assert repairer.stats.candidate_checks > 0
    assert (h_after, soft_after) <= (h_before, soft_before)


def test_ga_engine_calls_repair_on_conflicting_offspring(sample_dataset, monkeypatch):
    """TEST B: Deterministic integration test for GA + Repair using injected conflicting offspring."""
    random.seed(42)
    ga = GeneticAlgorithmEngine(sample_dataset, pop_size=4)

    # Chèn một cá thể con chắc chắn xung đột bằng cách thay tạm GAOperators.mutate
    def mock_mutate(sched, rooms, timeslots, rate, **kwargs):
        return Schedule(genes=[
            Gene("SEC01", "R101", 1),
            Gene("SEC02", "R101", 1),  # Xung đột cứng tại R101, tiết 1
            Gene("SEC03", "R102", 1),
        ])

    monkeypatch.setattr("ga.operators.GAOperators.mutate", mock_mutate)

    spy_calls = [0]
    original_repair = ga.repairer.repair

    def spy_repair(sched, max_attempts=15):
        spy_calls[0] += 1
        return original_repair(sched, max_attempts=max_attempts)

    monkeypatch.setattr(ga.repairer, "repair", spy_repair)

    res = ga.run(generations=2, use_repair=True, evaluation_budget=20)

    rep_stats = res.get("repair_stats", {})
    assert rep_stats.get("repair_enabled") is True
    assert spy_calls[0] > 0, "RepairEngine.repair must be invoked on conflicting offspring"
    assert rep_stats.get("repair_calls", 0) > 0


def test_repair_does_not_worsen_lexicographic_tuple(sample_dataset):
    random.seed(42)
    repairer = ScheduleRepairEngine(sample_dataset)
    evaluator = ConstraintEvaluator(sample_dataset)
    sched = Schedule(genes=[
        Gene("SEC01", "R101", 1),
        Gene("SEC02", "R101", 1),
        Gene("SEC03", "R102", 1),
    ])

    h_before, s_before = evaluator.calculate_fitness(sched)[1:]
    res = repairer.repair(sched)
    h_after, s_after = evaluator.calculate_fitness(res.schedule)[1:]

    assert (h_after, s_after) <= (h_before, s_before)


def test_export_prohibits_infeasible_schedule_unless_allowed(sample_dataset, tmp_path):
    random.seed(42)
    infeasible_sched = Schedule(genes=[
        Gene("SEC01", "R101", 1),
        Gene("SEC02", "R101", 1),  # Xung đột cứng
        Gene("SEC03", "R102", 1),
    ])

    out_file = tmp_path / "infeasible_test.xlsx"

    with pytest.raises(ValueError, match="Cannot export Excel schedule with hard violations"):
        export_schedule_to_excel(
            schedule=infeasible_sched,
            dataset=sample_dataset,
            output_path=out_file,
            allow_infeasible_export=False
        )

    # Phải thành công khi allow_infeasible_export=True
    exported_path = export_schedule_to_excel(
        schedule=infeasible_sched,
        dataset=sample_dataset,
        output_path=out_file,
        allow_infeasible_export=True
    )
    assert Path(exported_path).exists()
