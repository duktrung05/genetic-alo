"""Build a self-contained, Git-free Google Colab notebook suite.

The setup notebook embeds a compressed snapshot of the project.  It writes that
snapshot to Google Drive, while the remaining notebooks restore it into each
fresh Colab runtime before running tests, experiments, or the Streamlit demo.
"""

from __future__ import annotations

import hashlib
import io
import json
import os
from pathlib import Path
import shutil
import zipfile


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "colab"

TOP_LEVEL_FILES = {
    "Dockerfile",
    "README.md",
    "docker-compose.yml",
    "main.py",
    "main_benchmark.py",
    "requirements.txt",
    "ui_app.py",
}
SOURCE_DIRECTORIES = {
    "constraints",
    "data",
    "dataset",
    "domain",
    "evaluation",
    "ga",
    "outputs",
    "schedule_assistant",
    "scripts",
    "tests",
}
EXCLUDED_NAMES = {
    ".git",
    ".pytest_cache",
    ".venv",
    "__pycache__",
}
EXCLUDED_SUFFIXES = {".log", ".pyc", ".pyo", ".tmp"}


def markdown(source: str) -> dict:
    return {
        "cell_type": "markdown",
        "metadata": {},
        "source": source.strip().splitlines(keepends=True),
    }


def code(source: str) -> dict:
    return {
        "cell_type": "code",
        "execution_count": None,
        "metadata": {},
        "outputs": [],
        "source": source.strip().splitlines(keepends=True),
    }


def notebook(name: str, cells: list[dict]) -> dict:
    return {
        "cells": cells,
        "metadata": {
            "colab": {
                "name": name,
                "provenance": [],
            },
            "kernelspec": {
                "display_name": "Python 3",
                "name": "python3",
            },
            "language_info": {"name": "python"},
        },
        "nbformat": 4,
        "nbformat_minor": 0,
    }


def iter_payload_files() -> list[Path]:
    files: list[Path] = []
    for current_root, directories, filenames in os.walk(ROOT):
        current_path = Path(current_root)
        relative_root = current_path.relative_to(ROOT)
        directories[:] = [
            name
            for name in directories
            if name not in EXCLUDED_NAMES and name != "colab"
        ]
        if relative_root != Path(".") and relative_root.parts[0] not in SOURCE_DIRECTORIES:
            directories[:] = []
            continue
        for filename in filenames:
            path = current_path / filename
            relative = path.relative_to(ROOT)
            if path.suffix.lower() in EXCLUDED_SUFFIXES:
                continue
            if len(relative.parts) == 1:
                if relative.name in TOP_LEVEL_FILES:
                    files.append(path)
                continue
            if relative.parts[0] in SOURCE_DIRECTORIES:
                files.append(path)
    return sorted(files, key=lambda item: item.as_posix().lower())


def build_payload() -> tuple[bytes, str, int]:
    files = iter_payload_files()
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in files:
            archive.write(path, path.relative_to(ROOT).as_posix())
    payload = buffer.getvalue()
    return payload, hashlib.sha256(payload).hexdigest(), len(files)


def bootstrap_source(checksum: str = "") -> str:
    del checksum
    return '''
#@title Clone dự án từ GitHub và cài môi trường Colab
from google.colab import drive
drive.mount("/content/drive", force_remount=False)

import os
from pathlib import Path
import shutil
import subprocess
import sys

THU_MUC_COLAB_DRIVE = Path("/content/drive/MyDrive/Genetic_ALO_Colab")
THU_MUC_COLAB_DRIVE.mkdir(parents=True, exist_ok=True)
THU_MUC_DU_AN = Path("/content/genetic-alo")
THU_MUC_KET_QUA_DRIVE = THU_MUC_COLAB_DRIVE / "latest_outputs"
REPOSITORY_URL = "https://github.com/duktrung05/genetic-alo.git"

shutil.rmtree(THU_MUC_DU_AN, ignore_errors=True)
subprocess.run(
    [
        "git", "clone", "--depth", "1", "--branch", "main",
        REPOSITORY_URL, str(THU_MUC_DU_AN),
    ],
    check=True,
)

# Khôi phục output mới nhất từ Drive nếu notebook trước đã tạo kết quả.
if THU_MUC_KET_QUA_DRIVE.is_dir():
    shutil.copytree(
        THU_MUC_KET_QUA_DRIVE,
        THU_MUC_DU_AN / "outputs",
        dirs_exist_ok=True,
    )

subprocess.run(
    [
        sys.executable, "-m", "pip", "install", "--quiet",
        "--disable-pip-version-check", "-r",
        str(THU_MUC_DU_AN / "requirements.txt"),
    ],
    check=True,
)

os.chdir(THU_MUC_DU_AN)
if str(THU_MUC_DU_AN) not in sys.path:
    sys.path.insert(0, str(THU_MUC_DU_AN))

def dong_bo_ket_qua() -> Path:
    """Copy all current outputs to Drive so another notebook can reuse them."""
    THU_MUC_KET_QUA_DRIVE.mkdir(parents=True, exist_ok=True)
    shutil.copytree(
        THU_MUC_DU_AN / "outputs",
        THU_MUC_KET_QUA_DRIVE,
        dirs_exist_ok=True,
    )
    return THU_MUC_KET_QUA_DRIVE

print(f"✅ Đã clone dự án tại: {THU_MUC_DU_AN}")
print(f"✅ Python: {sys.version.split()[0]}")
subprocess.run(["git", "log", "-1", "--oneline"], cwd=THU_MUC_DU_AN, check=True)
print("✅ Dataset Excel nằm trong data/instances.")
'''


def write_notebook(filename: str, cells: list[dict]) -> None:
    path = OUTPUT_DIR / filename
    path.write_text(
        json.dumps(notebook(filename, cells), ensure_ascii=False, indent=1) + "\n",
        encoding="utf-8",
    )


def build_setup_notebook(checksum: str, file_count: int) -> None:
    setup_code = f'''
#@title Upload và giải nén dự án trực tiếp trên Google Colab
from google.colab import drive, files
drive.mount("/content/drive", force_remount=False)

import hashlib
import io
import os
from pathlib import Path
import shutil
import subprocess
import sys
import zipfile

MA_BAM_SOURCE = "{checksum}"
SO_TEP_DU_KIEN = {file_count}
TEN_FILE_SOURCE = "Genetic_ALO_Project.zip"

THU_MUC_COLAB_DRIVE = Path("/content/drive/MyDrive/Genetic_ALO_Colab")
THU_MUC_COLAB_DRIVE.mkdir(parents=True, exist_ok=True)
GOI_SOURCE = THU_MUC_COLAB_DRIVE / "project-source.zip"

# Chọn file Genetic_ALO_Project.zip đi kèm bộ notebook.
print(f"Hãy chọn file: {{TEN_FILE_SOURCE}}")
tep_da_tai = files.upload()
if TEN_FILE_SOURCE in tep_da_tai:
    du_lieu_zip = tep_da_tai[TEN_FILE_SOURCE]
elif len(tep_da_tai) == 1:
    ten_thuc_te, du_lieu_zip = next(iter(tep_da_tai.items()))
    print(f"Đang sử dụng file đã chọn: {{ten_thuc_te}}")
else:
    raise FileNotFoundError(f"Không tìm thấy {{TEN_FILE_SOURCE}} trong các file đã chọn.")

if hashlib.sha256(du_lieu_zip).hexdigest() != MA_BAM_SOURCE:
    raise ValueError(
        "File ZIP không đúng phiên bản của bộ notebook này. "
        "Hãy chọn Genetic_ALO_Project.zip nằm cùng bộ notebook."
    )

GOI_SOURCE.write_bytes(du_lieu_zip)
(THU_MUC_COLAB_DRIVE / "project-source.sha256").write_text(
    MA_BAM_SOURCE + "\\n", encoding="utf-8"
)

THU_MUC_DU_AN = Path("/content/genetic-alo-main")
shutil.rmtree(THU_MUC_DU_AN, ignore_errors=True)
THU_MUC_DU_AN.mkdir(parents=True)
with zipfile.ZipFile(io.BytesIO(du_lieu_zip), "r") as archive:
    archive.extractall(THU_MUC_DU_AN)

so_tep = sum(1 for path in THU_MUC_DU_AN.rglob("*") if path.is_file())
if so_tep != SO_TEP_DU_KIEN:
    raise RuntimeError(f"Dự kiến {{SO_TEP_DU_KIEN}} file nhưng giải nén được {{so_tep}} file.")

# Lưu các output mẫu để notebook demo có thể mở ngay.
THU_MUC_KET_QUA_DRIVE = THU_MUC_COLAB_DRIVE / "latest_outputs"
THU_MUC_KET_QUA_DRIVE.mkdir(parents=True, exist_ok=True)
shutil.copytree(
    THU_MUC_DU_AN / "outputs",
    THU_MUC_KET_QUA_DRIVE,
    dirs_exist_ok=True,
)

subprocess.run(
    [
        sys.executable, "-m", "pip", "install", "--quiet",
        "--disable-pip-version-check", "-r",
        str(THU_MUC_DU_AN / "requirements.txt"),
    ],
    check=True,
)
os.chdir(THU_MUC_DU_AN)
if str(THU_MUC_DU_AN) not in sys.path:
    sys.path.insert(0, str(THU_MUC_DU_AN))

from constraints import ConstraintEvaluator
from dataset import ExcelDatasetLoader
from ga import GeneticAlgorithmEngine

print(f"✅ Đã tích hợp và kiểm tra {{so_tep}} file source/data/output.")
print(f"✅ Gói source đã lưu tại: {{GOI_SOURCE}}")
print(f"✅ Thư mục chạy hiện tại: {{THU_MUC_DU_AN}}")
print("✅ Dataset Excel thật nằm trong data/instances và có thể mở từ thanh Files.")
print("➡️ Tiếp theo, mở notebook 01_Kiem_Thu_Va_Du_Lieu.ipynb.")
'''
    # The current Colab workflow clones the public repository directly.
    setup_code = bootstrap_source(checksum)
    write_notebook(
        "00_Thiet_Lap_Du_An_Colab.ipynb",
        [
            markdown(
                """
# 00 — Thiết lập toàn bộ dự án Genetic ALO trên Google Colab

Notebook này clone toàn bộ source code và dataset Excel từ repository GitHub
vào máy ảo Colab, cài dependencies và hiển thị commit đang chạy.

> Chạy **Runtime → Run all** và cấp quyền Google Drive để lưu kết quả.
"""
            ),
            code(setup_code),
            markdown(
                """
## Hoàn tất

Sau khi thấy ba dấu ✅, mở notebook `01_Kiem_Thu_Va_Du_Lieu.ipynb`.
Mỗi notebook còn lại cũng tự clone lại nhánh `main`, nên có thể chạy độc lập
trong một phiên Colab mới.
"""
            ),
        ],
    )


def build_test_notebook(checksum: str) -> None:
    write_notebook(
        "01_Kiem_Thu_Va_Du_Lieu.ipynb",
        [
            markdown(
                """
# 01 — Kiểm thử source code và dữ liệu

Notebook này khôi phục dự án từ Google Drive, xác minh hai dataset Excel, import
các module chính và chạy toàn bộ pytest. Chạy **Runtime → Run all**.
"""
            ),
            code(bootstrap_source(checksum)),
            code(
                r'''
#@title Kiểm tra cấu trúc và tính hợp lệ của dataset
from dataset import DatasetValidator, ExcelDatasetLoader

cac_tep_bat_buoc = [
    THU_MUC_DU_AN / "main.py",
    THU_MUC_DU_AN / "main_benchmark.py",
    THU_MUC_DU_AN / "ui_app.py",
    THU_MUC_DU_AN / "data/instances/instance_easy.xlsx",
    THU_MUC_DU_AN / "data/instances/instance_medium.xlsx",
]
tep_thieu = [str(path) for path in cac_tep_bat_buoc if not path.is_file()]
if tep_thieu:
    raise FileNotFoundError("Thiếu file bắt buộc:\n- " + "\n- ".join(tep_thieu))

for ten in ("easy", "medium"):
    path = THU_MUC_DU_AN / f"data/instances/instance_{ten}.xlsx"
    dataset = ExcelDatasetLoader.load_and_validate(str(path))
    report = DatasetValidator.validate_report(dataset)
    if not report["valid"]:
        raise ValueError(f"Dataset {ten} không hợp lệ: {report['errors']}")
    print(f"✅ Dataset {ten.upper()} hợp lệ: {path.name}")
'''
            ),
            code(
                r'''
#@title Xem trực tiếp dữ liệu Excel
import pandas as pd
from IPython.display import display

XEM_DATASET = "easy" #@param ["easy", "medium"]
SO_DONG_MOI_SHEET = 5 #@param {type:"integer"}

duong_dan_excel = THU_MUC_DU_AN / f"data/instances/instance_{XEM_DATASET}.xlsx"
tep_excel = pd.ExcelFile(duong_dan_excel)
print(f"Dataset thật: {duong_dan_excel}")
print("Các sheet:", tep_excel.sheet_names)

for ten_sheet in tep_excel.sheet_names:
    print(f"\n--- {ten_sheet} ---")
    display(pd.read_excel(duong_dan_excel, sheet_name=ten_sheet).head(SO_DONG_MOI_SHEET))
'''
            ),
            code(
                r'''
#@title Chạy toàn bộ pytest
CHAY_TOAN_BO_TEST = True #@param {type:"boolean"}

if CHAY_TOAN_BO_TEST:
    subprocess.run(
        [sys.executable, "-m", "pytest", "-q"],
        cwd=THU_MUC_DU_AN,
        check=True,
    )
    print("✅ Toàn bộ kiểm thử đã pass.")
else:
    print("ℹ️ Đã bỏ qua pytest.")
'''
            ),
        ],
    )


def build_production_notebook(checksum: str) -> None:
    write_notebook(
        "02_Chay_GA_Va_Xuat_Lich.ipynb",
        [
            markdown(
                """
# 02 — Chạy Genetic Algorithm và xuất thời khóa biểu

Notebook chạy luồng production **GA + Repair + Soft Local Search**, xuất Excel,
JSON và biểu đồ hội tụ. Kết quả được đồng bộ sang Google Drive để notebook demo
sử dụng đúng lịch vừa tạo.
"""
            ),
            code(bootstrap_source(checksum)),
            code(
                r'''
#@title Cấu hình và chạy thuật toán
TEN_DATASET = "easy" #@param ["easy", "medium"]
SEED = 42 #@param {type:"integer"}
KICH_THUOC_QUAN_THE = 60 #@param {type:"integer"}
NGAN_SACH_DANH_GIA = 1000 #@param {type:"integer"}
DUNG_SOFT_LOCAL_SEARCH = True #@param {type:"boolean"}
HO_SO_TRONG_SO = "balanced" #@param ["student_centric", "balanced", "resource_centric"]

duong_dan_input = THU_MUC_DU_AN / f"data/instances/instance_{TEN_DATASET}.xlsx"
thu_muc_ket_qua = THU_MUC_DU_AN / "outputs/production"
thu_muc_ket_qua.mkdir(parents=True, exist_ok=True)
duong_dan_output = thu_muc_ket_qua / "best_timetable.xlsx"

lenh_chay = [
    sys.executable,
    "main.py",
    "--input", str(duong_dan_input),
    "--output", str(duong_dan_output),
    "--seed", str(SEED),
    "--population-size", str(KICH_THUOC_QUAN_THE),
    "--search-evaluation-budget", str(NGAN_SACH_DANH_GIA),
    "--weight-profile", HO_SO_TRONG_SO,
]
if DUNG_SOFT_LOCAL_SEARCH:
    lenh_chay.append("--soft-local-search")

print("Lệnh thực thi:", " ".join(lenh_chay))
subprocess.run(lenh_chay, cwd=THU_MUC_DU_AN, check=True)
vi_tri_drive = dong_bo_ket_qua()
print(f"✅ File thời khóa biểu: {duong_dan_output}")
print(f"✅ Đã đồng bộ output sang: {vi_tri_drive}")
'''
            ),
            code(
                r'''
#@title Xem trước Excel và biểu đồ hội tụ
import pandas as pd
from IPython.display import Image, display
from openpyxl import load_workbook

if not duong_dan_output.is_file():
    raise FileNotFoundError("Chưa có file kết quả. Hãy chạy cell thuật toán trước.")

workbook = load_workbook(duong_dan_output, read_only=True, data_only=True)
print("Các sheet:", workbook.sheetnames)
ten_sheet = workbook.sheetnames[0]
workbook.close()

display(pd.read_excel(duong_dan_output, sheet_name=ten_sheet).head(20))

ma_phuong_phap = "ga_repair_sls" if DUNG_SOFT_LOCAL_SEARCH else "ga_repair"
duong_dan_bieu_do = thu_muc_ket_qua / f"convergence_{ma_phuong_phap}.png"
if duong_dan_bieu_do.is_file():
    display(Image(filename=str(duong_dan_bieu_do)))
'''
            ),
            code(
                r'''
#@title Tải toàn bộ kết quả về máy (tùy chọn)
TAI_KET_QUA = False #@param {type:"boolean"}

if TAI_KET_QUA:
    from google.colab import files
    tep_zip = shutil.make_archive(
        "/content/ket_qua_genetic_alo",
        "zip",
        root_dir=THU_MUC_DU_AN / "outputs",
    )
    files.download(tep_zip)
else:
    print("ℹ️ Kết quả đã nằm trên Google Drive; bật TAI_KET_QUA nếu muốn tải ZIP.")
'''
            ),
        ],
    )


def build_sensitivity_notebook(checksum: str) -> None:
    write_notebook(
        "03_Phan_Tich_Do_Nhay.ipynb",
        [
            markdown(
                """
# 03 — Phân tích độ nhạy trọng số

Notebook chạy thí nghiệm Phase 1.1 trên ba cấu hình trọng số và ba seed. Kết quả
gồm `raw_runs.csv` và `summary.json`, sau đó được lưu sang Google Drive.
"""
            ),
            code(bootstrap_source(checksum)),
            code(
                r'''
#@title Chạy phân tích độ nhạy
CHAY_PHAN_TICH_DO_NHAY = True #@param {type:"boolean"}
TEN_DATASET = "easy" #@param ["easy", "medium"]

thu_muc_do_nhay = THU_MUC_DU_AN / "outputs/benchmark/phase1_1_weight_sensitivity"
if CHAY_PHAN_TICH_DO_NHAY:
    subprocess.run(
        [
            sys.executable,
            "scripts/run_phase1_1_sensitivity.py",
            "--input", str(THU_MUC_DU_AN / f"data/instances/instance_{TEN_DATASET}.xlsx"),
            "--output-dir", str(thu_muc_do_nhay),
        ],
        cwd=THU_MUC_DU_AN,
        check=True,
    )
    print(f"✅ Đã đồng bộ sang: {dong_bo_ket_qua()}")
else:
    print("ℹ️ Đã bỏ qua phân tích độ nhạy.")
'''
            ),
            code(
                r'''
#@title Hiển thị bảng tổng hợp
import json
import pandas as pd
from IPython.display import display

summary_path = thu_muc_do_nhay / "summary.json"
if not summary_path.is_file():
    raise FileNotFoundError("Chưa có summary.json. Hãy chạy cell phân tích trước.")

payload = json.loads(summary_path.read_text(encoding="utf-8"))
rows = []
for profile, values in payload["summary"].items():
    rows.append({
        "profile": profile,
        "runs": values["runs"],
        "feasible_runs": values["feasible_runs"],
        "mean_total_soft_score": values["mean_total_soft_score"],
        "mean_runtime_seconds": values["mean_runtime_seconds"],
    })
display(pd.DataFrame(rows))
'''
            ),
        ],
    )


def build_benchmark_notebook(checksum: str) -> None:
    write_notebook(
        "04_Benchmark_So_Sanh.ipynb",
        [
            markdown(
                """
# 04 — Benchmark so sánh phương pháp

Phần đầu chạy benchmark nhanh phục vụ trình bày. Phần cuối là benchmark chính
thức gồm 60 lượt chạy; mặc định tắt để tránh vô tình tốn nhiều thời gian.
Checkpoint được đồng bộ sang Drive nên có thể tiếp tục sau khi Colab ngắt phiên.
"""
            ),
            code(bootstrap_source(checksum)),
            code(
                r'''
#@title Benchmark nhanh
CHAY_BENCHMARK_NHANH = True #@param {type:"boolean"}
CAC_PHUONG_PHAP = "ga_repair_sls,ga_repair,ga" #@param {type:"string"}
CAC_SEED = "0-2" #@param {type:"string"}
TEN_DATASET = "easy" #@param ["easy", "medium"]

if CHAY_BENCHMARK_NHANH:
    subprocess.run(
        [
            sys.executable, "main_benchmark.py",
            "--mode", "fast",
            "--methods", CAC_PHUONG_PHAP,
            "--seeds", CAC_SEED,
            "--data-source", "excel",
            "--input", str(THU_MUC_DU_AN / f"data/instances/instance_{TEN_DATASET}.xlsx"),
            "--experiment-name", "colab_fast",
        ],
        cwd=THU_MUC_DU_AN,
        check=True,
    )
    print(f"✅ Đã đồng bộ sang: {dong_bo_ket_qua()}")
else:
    print("ℹ️ Đã bỏ qua benchmark nhanh.")
'''
            ),
            code(
                r'''
#@title Xem kết quả benchmark nhanh
import pandas as pd
from IPython.display import display

thu_muc_benchmark_nhanh = THU_MUC_DU_AN / "outputs/benchmark/colab_fast"
tep_summary = thu_muc_benchmark_nhanh / "summary.csv"
if tep_summary.is_file():
    display(pd.read_csv(tep_summary))
else:
    print("Kết quả được tạo trong:", thu_muc_benchmark_nhanh)
    print([str(path.relative_to(THU_MUC_DU_AN)) for path in thu_muc_benchmark_nhanh.glob("*")])
'''
            ),
            code(
                r'''
#@title Benchmark chính thức 60 lượt chạy (tùy chọn)
CHAY_BENCHMARK_CUOI = False #@param {type:"boolean"}
CHAY_LAI_TU_DAU = False #@param {type:"boolean"}

if CHAY_BENCHMARK_CUOI:
    lenh = [sys.executable, "scripts/run_final_benchmark.py"]
    if CHAY_LAI_TU_DAU:
        lenh.append("--fresh")
    subprocess.run(lenh, cwd=THU_MUC_DU_AN, check=True)
    print(f"✅ Benchmark hoàn tất; đã đồng bộ sang: {dong_bo_ket_qua()}")
else:
    print("ℹ️ Benchmark 60 lượt đang tắt. Chỉ bật khi có đủ thời gian chạy.")
'''
            ),
        ],
    )


def build_demo_notebook(checksum: str) -> None:
    write_notebook(
        "05_Streamlit_Demo.ipynb",
        [
            markdown(
                """
# 05 — Mở giao diện Streamlit giống web demo

Notebook khôi phục đúng `ui_app.py` và output mới nhất, khởi động Streamlit trong
máy ảo Colab rồi tạo một đường dẫn TryCloudflare tạm thời. Không cần tài khoản
Cloudflare hoặc token. Hãy giữ phiên Colab hoạt động trong lúc trình bày.

> Đường dẫn là công khai và tạm thời; chỉ chia sẻ trong buổi demo.
"""
            ),
            code(bootstrap_source(checksum)),
            code(
                r'''
#@title Kiểm tra dữ liệu dùng cho demo
cac_tep_demo = [
    THU_MUC_DU_AN / "ui_app.py",
    THU_MUC_DU_AN / "data/instances/instance_easy.xlsx",
    THU_MUC_DU_AN / "outputs/production/schedule_query_data.json",
    THU_MUC_DU_AN / "outputs/production/best_timetable_metadata.json",
]
tep_thieu = [str(path) for path in cac_tep_demo if not path.is_file()]
if tep_thieu:
    raise FileNotFoundError(
        "Thiếu dữ liệu demo. Hãy chạy notebook 02 trước:\n- " + "\n- ".join(tep_thieu)
    )
print("✅ Dữ liệu demo đã sẵn sàng.")
'''
            ),
            code(
                r'''
#@title Khởi động Streamlit và tạo đường dẫn demo
MO_GIAO_DIEN_STREAMLIT = True #@param {type:"boolean"}

import re
import time
import urllib.request
from IPython.display import HTML, display

if not MO_GIAO_DIEN_STREAMLIT:
    print("ℹ️ Đã bỏ qua việc mở giao diện.")
else:
    # Dừng đúng các tiến trình do notebook này tạo nếu chạy lại cell.
    for ten_bien in ("tien_trinh_streamlit", "tien_trinh_tunnel"):
        tien_trinh_cu = globals().get(ten_bien)
        if tien_trinh_cu is not None and tien_trinh_cu.poll() is None:
            tien_trinh_cu.terminate()

    tep_log_streamlit = Path("/content/streamlit_colab.log")
    tep_log_tunnel = Path("/content/cloudflare_tunnel.log")
    log_streamlit = open(tep_log_streamlit, "w", encoding="utf-8")

    tien_trinh_streamlit = subprocess.Popen(
        [
            sys.executable, "-m", "streamlit", "run", "ui_app.py",
            "--server.headless=true",
            "--server.address=0.0.0.0",
            "--server.port=8501",
            "--browser.gatherUsageStats=false",
        ],
        cwd=THU_MUC_DU_AN,
        env={**os.environ, "GA_DEMO_EVALUATION_BUDGET": "100"},
        stdout=log_streamlit,
        stderr=subprocess.STDOUT,
        text=True,
    )

    # Chờ endpoint health thay vì ngủ một khoảng cố định.
    for _ in range(60):
        if tien_trinh_streamlit.poll() is not None:
            break
        try:
            with urllib.request.urlopen(
                "http://127.0.0.1:8501/_stcore/health", timeout=2
            ) as response:
                if response.status == 200 and response.read().decode().strip() == "ok":
                    break
        except Exception:
            time.sleep(1)
    else:
        raise TimeoutError("Streamlit không phản hồi health check sau 60 giây.")

    if tien_trinh_streamlit.poll() is not None:
        log_streamlit.flush()
        raise RuntimeError(tep_log_streamlit.read_text(errors="replace"))

    url_cong_khai = None
    for lan_thu in range(1, 3):
        log_tunnel = open(tep_log_tunnel, "w", encoding="utf-8")
        tien_trinh_tunnel = subprocess.Popen(
            [
                "npx", "--yes", "wrangler@latest", "tunnel", "quick-start",
                "http://127.0.0.1:8501",
            ],
            stdout=log_tunnel,
            stderr=subprocess.STDOUT,
            text=True,
        )
        for _ in range(90):
            time.sleep(1)
            log_tunnel.flush()
            noi_dung = tep_log_tunnel.read_text(encoding="utf-8", errors="replace")
            ket_qua = re.search(r"https://[a-zA-Z0-9-]+\.trycloudflare\.com", noi_dung)
            if ket_qua:
                url_cong_khai = ket_qua.group(0)
                break
            if tien_trinh_tunnel.poll() is not None:
                break
        if url_cong_khai:
            break
        if tien_trinh_tunnel.poll() is None:
            tien_trinh_tunnel.terminate()
        print(f"Thử tạo tunnel lần {lan_thu} chưa thành công; đang thử lại...")

    if not url_cong_khai:
        print(tep_log_tunnel.read_text(encoding="utf-8", errors="replace"))
        raise RuntimeError("Không tạo được link demo sau 2 lần thử.")

    print("✅ Streamlit health check: OK")
    print("✅ Link demo:", url_cong_khai)
    display(HTML(f'<a href="{url_cong_khai}" target="_blank" '
                 'style="font-size:20px;font-weight:bold">MỞ WEB DEMO</a>'))
'''
            ),
            markdown(
                """
## Phương án dự phòng khi mạng tunnel không ổn định

Notebook 02 vẫn chạy thuật toán, hiển thị bảng thời khóa biểu và biểu đồ ngay
trong Colab. Khi thi, hãy mở sẵn cả notebook 02 và 05; nếu link tạm thời gặp lỗi,
trình bày kết quả trực tiếp từ notebook 02.
"""
            ),
        ],
    )


def write_readme(checksum: str, file_count: int) -> None:
    del checksum
    content = f"""# Bộ Google Colab Genetic ALO

## Thứ tự chạy

1. `00_Thiet_Lap_Du_An_Colab.ipynb` — clone source/data từ GitHub và kiểm tra môi trường.
2. `01_Kiem_Thu_Va_Du_Lieu.ipynb` — hiển thị Excel, kiểm tra dataset và chạy pytest.
3. `02_Chay_GA_Va_Xuat_Lich.ipynb` — chạy GA + Repair + SLS, xuất lịch.
4. `03_Phan_Tich_Do_Nhay.ipynb` — thí nghiệm độ nhạy trọng số.
5. `04_Benchmark_So_Sanh.ipynb` — benchmark nhanh hoặc benchmark 60 lượt.
6. `05_Streamlit_Demo.ipynb` — mở đúng giao diện `ui_app.py` qua link tạm thời.

Repository hiện có khoảng **{file_count} file source, dataset Excel và output**.
Mỗi notebook tự clone nhánh `main` vào `/content/genetic-alo`, sau đó khôi phục
output mới nhất từ Google Drive nếu có. Dataset thật nằm trong `data/instances`.

## Cách nộp/chạy

- Giải nén bộ bàn giao và upload sáu file `.ipynb` lên Google Drive hoặc Colab.
- Mở và chạy theo thứ tự 00 → 05.
- Dùng runtime CPU; không cần GPU.
- Chạy notebook 02 trước notebook 05 để giao diện dùng output mới nhất.
- Quick Tunnel của notebook 05 chỉ dùng cho buổi demo; giữ tab Colab hoạt động.
"""
    (OUTPUT_DIR / "README.md").write_text(content, encoding="utf-8")


def build_archive() -> Path:
    # Build outside OUTPUT_DIR so the archive can never include itself.
    temporary_base = ROOT / "Genetic_ALO_Colab_Notebooks"
    temporary_path = Path(
        shutil.make_archive(str(temporary_base), "zip", ROOT, "colab")
    )
    archive_path = OUTPUT_DIR / temporary_path.name
    shutil.move(str(temporary_path), archive_path)
    return archive_path


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for old_notebook in OUTPUT_DIR.glob("*.ipynb"):
        old_notebook.unlink()
    old_archive = OUTPUT_DIR / "Genetic_ALO_Colab_Notebooks.zip"
    if old_archive.exists():
        old_archive.unlink()
    project_archive = OUTPUT_DIR / "Genetic_ALO_Project.zip"
    if project_archive.exists():
        project_archive.unlink()

    files = iter_payload_files()
    checksum = "github-main"
    file_count = len(files)
    build_setup_notebook(checksum, file_count)
    build_test_notebook(checksum)
    build_production_notebook(checksum)
    build_sensitivity_notebook(checksum)
    build_benchmark_notebook(checksum)
    build_demo_notebook(checksum)
    write_readme(checksum, file_count)
    archive_path = build_archive()

    print(f"Built 6 notebooks in: {OUTPUT_DIR}")
    print(f"Project files represented by the GitHub clone: {file_count}")
    print(f"Upload bundle: {archive_path}")


if __name__ == "__main__":
    main()
