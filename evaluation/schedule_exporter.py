import csv
import json
import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Union, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from domain import Schedule, expand_scheduling_activities
from constraints import ConstraintEvaluator, SoftConstraintConfig

CSV_HEADERS = [
    "activity_id",
    "section_id",
    "meeting_index",
    "meeting_count",
    "meeting",
    "class_code",
    "course_id",
    "course_code",
    "lecturer_id",
    "student_group_id",
    "room_id",
    "day",
    "start_period",
    "end_period",
    "start_time",
    "end_time",
    "duration_periods",
    "session",
    "room_type",
]

def export_schedule_to_csv(
    schedule: Schedule,
    dataset: dict,
    output_path: Union[str, Path],
    metadata: Optional[dict] = None,
    soft_config: Optional[SoftConstraintConfig] = None,
) -> str:
    """Kiểm tra tính hợp lệ và xuất lịch học tốt nhất ra file CSV được sắp xếp theo Ngày, Tiết bắt đầu, Mã phòng."""

    if schedule is None or not hasattr(schedule, "genes") or not isinstance(schedule.genes, list) or len(schedule.genes) == 0:
        raise ValueError("Cannot export empty or invalid schedule.")

    if not isinstance(dataset, dict) or "course_sections" not in dataset:
        raise ValueError("Invalid dataset supplied to exporter.")

    activities = expand_scheduling_activities(dataset["course_sections"])
    section_map = {activity.activity_id: activity for activity in activities}
    course_map = {c.course_id: c for c in dataset.get("courses", [])}
    room_map = {r.id: r for r in dataset["rooms"]}
    timeslot_map = {t.id: t for t in dataset["timeslots"]}

    genes = schedule.genes
    if len(genes) != len(section_map):
        raise ValueError(f"Gene count mismatch: expected {len(section_map)} activities, got {len(genes)} genes.")

    seen_section_ids = set()
    for gene in genes:
        sec_id = getattr(gene, "section_id", None)
        room_id = getattr(gene, "room_id", None)
        ts_id = getattr(gene, "timeslot_id", None)

        if sec_id not in section_map:
            raise ValueError(f"Invalid section ID in gene: {sec_id}")
        if sec_id in seen_section_ids:
            raise ValueError(f"Duplicate section ID found in schedule: {sec_id}")
        seen_section_ids.add(sec_id)

        if room_id not in room_map:
            raise ValueError(f"Invalid room ID in gene: {room_id}")
        if ts_id not in timeslot_map:
            raise ValueError(f"Invalid timeslot ID in gene: {ts_id}")

    if seen_section_ids != set(section_map.keys()):
        raise ValueError("Schedule is missing some course sections from dataset.")

    # Đánh giá lại bằng ConstraintEvaluator
    evaluator = ConstraintEvaluator(dataset, soft_config=soft_config)
    _, hard_violations, _ = evaluator.calculate_fitness(schedule)
    if hard_violations > 0:
        raise ValueError(f"Cannot export schedule with hard violations (hard_violations={hard_violations}).")

    # Ánh xạ thứ tự ngày dựa trên các khung giờ trong bộ dữ liệu

    day_order: Dict[str, int] = {}
    for ts in dataset["timeslots"]:
        if ts.day not in day_order:
            day_order[ts.day] = len(day_order)

    day_period_to_ts: Dict[Tuple[str, int], Any] = {
        (ts.day, ts.period): ts for ts in dataset["timeslots"]
    }

    rows: List[dict] = []
    for gene in genes:
        sec = section_map[gene.section_id]
        room = room_map[gene.room_id]
        start_ts = timeslot_map[gene.timeslot_id]

        duration = getattr(sec, "duration_periods", 1)
        start_period = start_ts.period
        end_period = start_period + duration - 1

        start_time = getattr(start_ts, "start_time", "")
        end_ts = day_period_to_ts.get((start_ts.day, end_period), start_ts)
        end_time = getattr(end_ts, "end_time", getattr(start_ts, "end_time", ""))

        session = getattr(start_ts, "session", "morning")
        room_type = getattr(room, "room_type", "NORMAL")

        row = {
            "activity_id": sec.activity_id,
            "section_id": sec.section_id,
            "meeting_index": sec.meeting_index,
            "meeting_count": sec.meeting_count,
            "meeting": f"{sec.meeting_index}/{sec.meeting_count}",
            "class_code": getattr(sec, "class_code", None) or "",
            "course_id": getattr(sec, "course_id", ""),
            "course_code": getattr(course_map.get(sec.course_id), "course_code", None) or "",
            "lecturer_id": getattr(sec, "lecturer_id", ""),
            "student_group_id": getattr(sec, "group_id", ""),
            "room_id": room.id,
            "day": start_ts.day,
            "start_period": start_period,
            "end_period": end_period,
            "start_time": start_time,
            "end_time": end_time,
            "duration_periods": duration,
            "session": session,
            "room_type": room_type,
        }
        rows.append(row)

    rows.sort(
        key=lambda r: (
            day_order.get(r["day"], 999),
            r["start_period"],
            str(r["room_id"]),
        )
    )

    out_file = Path(output_path)
    out_file.parent.mkdir(parents=True, exist_ok=True)

    with open(out_file, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    return str(out_file)

def export_schedule_to_excel(
    schedule: Schedule,
    dataset: dict,
    output_path: Union[str, Path],
    metadata: Optional[dict] = None,
    allow_infeasible_export: bool = False,
    soft_config: Optional[SoftConstraintConfig] = None,
) -> str:
    """Xuất thời khóa biểu ra file Excel 7 sheet chuẩn hóa.

    Các sheet:
    1. SUMMARY
    2. RAW_ASSIGNMENTS
    3. SCHEDULE_BY_GROUP
    4. SCHEDULE_BY_LECTURER
    5. SCHEDULE_BY_ROOM
    6. VIOLATIONS
    7. RUN_CONFIG
    """
    out_path = Path(output_path)
    
    if schedule is None or not hasattr(schedule, "genes") or not isinstance(schedule.genes, list) or len(schedule.genes) == 0:
        raise ValueError("Cannot export empty or invalid schedule to Excel.")

    if not isinstance(dataset, dict) or "course_sections" not in dataset or "rooms" not in dataset or "timeslots" not in dataset:
        raise ValueError("Invalid dataset supplied to Excel exporter.")

    # Đánh giá lại các vi phạm bằng bộ đánh giá hợp nhất
    evaluator = ConstraintEvaluator(dataset, soft_config=soft_config)
    unified = evaluator.evaluate_unified(schedule)
    hard_violations = unified.hard_violations
    soft_penalty = unified.soft_penalty
    hard_details = unified.hard_details
    soft_breakdown = unified.soft_breakdown
    instance_violations = unified.instance_violations
    score, _, _ = evaluator.calculate_fitness(schedule)

    if hard_violations > 0:
        if not allow_infeasible_export:
            if out_path.exists():
                out_path.unlink()
            raise ValueError(f"Cannot export Excel schedule with hard violations (hard_violations={hard_violations}).")

        # Thêm hậu tố _INFEASIBLE nếu chưa có
        if "_INFEASIBLE" not in out_path.stem:
            out_path = out_path.parent / f"{out_path.stem}_INFEASIBLE.xlsx"

    meta = metadata or {}
    activities = expand_scheduling_activities(dataset["course_sections"])
    section_map = {activity.activity_id: activity for activity in activities}
    room_map = {r.id: r for r in dataset["rooms"]}
    timeslot_map = {t.id: t for t in dataset["timeslots"]}
    lecturer_map = {l.id: l for l in dataset.get("lecturers", [])}
    group_map = {g.id: g for g in dataset.get("student_groups", [])}
    course_map = {c.course_id: c for c in dataset.get("courses", [])}

    day_order: Dict[str, int] = {}
    for ts in dataset["timeslots"]:
        if ts.day not in day_order:
            day_order[ts.day] = len(day_order)

    day_period_to_ts: Dict[Tuple[str, int], Any] = {
        (ts.day, ts.period): ts for ts in dataset["timeslots"]
    }

    wb = openpyxl.Workbook()
    # Xóa trang tính mặc định
    wb.remove(wb.active)

    bold_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    def format_sheet(ws):
        ws.freeze_panes = "A2"
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = header_fill

    # --- 1. TRANG TÍNH SUMMARY ---
    ws_sum = wb.create_sheet(title="SUMMARY")
    ws_sum.append(["Metric", "Value"])
    rep_enabled = meta.get("repair_enabled", meta.get("use_repair", True))
    rep_calls = meta.get("repair_calls", 0)

    summary_rows = [
        ("dataset_source", meta.get("dataset_source", meta.get("dataset_preset", "Excel"))),
        ("dataset_path", meta.get("dataset_path", meta.get("input_file", "data/instances/instance_easy.xlsx"))),
        ("dataset_name", meta.get("dataset_name", "instance_easy.xlsx")),
        ("dataset_version", meta.get("dataset_version", "1.0")),
        ("dataset_hash", meta.get("dataset_hash", "N/A")),
        ("input_file", meta.get("input_file", "data/instances/instance_easy.xlsx")),
        ("normalized_json_path", meta.get("normalized_json_path", "outputs/datasets/instance_easy.normalized.json")),
        ("algorithm", meta.get("method", "GA + Repair")),
        ("seed", meta.get("seed", 0)),
        ("population_size", meta.get("pop_size", 60)),
        ("generations", meta.get("generations", meta.get("generations_run", 80))),
        ("evaluation_budget", meta.get("evaluation_budget", 1000)),
        ("hard_violations", hard_violations),
        ("soft_penalty", soft_penalty),
        ("score", score),
        ("total_score", score),
        ("feasible", hard_violations == 0),
        ("is_feasible", hard_violations == 0),
        ("runtime", meta.get("runtime_seconds", 0.0)),
        ("runtime_seconds", meta.get("runtime_seconds", 0.0)),
        ("fitness_evaluations", meta.get("fitness_evaluations", 0)),
        ("generation_to_first_feasible", meta.get("generation_to_first_feasible", 0 if hard_violations == 0 else "N/A")),
        ("time_to_first_feasible", meta.get("time_to_first_feasible", 0.0 if hard_violations == 0 else "N/A")),
        ("repair_enabled", rep_enabled),
        ("repair_trigger_policy", meta.get("repair_trigger_policy", "Offspring Mutation Constraint Satisfaction")),
        ("repair_calls", rep_calls),
        ("repair_attempts", meta.get("repair_attempts", rep_calls)),
        ("repair_successes", meta.get("repair_successes", 0)),
        ("repair_failures", meta.get("repair_failures", 0)),
        ("sections_repaired", meta.get("sections_repaired", 0)),
        ("sections_failed", meta.get("sections_failed", 0)),
        ("candidate_checks", meta.get("candidate_checks", 0)),
        ("hard_before_repair", meta.get("hard_before_repair", 0)),
        ("hard_after_repair", meta.get("hard_after_repair", 0)),
        ("soft_before_repair", meta.get("soft_before_repair", 0)),
        ("soft_after_repair", meta.get("soft_after_repair", 0)),
        ("repair_runtime_seconds", meta.get("repair_runtime_seconds", 0.0)),
    ]
    if rep_enabled and rep_calls == 0:
        summary_rows.append(("repair_note", "Repair was enabled but not triggered because no repairable violations were detected."))

    summary_rows.append(("generated_at", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    for k, v in summary_rows:
        ws_sum.append([k, str(v)])
    format_sheet(ws_sum)

    # Chuẩn bị các hàng gene
    assignments = []
    for gene in schedule.genes:
        sec = section_map[gene.section_id]
        room = room_map[gene.room_id]
        start_ts = timeslot_map[gene.timeslot_id]
        dur = getattr(sec, "duration_periods", 1)
        start_p = start_ts.period
        end_p = start_p + dur - 1

        end_ts = day_period_to_ts.get((start_ts.day, end_p), start_ts)
        start_time = getattr(start_ts, "start_time", "07:00")
        end_time = getattr(end_ts, "end_time", "07:50")

        periods_str = f"Tiết {start_p}" if dur == 1 else f"Tiết {start_p}-{end_p}"
        time_range = f"{start_time}-{end_time}"

        lec = lecturer_map.get(sec.lecturer_id)
        grp = group_map.get(sec.group_id)
        crs = course_map.get(sec.course_id)

        assignments.append({
            "activity_id": sec.activity_id,
            "section_id": sec.section_id,
            "meeting_index": sec.meeting_index,
            "meeting_count": sec.meeting_count,
            "meeting": f"{sec.meeting_index}/{sec.meeting_count}",
            "class_code": getattr(sec, "class_code", None) or "",
            "course_id": sec.course_id,
            "course_code": getattr(crs, "course_code", None) or "",
            "course_name": getattr(sec, "course_name", getattr(crs, "name", sec.course_id)),
            "student_group_id": sec.group_id,
            "student_group_name": getattr(grp, "name", sec.group_id),
            "lecturer_id": sec.lecturer_id,
            "lecturer_name": getattr(lec, "name", sec.lecturer_id),
            "student_count": sec.student_count,
            "room_id": room.id,
            "room_number": getattr(room, "name", room.id),
            "building": getattr(room, "name", "").split("-")[0] if "-" in getattr(room, "name", "") else "",
            "campus_id": getattr(room, "campus_id", None) or "",
            "room_type": getattr(room, "room_type", "NORMAL"),
            "day_no": day_order.get(start_ts.day, 99),
            "day_name": start_ts.day,
            "start_period": start_p,
            "end_period": end_p,
            "duration_periods": dur,
            "periods": periods_str,
            "start_time": start_time,
            "end_time": end_time,
            "shift": getattr(start_ts, "session", "morning"),
            "required_room_type": getattr(sec, "required_room_type", "NORMAL"),
            "preferred_campus_id": getattr(sec, "preferred_campus_id", None) or "",
            "preferred_shift": getattr(sec, "preferred_shift", None) or "",
            "time_range": time_range,
            "room_display": getattr(room, "name", room.id),
        })

    # --- 2. TRANG TÍNH RAW_ASSIGNMENTS ---
    ws_raw = wb.create_sheet(title="RAW_ASSIGNMENTS")
    raw_cols = [
        "activity_id", "section_id", "meeting_index", "meeting_count", "meeting",
        "class_code", "course_id", "course_code", "course_name",
        "student_group_id", "student_group_name", "lecturer_id", "lecturer_name",
        "student_count", "room_id", "room_number", "building", "campus_id",
        "room_type", "day_no", "day_name", "start_period", "end_period",
        "duration_periods", "periods", "start_time", "end_time", "shift",
        "required_room_type", "preferred_campus_id", "preferred_shift"
    ]
    ws_raw.append(raw_cols)

    # Sắp xếp các phân công thô theo section_id
    raw_sorted = sorted(assignments, key=lambda a: (a["section_id"], a["meeting_index"]))
    for a in raw_sorted:
        ws_raw.append([a[c] for c in raw_cols])
    format_sheet(ws_raw)

    # --- 3. TRANG TÍNH SCHEDULE_BY_GROUP ---
    ws_grp = wb.create_sheet(title="SCHEDULE_BY_GROUP")
    grp_cols = ["group_id", "group_name", "day_name", "periods", "time_range", "class_code", "meeting", "course_code", "course_name", "lecturer_name", "room_display", "campus_id"]
    ws_grp.append(grp_cols)
    grp_sorted = sorted(assignments, key=lambda a: (a["student_group_id"], a["day_no"], a["start_period"]))
    for a in grp_sorted:
        ws_grp.append([a["student_group_id"], a["student_group_name"], a["day_name"], a["periods"], a["time_range"], a["class_code"], a["meeting"], a["course_code"], a["course_name"], a["lecturer_name"], a["room_display"], a["campus_id"]])
    format_sheet(ws_grp)

    # --- 4. TRANG TÍNH SCHEDULE_BY_LECTURER ---
    ws_lec = wb.create_sheet(title="SCHEDULE_BY_LECTURER")
    lec_cols = ["lecturer_id", "lecturer_name", "day_name", "periods", "time_range", "class_code", "meeting", "course_code", "course_name", "student_group_name", "room_display", "campus_id"]
    ws_lec.append(lec_cols)
    lec_sorted = sorted(assignments, key=lambda a: (a["lecturer_id"], a["day_no"], a["start_period"]))
    for a in lec_sorted:
        ws_lec.append([a["lecturer_id"], a["lecturer_name"], a["day_name"], a["periods"], a["time_range"], a["class_code"], a["meeting"], a["course_code"], a["course_name"], a["student_group_name"], a["room_display"], a["campus_id"]])
    format_sheet(ws_lec)

    # --- 5. TRANG TÍNH SCHEDULE_BY_ROOM ---
    ws_rm = wb.create_sheet(title="SCHEDULE_BY_ROOM")
    rm_cols = ["room_id", "room_display", "campus_id", "day_name", "periods", "time_range", "class_code", "meeting", "course_code", "course_name", "lecturer_name", "student_group_name"]
    ws_rm.append(rm_cols)
    rm_sorted = sorted(assignments, key=lambda a: (a["room_id"], a["day_no"], a["start_period"]))
    for a in rm_sorted:
        ws_rm.append([a["room_id"], a["room_display"], a["campus_id"], a["day_name"], a["periods"], a["time_range"], a["class_code"], a["meeting"], a["course_code"], a["course_name"], a["lecturer_name"], a["student_group_name"]])
    format_sheet(ws_rm)

    # --- 6. TRANG TÍNH VIOLATIONS ---
    ws_viol = wb.create_sheet(title="VIOLATIONS")
    viol_cols = [
        "violation_type",
        "severity",
        "constraint_name",
        "section_ids",
        "lecturer_id",
        "student_group_ids",
        "room_id",
        "day",
        "periods",
        "raw_count",
        "weight",
        "weighted_penalty",
        "denominator",
        "normalized_penalty",
        "description"
    ]
    ws_viol.append(viol_cols)

    # 6a. Vi phạm cứng
    if hard_violations == 0:
        ws_viol.append(["INFO", "NONE", "hard_constraints", "-", "-", "-", "-", "-", "-", 0, 1000, 0, "-", "-", "No hard violations detected"])
    else:
        for k, v in hard_details.items():
            if v > 0:
                ws_viol.append(["HARD", "HIGH", k, "-", "-", "-", "-", "-", "-", v, 1000, v * 1000, "-", "-", f"Hard constraint violation: {k} (count={v})"])

    # 6b. Vi phạm mềm (phân tích từng trường hợp)
    if len(instance_violations) == 0:
        ws_viol.append(["INFO", "NONE", "soft_constraints", "-", "-", "-", "-", "-", "-", 0, 0, 0, 0, 0, "No soft constraint violations detected"])
    else:
        for item in instance_violations:
            r_cnt = item.get("raw_count", 1)
            w = item.get("weight", 0)
            wp = item.get("weighted_penalty", r_cnt * w)
            ws_viol.append([
                item.get("violation_type", "SOFT"),
                item.get("severity", "LOW"),
                item.get("constraint_name", ""),
                item.get("section_ids", "-"),
                item.get("lecturer_id", "-"),
                item.get("student_group_ids", "-"),
                item.get("room_id", "-"),
                item.get("day", "-"),
                item.get("periods", "-"),
                r_cnt,
                w,
                wp,
                item.get("denominator", 0),
                item.get("normalized_penalty", 0),
                item.get("description", "")
            ])

    # 6c. Hàng tổng kết TOTAL SOFT PENALTY (bắt buộc bằng SUMMARY.soft_penalty)
    total_raw_soft_count = sum(item.raw_count for item in soft_breakdown)
    ws_viol.append([
        "SUMMARY",
        "TOTAL",
        "TOTAL_SOFT_PENALTY",
        "-",
        "-",
        "-",
        "-",
        "-",
        "-",
        total_raw_soft_count,
        "-",
        soft_penalty,
        "-",
        "-",
        f"TOTAL SOFT PENALTY (Sum of all soft constraint weighted penalties = {soft_penalty})"
    ])

    format_sheet(ws_viol)

    # --- 7. TRANG TÍNH RUN_CONFIG ---
    ws_cfg = wb.create_sheet(title="RUN_CONFIG")
    ws_cfg.append(["Parameter", "Value"])
    soft_cfg = evaluator.soft_checker.config
    cfg_rows = [
        ("primary_method", meta.get("primary_method", "ga_repair")),
        ("selected_methods", meta.get("selected_methods", "ga_repair")),
        ("search_evaluation_budget", meta.get("evaluation_budget", meta.get("search_evaluation_budget", 1000))),
        ("population_size", meta.get("pop_size", 60)),
        ("generations", meta.get("generations", 80)),
        ("crossover_rate", meta.get("crossover_rate", 0.8)),
        ("mutation_rate", meta.get("mutation_rate", 0.2)),
        ("hard_weight", meta.get("hard_weight", 1000)),
        ("soft_weight", meta.get("soft_weight", 1)),
        ("same_session_rule", True),
        ("effective_soft_constraints", json.dumps(soft_cfg.to_metadata(), ensure_ascii=False, sort_keys=True)),
        ("S1_compact_student_schedule_weight", soft_cfg.get_weight("compact_student_schedule")),
        ("S1_compact_student_schedule_enabled", soft_cfg.is_enabled("compact_student_schedule")),
        ("S2_late_day_periods_weight", soft_cfg.get_weight("late_day_periods")),
        ("S2_late_day_periods_enabled", soft_cfg.is_enabled("late_day_periods")),
        ("S3_preferred_shift_mismatch_weight", soft_cfg.get_weight("preferred_shift_mismatch")),
        ("S3_preferred_shift_mismatch_enabled", soft_cfg.is_enabled("preferred_shift_mismatch")),
        ("S4_room_seat_waste_weight", soft_cfg.get_weight("room_seat_waste")),
        ("S4_room_seat_waste_enabled", soft_cfg.is_enabled("room_seat_waste")),
        ("S5_consecutive_cross_campus_weight", soft_cfg.get_weight("consecutive_cross_campus")),
        ("S5_consecutive_cross_campus_enabled", soft_cfg.is_enabled("consecutive_cross_campus")),
        ("S6_preferred_campus_mismatch_weight", soft_cfg.get_weight("preferred_campus_mismatch")),
        ("S6_preferred_campus_mismatch_enabled", soft_cfg.is_enabled("preferred_campus_mismatch")),
        ("S7_student_home_campus_mismatch_weight", soft_cfg.get_weight("student_home_campus_mismatch")),
        ("S7_student_home_campus_mismatch_enabled", soft_cfg.is_enabled("student_home_campus_mismatch")),
        ("repair_enabled", meta.get("use_repair", True)),
    ]
    for k, v in cfg_rows:
        ws_cfg.append([k, str(v)])
    format_sheet(ws_cfg)


    # --- 8. TRANG TÍNH RUN_METRICS (nếu có chi tiết nhiều lần chạy) ---
    all_runs_flat = meta.get("all_runs_flat")
    if all_runs_flat:
        ws_m = wb.create_sheet(title="RUN_METRICS")
        headers = [
            "method", "seed", "feasible", "final_hard_violations", "final_soft_penalty",
            "runtime_seconds", "time_to_first_feasible_seconds", "search_fitness_evaluations",
            "search_hard_constraint_evaluations", "search_soft_constraint_evaluations", "search_constraint_evaluations",
            "internal_hard_constraint_evaluations", "internal_soft_constraint_evaluations", "internal_constraint_evaluations",
            "reporting_hard_constraint_evaluations", "reporting_soft_constraint_evaluations", "reporting_constraint_evaluations",
            "total_constraint_evaluations",
            "candidate_checks", "repair_calls", "repair_improved", "repair_unchanged", "repair_failed",
            "first_feasible_generation", "first_feasible_search_evaluation", "first_feasible_total_constraint_evaluation"
        ]
        ws_m.append(headers)
        for r in all_runs_flat:
            row = [
                r.get("method", ""),
                r.get("seed", ""),
                r.get("is_hard_feasible", r.get("hard_violations", 1) == 0),
                r.get("hard_violations", 0),
                r.get("soft_penalty", 0),
                round(r.get("runtime_seconds", 0.0), 4),
                round(r.get("time_to_first_feasible_seconds"), 4) if r.get("time_to_first_feasible_seconds") is not None and r.get("time_to_first_feasible_seconds") != "N/A" else "",
                r.get("search_fitness_evaluations", r.get("fitness_evaluations", 0)),
                r.get("search_hard_constraint_evaluations", 0),
                r.get("search_soft_constraint_evaluations", 0),
                r.get("search_constraint_evaluations", 0),
                r.get("internal_hard_constraint_evaluations", 0),
                r.get("internal_soft_constraint_evaluations", 0),
                r.get("internal_constraint_evaluations", 0),
                r.get("reporting_hard_constraint_evaluations", 0),
                r.get("reporting_soft_constraint_evaluations", 0),
                r.get("reporting_constraint_evaluations", 0),
                r.get("total_constraint_evaluations", 0),
                r.get("candidate_checks", 0),
                r.get("repair_calls", 0),
                r.get("repair_improved", 0),
                r.get("repair_unchanged", 0),
                r.get("repair_failed", 0),
                r.get("first_feasible_generation") if r.get("first_feasible_generation") is not None and r.get("first_feasible_generation") != "N/A" else "",
                r.get("first_feasible_search_evaluation") if r.get("first_feasible_search_evaluation") is not None and r.get("first_feasible_search_evaluation") != "N/A" else "",
                r.get("first_feasible_total_constraint_evaluation") if r.get("first_feasible_total_constraint_evaluation") is not None and r.get("first_feasible_total_constraint_evaluation") != "N/A" else "",
            ]
            ws_m.append(row)
        format_sheet(ws_m)

    # --- 9. TRANG TÍNH BENCHMARK_SUMMARY (nếu có phần tổng hợp) ---
    summary_list = meta.get("summary_list", meta.get("summary"))
    if summary_list:
        ws_s = wb.create_sheet(title="BENCHMARK_SUMMARY")
        sum_headers = [
            "method", "run_count", "feasible_rate", "median_final_hard", "median_final_soft",
            "mean_final_soft", "median_runtime_seconds", "median_time_to_first_feasible_seconds",
            "median_search_fitness_evaluations", "median_search_constraint_evaluations",
            "median_internal_constraint_evaluations", "median_reporting_constraint_evaluations",
            "median_total_constraint_evaluations",
            "median_candidate_checks", "median_repair_calls", "total_repair_calls",
            "total_repair_improved", "total_repair_unchanged", "total_repair_failed",
            "improvement_rate", "non_failure_rate"
        ]
        ws_s.append(sum_headers)
        for s in summary_list:
            med_ttff = s.get("median_time_to_first_feasible_seconds", s.get("time_to_first_feasible_median"))
            imp_rate = s.get("improvement_rate")
            non_fail = s.get("non_failure_rate")
            search_fit = s.get("median_search_fitness_evaluations", s.get("search_evaluations_median", 0))
            row = [
                s.get("method", ""),
                s.get("run_count", s.get("runs", 0)),
                round(s.get("feasible_rate", s.get("hard_feasible_rate", 0.0)), 4),
                s.get("median_final_hard", s.get("median_hard", 0.0)),
                s.get("median_final_soft", s.get("median_soft_penalty", 0.0)),
                round(s.get("mean_final_soft", s.get("mean_soft_penalty", 0.0)), 2),
                round(s.get("median_runtime_seconds", s.get("runtime_median", 0.0)), 4),
                round(med_ttff, 4) if med_ttff is not None and med_ttff != "N/A" else "",
                search_fit,
                s.get("median_search_constraint_evaluations", search_fit * 2),
                s.get("median_internal_constraint_evaluations", 0),
                s.get("median_reporting_constraint_evaluations", 0),
                s.get("median_total_constraint_evaluations", s.get("total_constraint_evaluations_median", 0)),
                s.get("median_candidate_checks", s.get("candidate_checks_median", 0)),
                s.get("median_repair_calls", s.get("repair_calls_median", 0)),
                s.get("total_repair_calls", s.get("repair_calls_total", 0)),
                s.get("total_repair_improved", s.get("repair_improved_total", 0)),
                s.get("total_repair_unchanged", s.get("repair_unchanged_total", 0)),
                s.get("total_repair_failed", s.get("repair_failed_total", 0)),
                round(imp_rate, 4) if imp_rate is not None else "",
                round(non_fail, 4) if non_fail is not None else "",
            ]
            ws_s.append(row)
        format_sheet(ws_s)


    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return str(out_path)



def export_metadata_to_json(
    metadata: dict,
    output_path: Union[str, Path],
) -> str:
    """Xuất dictionary metadata ra file JSON."""
    out_file = Path(output_path)
    out_file.parent.mkdir(parents=True, exist_ok=True)

    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)

    return str(out_file)
