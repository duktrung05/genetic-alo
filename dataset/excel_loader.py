"""Excel Dataset Loader Module.

Parses and validates timetable scheduling dataset from Microsoft Excel workbook format (.xlsx).
"""

import os
import json
import datetime
from typing import Dict, List, Set, Optional, Any
import openpyxl
from domain import Campus, Timeslot, Room, Lecturer, StudentGroup, Course, CourseSection, ConstraintDefinition
from .validator import DatasetValidator

class ExcelValidationError(ValueError):
    """Custom exception raised when Excel data fails validation rules with precise row/col context."""
    pass


class ExcelDatasetLoader:
    """Loader responsible for reading timetable scheduling entities from Excel workbooks."""

    # Ánh xạ chuẩn từ constraint_id trong Excel sang khóa kỹ thuật nội bộ
    SOFT_CONSTRAINT_KEY_BY_ID: Dict[str, str] = {
        "S1": "compact_student_schedule",
        "S2": "late_day_periods",
        "S3": "preferred_shift_mismatch",
        "S4": "room_seat_waste",
        "S5": "consecutive_cross_campus",
        "S6": "preferred_campus_mismatch",
        "S7": "student_home_campus_mismatch",
    }
    SUPPORTED_SOFT_IDS: frozenset = frozenset(SOFT_CONSTRAINT_KEY_BY_ID.keys())

    SHIFT_MAP = {
        "Sáng": "morning",
        "Chiều": "afternoon",
        "Tối": "evening",
        "MORNING": "morning",
        "AFTERNOON": "afternoon",
        "EVENING": "evening",
        "morning": "morning",
        "afternoon": "afternoon",
        "evening": "evening",
        "Morning": "morning",
        "Afternoon": "afternoon",
        "Evening": "evening",
    }

    VALID_SHIFTS = frozenset({"morning", "afternoon", "evening"})
    VALID_DIFFICULTIES = frozenset({"EASY", "MEDIUM", "HARD"})
    VALID_ROOM_TYPES = frozenset({"NORMAL", "LAB"})

    IGNORED_SHEETS = {
        "README",
        "BASELINE_SCHEDULE",
        "BEST_SCHEDULE",
        "SCHEDULE_BY_GROUP",
        "SCHEDULE_BY_LECTURER",
        "SCHEDULE_BY_ROOM",
    }

    @classmethod
    def _normalize_optional_str(cls, value) -> Optional[str]:
        """Convert cell value to stripped string or None if empty/NaN."""
        if value is None:
            return None
        text = str(value).strip()
        if not text or text.lower() == "nan":
            return None
        return text

    @classmethod
    def _required_str(cls, value, sheet: str, row_idx: int, col: str) -> str:
        text = cls._normalize_optional_str(value)
        if text is None:
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}': value must not be blank"
            )
        return text

    @classmethod
    def _parse_positive_int(cls, value, sheet: str, row_idx: int, col: str) -> int:
        if value is None or isinstance(value, bool):
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}', Value '{value}': "
                "Must be a positive integer"
            )
        try:
            number = float(value)
        except (ValueError, TypeError):
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}', Value '{value}': "
                "Must be a positive integer"
            )
        if not number.is_integer():
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}', Value '{value}': "
                "Fractional value not allowed; must be a positive integer"
            )
        result = int(number)
        if result < 1:
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}', Value '{value}': "
                "Must be > 0"
            )
        return result

    @classmethod
    def _parse_difficulty(cls, value, sheet: str, row_idx: int, col: str) -> str:
        difficulty = cls._required_str(value, sheet, row_idx, col).upper()
        if difficulty not in cls.VALID_DIFFICULTIES:
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}', Value '{value}': "
                f"Unknown difficulty '{value}'. Allowed values: {sorted(cls.VALID_DIFFICULTIES)}"
            )
        return difficulty

    @classmethod
    def _parse_time(cls, value, sheet: str, row_idx: int, col: str) -> str:
        try:
            if isinstance(value, datetime.datetime):
                parsed = value.time()
            elif isinstance(value, datetime.time):
                parsed = value
            else:
                text = cls._required_str(value, sheet, row_idx, col)
                parsed = datetime.time.fromisoformat(text)
        except (TypeError, ValueError):
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}', Value '{value}': "
                "Invalid time; expected HH:MM or HH:MM:SS"
            )
        return parsed.strftime("%H:%M:%S") if parsed.second else parsed.strftime("%H:%M")

    @classmethod
    def _parse_meetings_per_week(cls, value, row_idx: int) -> int:
        """Parse meetings_per_week from cell value.

        Rules:
          - blank/None  → default 1
          - integer     → use as-is (must be >= 1)
          - float N.0   → int(N) if fractional part is 0
          - non-integer float → ExcelValidationError
          - non-numeric → ExcelValidationError
        """
        if value is None:
            return 1
        if isinstance(value, bool):
            raise ExcelValidationError(
                f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'meetings_per_week', "
                f"Value '{value}': Must be a positive integer, not a boolean"
            )
        try:
            f = float(value)
        except (ValueError, TypeError):
            raise ExcelValidationError(
                f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'meetings_per_week', "
                f"Value '{value}': Must be a positive integer (e.g. 1, 2)"
            )
        if f != int(f):
            raise ExcelValidationError(
                f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'meetings_per_week', "
                f"Value '{value}': Non-integer value not allowed (got {f})"
            )
        result = int(f)
        if result < 1:
            raise ExcelValidationError(
                f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'meetings_per_week', "
                f"Value '{value}': Must be >= 1 (got {result})"
            )
        return result

    @classmethod
    def _parse_weight_int(cls, value, sheet: str, row_idx: int, col: str) -> int:
        """Parse constraint weight from cell: accepts int, float N.0, or str of those.

        Raises ExcelValidationError for:
          - blank / None
          - non-numeric string (e.g. 'abc')
          - non-integral float (e.g. 2.5)
          - negative value
        """
        if value is None:
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}': weight must not be blank"
            )
        raw_str = str(value).strip()
        if not raw_str or raw_str.lower() == "nan":
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}': weight must not be blank"
            )
        try:
            f = float(raw_str)
        except (ValueError, TypeError):
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}', "
                f"Value '{value}': weight must be a non-negative integer (e.g. 10, 5)"
            )
        if f != int(f):
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}', "
                f"Value '{value}': Non-integer weight not allowed (got {f}). "
                f"Only integers like 10, 5, 0 are accepted."
            )
        result = int(f)
        if result < 0:
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}', "
                f"Value '{value}': weight cannot be negative (got {result})"
            )
        return result

    @classmethod
    def _parse_enabled_bool(cls, value, sheet: str, row_idx: int, col: str) -> bool:
        """Parse enabled flag: True/False, 'true'/'false', 1/0.

        Raises ExcelValidationError for any unrecognized value.
        NOTE: bool('False') == True is a Python gotcha — we handle string comparison
        explicitly.
        """
        if value is None:
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}': enabled must not be blank"
            )
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            if value == 1:
                return True
            if value == 0:
                return False
            raise ExcelValidationError(
                f"Sheet '{sheet}', Row {row_idx}, Column '{col}', "
                f"Value '{value}': enabled must be True/False or 1/0"
            )
        s = str(value).strip().lower()
        if s in ("true", "1"):
            return True
        if s in ("false", "0"):
            return False
        raise ExcelValidationError(
            f"Sheet '{sheet}', Row {row_idx}, Column '{col}', "
            f"Value '{value}': enabled must be True/False, 'true'/'false', 1, or 0"
        )

    @classmethod
    def _parse_constraints_sheet(cls, wb) -> List[ConstraintDefinition]:
        """Parse the CONSTRAINTS sheet into a list of ConstraintDefinition objects.

        Validates:
        - Required headers present
        - No duplicate constraint_id
        - Soft constraint IDs must be in SUPPORTED_SOFT_IDS
        - weight: non-negative integer
        - enabled: boolean
        - constraint_type: SOFT or HARD (case-insensitive, normalized to upper)

        Returns empty list if CONSTRAINTS sheet is missing (non-Excel datasets).
        """
        if "CONSTRAINTS" not in wb.sheetnames:
            return []

        ws = wb["CONSTRAINTS"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Kiểm tra tiêu đề
        req_cols = ["constraint_id", "constraint_type", "constraint_name", "weight", "enabled"]
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        for col in req_cols:
            if col not in header:
                raise ExcelValidationError(
                    f"Sheet 'CONSTRAINTS', Row 1, Column '{col}': "
                    f"Required header column is missing. "
                    f"Expected columns: {req_cols}"
                )

        def _col(name: str):
            return header.index(name)

        seen_ids: Set[str] = set()
        definitions: List[ConstraintDefinition] = []

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or row[0] is None:
                continue
            # Chuẩn hóa hàng thành từ điển theo tiêu đề
            row_dict = {
                h: (row[i] if i < len(row) else None)
                for i, h in enumerate(header) if h
            }

            c_id_raw = cls._normalize_optional_str(row_dict.get("constraint_id"))
            if not c_id_raw:
                raise ExcelValidationError(
                    f"Sheet 'CONSTRAINTS', Row {row_idx}, Column 'constraint_id': "
                    f"constraint_id must not be empty"
                )
            c_id = c_id_raw.strip()

            if c_id in seen_ids:
                raise ExcelValidationError(
                    f"Sheet 'CONSTRAINTS', Row {row_idx}, Column 'constraint_id', "
                    f"Value '{c_id}': Duplicate constraint_id"
                )
            seen_ids.add(c_id)

            # Chuẩn hóa constraint_type
            c_type_raw = cls._normalize_optional_str(row_dict.get("constraint_type"))
            if not c_type_raw:
                raise ExcelValidationError(
                    f"Sheet 'CONSTRAINTS', Row {row_idx}, Column 'constraint_type': "
                    f"constraint_type must not be empty"
                )
            c_type = c_type_raw.strip().upper()
            if c_type not in ("SOFT", "HARD"):
                raise ExcelValidationError(
                    f"Sheet 'CONSTRAINTS', Row {row_idx}, Column 'constraint_type', "
                    f"Value '{c_type_raw}': Must be 'SOFT' or 'HARD'"
                )

            # Kiểm tra các mã ràng buộc mềm
            if c_type == "SOFT" and c_id not in cls.SUPPORTED_SOFT_IDS:
                raise ExcelValidationError(
                    f"Sheet 'CONSTRAINTS', Row {row_idx}, Column 'constraint_id', "
                    f"Value '{c_id}': Unsupported soft constraint_id='{c_id}'. "
                    f"Supported IDs: {sorted(cls.SUPPORTED_SOFT_IDS)}."
                )

            c_name = cls._normalize_optional_str(row_dict.get("constraint_name")) or c_id

            weight = cls._parse_weight_int(
                row_dict.get("weight"), "CONSTRAINTS", row_idx, "weight"
            )
            enabled = cls._parse_enabled_bool(
                row_dict.get("enabled"), "CONSTRAINTS", row_idx, "enabled"
            )

            # Giai đoạn 2.2 không thiết kế lại HardConstraintChecker. Các hàng HARD
            # là siêu dữ liệu khai báo/kiểm tra: trọng số trong sổ làm việc không điều
            # chỉnh từng phép kiểm tra cứng và mọi phép kiểm tra cứng vẫn bắt buộc.
            # Việc từ chối hàng HARD bị tắt ngăn sổ làm việc khai báo những ngữ nghĩa
            # mà bộ kiểm tra không thể đáp ứng.
            if c_type == "HARD" and not enabled:
                raise ExcelValidationError(
                    f"Sheet 'CONSTRAINTS', Row {row_idx}, Column 'enabled', "
                    f"Value '{row_dict.get('enabled')}': HARD constraints are always "
                    "enforced; enabled=False is not supported"
                )

            definitions.append(ConstraintDefinition(
                constraint_id=c_id,
                constraint_type=c_type,
                constraint_name=c_name,
                weight=weight,
                enabled=enabled,
            ))

        return definitions

    @classmethod
    def load(cls, excel_path: str = "data/instances/instance_easy.xlsx") -> dict:
        """Load and parse dataset dictionary from specified Excel file path.

        Args:
            excel_path: Path to the Excel workbook file.

        Returns:
            Dataset dictionary containing entity lists:
            {"timeslots", "rooms", "lecturers", "student_groups", "courses", "course_sections"}
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel dataset file not found at path: '{excel_path}'")

        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # 0. Đọc trang tính CONSTRAINTS (trước các trang tính khác)
        constraint_definitions = cls._parse_constraints_sheet(wb)

        # Danh mục CAMPUS là nguồn chuẩn cho mọi khóa ngoại về cơ sở.
        if "CAMPUSES" not in wb.sheetnames:
            raise ExcelValidationError(
                "Sheet 'CAMPUSES', Row 0, Column 'sheet': Sheet is missing from workbook"
            )
        ws_campus = wb["CAMPUSES"]
        rows_campus = list(ws_campus.iter_rows(values_only=True))
        if not rows_campus:
            raise ExcelValidationError("Sheet 'CAMPUSES', Row 1, Column 'all': Sheet is empty")
        header_campus = [str(h).strip() if h is not None else "" for h in rows_campus[0]]
        for col in ("campus_id", "campus_name"):
            if col not in header_campus:
                raise ExcelValidationError(
                    f"Sheet 'CAMPUSES', Row 1, Column '{col}': Required header column is missing"
                )
        campuses: List[Campus] = []
        campus_ids: Set[str] = set()
        for row_idx, row in enumerate(rows_campus[1:], start=2):
            if not row or all(value is None for value in row):
                continue
            campus_id = cls._required_str(
                row[header_campus.index("campus_id")], "CAMPUSES", row_idx, "campus_id"
            )
            campus_name = cls._required_str(
                row[header_campus.index("campus_name")], "CAMPUSES", row_idx, "campus_name"
            )
            if campus_id in campus_ids:
                raise ExcelValidationError(
                    f"Sheet 'CAMPUSES', Row {row_idx}, Column 'campus_id', "
                    f"Value '{campus_id}': Duplicate campus_id"
                )
            campus_ids.add(campus_id)
            campuses.append(Campus(id=campus_id, name=campus_name))

        # Ghi nhật ký các trang tính bị bỏ qua nếu có
        for sheet_name in wb.sheetnames:
            if sheet_name in cls.IGNORED_SHEETS:
                # Trang tính đầu ra/tài liệu được chủ ý bỏ qua
                pass

        # 1. Đọc TIMESLOTS
        if "TIMESLOTS" not in wb.sheetnames:
            raise ExcelValidationError("Sheet 'TIMESLOTS', Row 0, Column 'sheet': Sheet is missing from workbook")
        ws_ts = wb["TIMESLOTS"]
        rows_ts = list(ws_ts.iter_rows(values_only=True))
        if not rows_ts:
            raise ExcelValidationError("Sheet 'TIMESLOTS', Row 1, Column 'all': Sheet is empty")

        header_ts = [str(h).strip() if h is not None else "" for h in rows_ts[0]]
        req_ts_cols = ["timeslot_id", "day_name", "period_no", "shift", "start_time", "end_time"]
        for col in req_ts_cols:
            if col not in header_ts:
                raise ExcelValidationError(f"Sheet 'TIMESLOTS', Row 1, Column '{col}': Required header column is missing")

        code_to_ts_id: Dict[str, int] = {}
        timeslots: List[Timeslot] = []

        for row_idx, row in enumerate(rows_ts[1:], start=2):
            if not row or row[0] is None:
                continue
            code = cls._required_str(
                row[header_ts.index("timeslot_id")], "TIMESLOTS", row_idx, "timeslot_id"
            )
            if code in code_to_ts_id:
                raise ExcelValidationError(f"Sheet 'TIMESLOTS', Row {row_idx}, Column 'timeslot_id', Value '{code}': Duplicate timeslot_id")

            day_name = cls._required_str(
                row[header_ts.index("day_name")], "TIMESLOTS", row_idx, "day_name"
            )
            period_no = cls._parse_positive_int(
                row[header_ts.index("period_no")], "TIMESLOTS", row_idx, "period_no"
            )

            shift = cls._required_str(
                row[header_ts.index("shift")], "TIMESLOTS", row_idx, "shift"
            )
            session = cls.SHIFT_MAP.get(shift)
            if session is None:
                raise ExcelValidationError(
                    f"Sheet 'TIMESLOTS', Row {row_idx}, Column 'shift', Value '{shift}': "
                    f"Unknown shift '{shift}'"
                )
            ts_id = len(timeslots)
            code_to_ts_id[code] = ts_id
            start_t = cls._parse_time(
                row[header_ts.index("start_time")], "TIMESLOTS", row_idx, "start_time"
            )
            end_t = cls._parse_time(
                row[header_ts.index("end_time")], "TIMESLOTS", row_idx, "end_time"
            )
            if datetime.time.fromisoformat(start_t) >= datetime.time.fromisoformat(end_t):
                raise ExcelValidationError(
                    f"Sheet 'TIMESLOTS', Row {row_idx}: start_time '{start_t}' "
                    f"must be before end_time '{end_t}'"
                )
            if any(t.day == day_name and t.period == period_no for t in timeslots):
                raise ExcelValidationError(
                    f"Sheet 'TIMESLOTS', Row {row_idx}: Duplicate "
                    f"(day_name, period_no) ('{day_name}', {period_no})"
                )

            ts = Timeslot(
                id=ts_id,
                day=day_name,
                period=period_no,
                start_time=start_t,
                end_time=end_t,
                session=session,
                external_id=code,
            )
            timeslots.append(ts)

        # 2. Đọc ROOMS
        if "ROOMS" not in wb.sheetnames:
            raise ExcelValidationError("Sheet 'ROOMS', Row 0, Column 'sheet': Sheet is missing from workbook")
        ws_rm = wb["ROOMS"]
        rows_rm = list(ws_rm.iter_rows(values_only=True))
        if not rows_rm:
            raise ExcelValidationError("Sheet 'ROOMS', Row 1, Column 'all': Sheet is empty")

        header_rm = [str(h).strip() if h is not None else "" for h in rows_rm[0]]
        req_rm_cols = ["room_id", "capacity", "room_type"]
        for col in req_rm_cols:
            if col not in header_rm:
                raise ExcelValidationError(f"Sheet 'ROOMS', Row 1, Column '{col}': Required header column is missing")

        rooms: List[Room] = []
        room_ids: Set[str] = set()

        for row_idx, r in enumerate(rows_rm[1:], start=2):
            if not r or r[0] is None:
                continue
            r_id = str(r[header_rm.index("room_id")]).strip()
            if not r_id:
                raise ExcelValidationError(f"Sheet 'ROOMS', Row {row_idx}, Column 'room_id', Value '{r_id}': Invalid empty room_id")
            if r_id in room_ids:
                raise ExcelValidationError(f"Sheet 'ROOMS', Row {row_idx}, Column 'room_id', Value '{r_id}': Duplicate room_id")
            room_ids.add(r_id)

            cap = cls._parse_positive_int(
                r[header_rm.index("capacity")], "ROOMS", row_idx, "capacity"
            )

            r_type = str(r[header_rm.index("room_type")]).strip()
            if r_type not in cls.VALID_ROOM_TYPES:
                raise ExcelValidationError(f"Sheet 'ROOMS', Row {row_idx}, Column 'room_type', Value '{r_type}': Invalid room_type. Must be 'NORMAL' or 'LAB'")

            bld = str(r[header_rm.index("building")]).strip() if "building" in header_rm and r[header_rm.index("building")] else ""
            r_num = str(r[header_rm.index("room_number")]).strip() if "room_number" in header_rm and r[header_rm.index("room_number")] else r_id
            name = f"{bld}-{r_num}" if bld else r_num

            campus_id = cls._normalize_optional_str(
                r[header_rm.index("campus_id")] if "campus_id" in header_rm else None
            )
            if campus_id is not None and campus_id not in campus_ids:
                raise ExcelValidationError(
                    f"Sheet 'ROOMS', Row {row_idx}, Column 'campus_id', "
                    f"Value '{campus_id}': Unknown campus_id"
                )

            room = Room(
                id=r_id,
                name=name,
                capacity=cap,
                room_type=r_type,
                campus_id=campus_id,
            )
            rooms.append(room)

        # 3. Đọc LECTURER_AVAILABILITY
        lec_avail_map: Dict[str, Optional[frozenset]] = {}
        if "LECTURER_AVAILABILITY" not in wb.sheetnames:
            raise ExcelValidationError(
                "Sheet 'LECTURER_AVAILABILITY', Row 0, Column 'sheet': "
                "Sheet is missing from canonical workbook"
            )
        ws_avail = wb["LECTURER_AVAILABILITY"]
        rows_avail = list(ws_avail.iter_rows(values_only=True))
        if not rows_avail:
            raise ExcelValidationError(
                "Sheet 'LECTURER_AVAILABILITY', Row 1, Column 'all': Sheet is empty"
            )
        avail_header = [str(h).strip() if h is not None else "" for h in rows_avail[0]]
        if len(avail_header) < 2 or avail_header[:2] != ["lecturer_id", "lecturer_name"]:
            raise ExcelValidationError(
                "Sheet 'LECTURER_AVAILABILITY', Row 1: first columns must be "
                "'lecturer_id', 'lecturer_name'"
            )
        actual_timeslot_columns = avail_header[2:]
        # Mẫu Excel có thể chứa định dạng ở các cột không dùng, khiến openpyxl
        # trả về các tiêu đề None ở cuối. Đây là phần đệm của trang tính, không
        # phải cột thời gian rảnh. Ô trống ở giữa vẫn không hợp lệ.
        while actual_timeslot_columns and not actual_timeslot_columns[-1]:
            actual_timeslot_columns.pop()
        expected_timeslot_columns = list(code_to_ts_id)
        if any(not name for name in actual_timeslot_columns):
            raise ExcelValidationError(
                "Sheet 'LECTURER_AVAILABILITY', Row 1: blank timeslot columns "
                "are not allowed"
            )
        if len(actual_timeslot_columns) != len(set(actual_timeslot_columns)):
            raise ExcelValidationError(
                "Sheet 'LECTURER_AVAILABILITY', Row 1: duplicate timeslot columns"
            )
        missing_columns = sorted(set(expected_timeslot_columns) - set(actual_timeslot_columns))
        unknown_columns = sorted(set(actual_timeslot_columns) - set(expected_timeslot_columns))
        if missing_columns:
            raise ExcelValidationError(
                "Sheet 'LECTURER_AVAILABILITY', Row 1: missing timeslot columns: "
                f"{missing_columns}"
            )
        if unknown_columns:
            raise ExcelValidationError(
                "Sheet 'LECTURER_AVAILABILITY', Row 1: unknown timeslot columns: "
                f"{unknown_columns}"
            )
        if len(actual_timeslot_columns) != len(expected_timeslot_columns):
            raise ExcelValidationError(
                "Sheet 'LECTURER_AVAILABILITY', Row 1: timeslot-column count "
                f"must be exactly {len(expected_timeslot_columns)}, got "
                f"{len(actual_timeslot_columns)}"
            )

        for row_idx, r in enumerate(rows_avail[1:], start=2):
            if not r or all(value is None for value in r):
                continue
            l_id = cls._required_str(r[0], "LECTURER_AVAILABILITY", row_idx, "lecturer_id")
            if l_id in lec_avail_map:
                raise ExcelValidationError(
                    f"Sheet 'LECTURER_AVAILABILITY', Row {row_idx}, Column 'lecturer_id', "
                    f"Value '{l_id}': Duplicate lecturer_id"
                )
            avail_ts_set: Set[int] = set()
            for col_name in expected_timeslot_columns:
                col_idx = avail_header.index(col_name)
                is_avail = r[col_idx] if col_idx < len(r) else None
                if not isinstance(is_avail, bool):
                    raise ExcelValidationError(
                        f"Sheet 'LECTURER_AVAILABILITY', Row {row_idx}, "
                        f"Column '{col_name}', Value '{is_avail}': "
                        "Availability must be boolean True/False"
                    )
                if is_avail:
                    avail_ts_set.add(code_to_ts_id[col_name])
            lec_avail_map[l_id] = (
                None if len(avail_ts_set) == len(timeslots) else frozenset(avail_ts_set)
            )

        # 4. Đọc LECTURERS
        if "LECTURERS" not in wb.sheetnames:
            raise ExcelValidationError("Sheet 'LECTURERS', Row 0, Column 'sheet': Sheet is missing from workbook")
        ws_lec = wb["LECTURERS"]
        rows_lec = list(ws_lec.iter_rows(values_only=True))
        if not rows_lec:
            raise ExcelValidationError("Sheet 'LECTURERS', Row 1, Column 'all': Sheet is empty")

        header_lec = [str(h).strip() if h is not None else "" for h in rows_lec[0]]
        req_lec_cols = ["lecturer_id", "lecturer_name"]
        for col in req_lec_cols:
            if col not in header_lec:
                raise ExcelValidationError(f"Sheet 'LECTURERS', Row 1, Column '{col}': Required header column is missing")

        lecturers: List[Lecturer] = []
        lecturer_ids: Set[str] = set()

        for row_idx, r in enumerate(rows_lec[1:], start=2):
            if not r or r[0] is None:
                continue
            l_id = str(r[header_lec.index("lecturer_id")]).strip()
            if not l_id:
                raise ExcelValidationError(f"Sheet 'LECTURERS', Row {row_idx}, Column 'lecturer_id', Value '{l_id}': Invalid empty lecturer_id")
            if l_id in lecturer_ids:
                raise ExcelValidationError(f"Sheet 'LECTURERS', Row {row_idx}, Column 'lecturer_id', Value '{l_id}': Duplicate lecturer_id")
            lecturer_ids.add(l_id)

            l_name = str(r[header_lec.index("lecturer_name")]).strip()
            avail = lec_avail_map.get(l_id)
            if "limited_availability" in header_lec:
                limited = r[header_lec.index("limited_availability")]
                if not isinstance(limited, bool):
                    raise ExcelValidationError(
                        f"Sheet 'LECTURERS', Row {row_idx}, Column 'limited_availability', "
                        f"Value '{limited}': Must be boolean True/False"
                    )
                if limited != (avail is not None):
                    raise ExcelValidationError(
                        f"Sheet 'LECTURERS', Row {row_idx}, Column 'limited_availability': "
                        "Value does not match LECTURER_AVAILABILITY matrix"
                    )
            lec = Lecturer(
                id=l_id,
                name=l_name,
                available_timeslot_ids=avail,
            )
            lecturers.append(lec)

        unknown_availability_lecturers = sorted(set(lec_avail_map) - lecturer_ids)
        missing_availability_lecturers = sorted(lecturer_ids - set(lec_avail_map))
        if unknown_availability_lecturers:
            raise ExcelValidationError(
                "Sheet 'LECTURER_AVAILABILITY': unknown lecturer_id values: "
                f"{unknown_availability_lecturers}"
            )
        if missing_availability_lecturers:
            raise ExcelValidationError(
                "Sheet 'LECTURER_AVAILABILITY': missing lecturer rows: "
                f"{missing_availability_lecturers}"
            )

        # 5. Đọc STUDENT_GROUPS
        if "STUDENT_GROUPS" not in wb.sheetnames:
            raise ExcelValidationError("Sheet 'STUDENT_GROUPS', Row 0, Column 'sheet': Sheet is missing from workbook")
        ws_grp = wb["STUDENT_GROUPS"]
        rows_grp = list(ws_grp.iter_rows(values_only=True))
        if not rows_grp:
            raise ExcelValidationError("Sheet 'STUDENT_GROUPS', Row 1, Column 'all': Sheet is empty")

        header_grp = [str(h).strip() if h is not None else "" for h in rows_grp[0]]
        req_grp_cols = ["group_id", "group_name", "size"]
        for col in req_grp_cols:
            if col not in header_grp:
                raise ExcelValidationError(f"Sheet 'STUDENT_GROUPS', Row 1, Column '{col}': Required header column is missing")

        student_groups: List[StudentGroup] = []
        group_ids: Set[str] = set()

        for row_idx, r in enumerate(rows_grp[1:], start=2):
            if not r or r[0] is None:
                continue
            g_id = str(r[header_grp.index("group_id")]).strip()
            if not g_id:
                raise ExcelValidationError(f"Sheet 'STUDENT_GROUPS', Row {row_idx}, Column 'group_id', Value '{g_id}': Invalid empty group_id")
            if g_id in group_ids:
                raise ExcelValidationError(f"Sheet 'STUDENT_GROUPS', Row {row_idx}, Column 'group_id', Value '{g_id}': Duplicate group_id")
            group_ids.add(g_id)

            g_name = str(r[header_grp.index("group_name")]).strip()
            size = cls._parse_positive_int(
                r[header_grp.index("size")], "STUDENT_GROUPS", row_idx, "size"
            )

            home_campus_id = cls._normalize_optional_str(
                r[header_grp.index("home_campus_id")] if "home_campus_id" in header_grp else None
            )
            if home_campus_id is not None and home_campus_id not in campus_ids:
                raise ExcelValidationError(
                    f"Sheet 'STUDENT_GROUPS', Row {row_idx}, Column 'home_campus_id', "
                    f"Value '{home_campus_id}': Unknown campus_id"
                )

            grp = StudentGroup(
                id=g_id,
                name=g_name,
                student_count=size,
                home_campus_id=home_campus_id,
            )
            student_groups.append(grp)

        # 6. Đọc COURSES
        if "COURSES" not in wb.sheetnames:
            raise ExcelValidationError("Sheet 'COURSES', Row 0, Column 'sheet': Sheet is missing from workbook")
        ws_crs = wb["COURSES"]
        rows_crs = list(ws_crs.iter_rows(values_only=True))
        if not rows_crs:
            raise ExcelValidationError("Sheet 'COURSES', Row 1, Column 'all': Sheet is empty")

        header_crs = [str(h).strip() if h is not None else "" for h in rows_crs[0]]
        req_crs_cols = ["course_id", "course_code", "course_name", "difficulty"]
        for col in req_crs_cols:
            if col not in header_crs:
                raise ExcelValidationError(f"Sheet 'COURSES', Row 1, Column '{col}': Required header column is missing")

        courses: List[Course] = []
        course_ids: Set[str] = set()
        course_codes: Set[str] = set()
        course_by_id: Dict[str, Course] = {}
        difficulty_by_course_id: Dict[str, str] = {}

        for row_idx, r in enumerate(rows_crs[1:], start=2):
            if not r or r[0] is None:
                continue
            c_id = str(r[header_crs.index("course_id")]).strip()
            if not c_id:
                raise ExcelValidationError(f"Sheet 'COURSES', Row {row_idx}, Column 'course_id', Value '{c_id}': Invalid empty course_id")
            if c_id in course_ids:
                raise ExcelValidationError(f"Sheet 'COURSES', Row {row_idx}, Column 'course_id', Value '{c_id}': Duplicate course_id")
            course_ids.add(c_id)

            course_code = cls._required_str(
                r[header_crs.index("course_code")], "COURSES", row_idx, "course_code"
            )
            if course_code in course_codes:
                raise ExcelValidationError(
                    f"Sheet 'COURSES', Row {row_idx}, Column 'course_code', "
                    f"Value '{course_code}': Duplicate course_code"
                )
            course_codes.add(course_code)
            c_name = cls._required_str(
                r[header_crs.index("course_name")], "COURSES", row_idx, "course_name"
            )
            diff = cls._parse_difficulty(
                r[header_crs.index("difficulty")], "COURSES", row_idx, "difficulty"
            )
            if "default_room_type" in header_crs:
                default_room_type = cls._required_str(
                    r[header_crs.index("default_room_type")],
                    "COURSES", row_idx, "default_room_type",
                )
                if default_room_type not in cls.VALID_ROOM_TYPES:
                    raise ExcelValidationError(
                        f"Sheet 'COURSES', Row {row_idx}, Column 'default_room_type', "
                        f"Value '{default_room_type}': Invalid room_type"
                    )
            crs = Course(
                course_id=c_id,
                name=c_name,
                credits=3,
                is_difficult=(diff == "HARD"),
                course_code=course_code,
            )
            courses.append(crs)
            course_by_id[c_id] = crs
            difficulty_by_course_id[c_id] = diff

        # 7. Đọc COURSE_SECTIONS
        if "COURSE_SECTIONS" not in wb.sheetnames:
            raise ExcelValidationError("Sheet 'COURSE_SECTIONS', Row 0, Column 'sheet': Sheet is missing from workbook")
        ws_sec = wb["COURSE_SECTIONS"]
        rows_sec = list(ws_sec.iter_rows(values_only=True))
        if not rows_sec:
            raise ExcelValidationError("Sheet 'COURSE_SECTIONS', Row 1, Column 'all': Sheet is empty")

        header_sec = [str(h).strip() if h is not None else "" for h in rows_sec[0]]
        req_sec_cols = [
            "section_id", "class_code", "course_id", "course_code",
            "lecturer_id", "student_group_id", "student_count",
            "required_room_type", "duration_periods",
        ]
        for col in req_sec_cols:
            if col not in header_sec:
                raise ExcelValidationError(f"Sheet 'COURSE_SECTIONS', Row 1, Column '{col}': Required header column is missing")

        course_sections: List[CourseSection] = []
        section_ids: Set[str] = set()
        class_codes: Set[str] = set()
        lecturer_by_id = {lecturer.id: lecturer for lecturer in lecturers}
        group_by_id = {group.id: group for group in student_groups}

        for row_idx, r in enumerate(rows_sec[1:], start=2):
            if not r or r[0] is None:
                continue
            sec_id = str(r[header_sec.index("section_id")]).strip()
            if not sec_id:
                raise ExcelValidationError(f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'section_id', Value '{sec_id}': Invalid empty section_id")
            if sec_id in section_ids:
                raise ExcelValidationError(f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'section_id', Value '{sec_id}': Duplicate section_id")
            section_ids.add(sec_id)

            class_code = cls._required_str(
                r[header_sec.index("class_code")],
                "COURSE_SECTIONS", row_idx, "class_code",
            )
            if class_code in class_codes:
                raise ExcelValidationError(
                    f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'class_code', "
                    f"Value '{class_code}': Duplicate class_code"
                )
            class_codes.add(class_code)

            c_id = str(r[header_sec.index("course_id")]).strip()
            if c_id not in course_ids:
                raise ExcelValidationError(f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'course_id', Value '{c_id}': Referenced course_id does not exist in COURSES sheet")
            section_course_code = cls._required_str(
                r[header_sec.index("course_code")],
                "COURSE_SECTIONS", row_idx, "course_code",
            )
            if section_course_code != course_by_id[c_id].course_code:
                raise ExcelValidationError(
                    f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'course_code', "
                    f"Value '{section_course_code}': does not match COURSES master "
                    f"value '{course_by_id[c_id].course_code}'"
                )

            l_id = str(r[header_sec.index("lecturer_id")]).strip()
            if l_id not in lecturer_ids:
                raise ExcelValidationError(f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'lecturer_id', Value '{l_id}': Referenced lecturer_id does not exist in LECTURERS sheet")

            g_id = str(r[header_sec.index("student_group_id")]).strip()
            if g_id not in group_ids:
                raise ExcelValidationError(f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'student_group_id', Value '{g_id}': Referenced student_group_id does not exist in STUDENT_GROUPS sheet")

            st_cnt = cls._parse_positive_int(
                r[header_sec.index("student_count")],
                "COURSE_SECTIONS", row_idx, "student_count",
            )

            req_type = str(r[header_sec.index("required_room_type")]).strip()
            if req_type not in cls.VALID_ROOM_TYPES:
                raise ExcelValidationError(f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'required_room_type', Value '{req_type}': Invalid required_room_type. Must be 'NORMAL' or 'LAB'")

            dur = cls._parse_positive_int(
                r[header_sec.index("duration_periods")],
                "COURSE_SECTIONS", row_idx, "duration_periods",
            )

            c_name = course_by_id[c_id].name
            if "course_name" in header_sec:
                supplied_course_name = cls._required_str(
                    r[header_sec.index("course_name")],
                    "COURSE_SECTIONS", row_idx, "course_name",
                )
                if supplied_course_name != c_name:
                    raise ExcelValidationError(
                        f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'course_name': "
                        "does not match COURSES master"
                    )
            diff = difficulty_by_course_id[c_id]
            if "difficulty" in header_sec:
                supplied_difficulty = cls._parse_difficulty(
                    r[header_sec.index("difficulty")],
                    "COURSE_SECTIONS", row_idx, "difficulty",
                )
                if supplied_difficulty != difficulty_by_course_id[c_id]:
                    raise ExcelValidationError(
                        f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'difficulty': "
                        "does not match COURSES master"
                    )
                diff = supplied_difficulty

            # Các trường tùy chọn về ưu tiên
            preferred_campus_id = cls._normalize_optional_str(
                r[header_sec.index("preferred_campus_id")] if "preferred_campus_id" in header_sec else None
            )
            if preferred_campus_id is not None and preferred_campus_id not in campus_ids:
                raise ExcelValidationError(
                    f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'preferred_campus_id', "
                    f"Value '{preferred_campus_id}': Unknown campus_id"
                )

            if "student_group_name" in header_sec:
                supplied_group_name = cls._required_str(
                    r[header_sec.index("student_group_name")],
                    "COURSE_SECTIONS", row_idx, "student_group_name",
                )
                if supplied_group_name != group_by_id[g_id].name:
                    raise ExcelValidationError(
                        f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'student_group_name': "
                        "does not match STUDENT_GROUPS master"
                    )
            if "lecturer_name" in header_sec:
                supplied_lecturer_name = cls._required_str(
                    r[header_sec.index("lecturer_name")],
                    "COURSE_SECTIONS", row_idx, "lecturer_name",
                )
                if supplied_lecturer_name != lecturer_by_id[l_id].name:
                    raise ExcelValidationError(
                        f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'lecturer_name': "
                        "does not match LECTURERS master"
                    )

            # Chuẩn hóa preferred_shift qua SHIFT_MAP rồi kiểm tra
            raw_shift = cls._normalize_optional_str(
                r[header_sec.index("preferred_shift")] if "preferred_shift" in header_sec else None
            )
            preferred_shift: Optional[str] = None
            if raw_shift is not None:
                preferred_shift = cls.SHIFT_MAP.get(raw_shift, raw_shift.lower())
                if preferred_shift not in cls.VALID_SHIFTS:
                    raise ExcelValidationError(
                        f"Sheet 'COURSE_SECTIONS', Row {row_idx}, Column 'preferred_shift', "
                        f"Value '{raw_shift}': Invalid shift. Allowed values: "
                        f"{sorted(cls.VALID_SHIFTS)}"
                    )

            meetings_per_week = cls._parse_meetings_per_week(
                r[header_sec.index("meetings_per_week")] if "meetings_per_week" in header_sec else None,
                row_idx,
            )

            sec = CourseSection(
                section_id=sec_id,
                course_id=c_id,
                course_name=c_name,
                lecturer_id=l_id,
                group_id=g_id,
                student_count=st_cnt,
                is_difficult=(diff == "HARD"),
                required_room_type=req_type,
                duration_periods=dur,
                preferred_campus_id=preferred_campus_id,
                preferred_shift=preferred_shift,
                meetings_per_week=meetings_per_week,
                class_code=class_code,
            )
            course_sections.append(sec)

        return {
            "campuses": campuses,
            "timeslots": timeslots,
            "rooms": rooms,
            "lecturers": lecturers,
            "student_groups": student_groups,
            "courses": courses,
            "course_sections": course_sections,
            "constraints": constraint_definitions,
        }

    @classmethod
    def load_and_validate(cls, excel_path: str = "data/instances/instance_easy.xlsx") -> dict:
        """Load dataset from Excel file and run DatasetValidator validation.

        Args:
            excel_path: Path to the Excel workbook file.

        Returns:
            Validated dataset dictionary.
        """
        dataset = cls.load(excel_path)
        DatasetValidator.validate(dataset)
        return dataset

    @classmethod
    def export_normalized_json(cls, dataset: dict, output_path: str = "outputs/datasets/instance_easy.normalized.json") -> str:
        """Serialize dataset dictionary into normalized JSON snapshot file.

        Args:
            dataset: Dataset dictionary.
            output_path: Target JSON file path.

        Returns:
            Absolute path of written JSON snapshot file.
        """
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        data = {
            "timeslots": [
                {
                    "id": t.id,
                    "external_id": getattr(t, "external_id", None),
                    "day": t.day,
                    "period": t.period,
                    "start_time": t.start_time,
                    "end_time": t.end_time,
                    "session": t.session,
                }
                for t in dataset["timeslots"]
            ],
            "campuses": [
                {"id": campus.id, "name": campus.name}
                for campus in dataset.get("campuses", [])
            ],
            "rooms": [
                {
                    "id": r.id,
                    "name": r.name,
                    "capacity": r.capacity,
                    "room_type": getattr(r, "room_type", "NORMAL"),
                    "campus_id": getattr(r, "campus_id", None),
                }
                for r in dataset["rooms"]
            ],
            "lecturers": [
                {
                    "id": l.id,
                    "name": l.name,
                    "available_timeslot_ids": (
                        sorted(list(l.available_timeslot_ids))
                        if l.available_timeslot_ids is not None
                        else None
                    ),
                }
                for l in dataset.get("lecturers", [])
            ],
            "student_groups": [
                {
                    "id": g.id,
                    "name": g.name,
                    "student_count": g.student_count,
                    "home_campus_id": getattr(g, "home_campus_id", None),
                }
                for g in dataset.get("student_groups", [])
            ],
            "courses": [
                {
                    "course_id": c.course_id,
                    "course_code": getattr(c, "course_code", None),
                    "name": c.name,
                    "credits": getattr(c, "credits", 3),
                    "is_difficult": getattr(c, "is_difficult", False),
                }
                for c in dataset.get("courses", [])
            ],
            "course_sections": [
                {
                    "section_id": s.section_id,
                    "class_code": getattr(s, "class_code", None),
                    "course_id": s.course_id,
                    "course_name": s.course_name,
                    "lecturer_id": s.lecturer_id,
                    "group_id": s.group_id,
                    "student_count": s.student_count,
                    "is_difficult": getattr(s, "is_difficult", False),
                    "required_room_type": getattr(s, "required_room_type", "NORMAL"),
                    "duration_periods": getattr(s, "duration_periods", 1),
                    "preferred_campus_id": getattr(s, "preferred_campus_id", None),
                    "preferred_shift": getattr(s, "preferred_shift", None),
                    "meetings_per_week": getattr(s, "meetings_per_week", 1),
                }
                for s in dataset["course_sections"]
            ],
            "constraints": [
                {
                    "constraint_id": c.constraint_id,
                    "constraint_type": c.constraint_type,
                    "constraint_name": c.constraint_name,
                    "weight": c.weight,
                    "enabled": c.enabled,
                }
                for c in dataset.get("constraints", [])
            ],
        }

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return os.path.abspath(output_path)

    @classmethod
    def load_normalized_json(cls, json_path: str = "outputs/datasets/instance_easy.normalized.json") -> dict:
        """Deserialize dataset dictionary from normalized JSON snapshot file."""
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON snapshot file not found: '{json_path}'")

        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        timeslots = [Timeslot(**t) for t in data["timeslots"]]
        rooms = [Room(**r) for r in data["rooms"]]
        if "campuses" in data:
            campuses = [Campus(**campus) for campus in data["campuses"]]
        else:
            # Các bản chụp chuẩn hóa cũ có trước danh mục cơ sở. Duy trì tương
            # thích bằng cách chỉ tạo các mã cơ sở đã lưu trên phòng; Excel chuẩn
            # không bao giờ dùng phương án dự phòng này.
            legacy_ids = sorted({room.campus_id for room in rooms if room.campus_id})
            campuses = [Campus(id=campus_id, name=campus_id) for campus_id in legacy_ids]
        lecturers = [
            Lecturer(
                id=l["id"],
                name=l["name"],
                available_timeslot_ids=(
                    frozenset(l["available_timeslot_ids"])
                    if l["available_timeslot_ids"] is not None
                    else None
                ),
            )
            for l in data["lecturers"]
        ]
        student_groups = [StudentGroup(**g) for g in data["student_groups"]]
        courses = [
            Course(**({**course, "course_code": course.get("course_code") or course["course_id"]}))
            for course in data.get("courses", [])
        ]
        course_sections = [
            CourseSection(**({**section, "class_code": section.get("class_code")}))
            for section in data["course_sections"]
        ]
        constraints = [
            ConstraintDefinition(**c)
            for c in data.get("constraints", [])
        ]

        dataset = {
            "campuses": campuses,
            "timeslots": timeslots,
            "rooms": rooms,
            "lecturers": lecturers,
            "student_groups": student_groups,
            "courses": courses,
            "course_sections": course_sections,
            "constraints": constraints,
        }
        DatasetValidator.validate(dataset)
        return dataset

    @classmethod
    def analyze_excel_dataset(cls, dataset: dict) -> dict:
        """Compute detailed statistics report of Excel dataset entities and constraint tightness."""
        sections = dataset["course_sections"]
        lecturers = dataset["lecturers"]
        groups = dataset["student_groups"]
        rooms = dataset["rooms"]
        timeslots = dataset["timeslots"]

        # Phân bố thời lượng
        dur_dist = {2: 0, 3: 0, 4: 0}
        lab_count = 0
        for s in sections:
            dur = getattr(s, "duration_periods", 1)
            dur_dist[dur] = dur_dist.get(dur, 0) + 1
            if getattr(s, "required_room_type", "NORMAL") == "LAB":
                lab_count += 1

        # Khối lượng giảng dạy theo giảng viên (lớp học phần và tổng số tiết)
        lec_load = {}
        for l in lecturers:
            l_secs = [s for s in sections if s.lecturer_id == l.id]
            total_p = sum(
                getattr(s, "duration_periods", 1)
                * getattr(s, "meetings_per_week", 1)
                for s in l_secs
            )
            lec_load[l.id] = {"name": l.name, "sections": len(l_secs), "total_periods": total_p}

        # Khối lượng học tập theo nhóm sinh viên
        grp_load = {}
        for g in groups:
            g_secs = [s for s in sections if s.group_id == g.id]
            total_p = sum(
                getattr(s, "duration_periods", 1)
                * getattr(s, "meetings_per_week", 1)
                for s in g_secs
            )
            grp_load[g.id] = {"name": g.name, "sections": len(g_secs), "total_periods": total_p}

        # Số lượng ứng viên nhỏ nhất / lớn nhất / trung bình
        from .timeslot_factory import get_occupied_periods, is_valid_period_block
        from collections import defaultdict
        day_period_to_ts_id = {(ts.day, ts.period): ts.id for ts in timeslots}
        day_available_periods = defaultdict(set)
        for ts in timeslots:
            day_available_periods[ts.day].add(ts.period)

        cand_counts = []
        for s in sections:
            dur = getattr(s, "duration_periods", 1)
            req_type = getattr(s, "required_room_type", "NORMAL")
            valid_rms = [r for r in rooms if r.capacity >= s.student_count and getattr(r, "room_type", "NORMAL") == req_type]

            lec = next((l for l in lecturers if l.id == s.lecturer_id), None)
            avail_ts = lec.available_timeslot_ids if lec else None

            valid_ts = [
                t for t in timeslots
                if is_valid_period_block(t.period, dur, day_available_periods.get(t.day))
                and (avail_ts is None or all(day_period_to_ts_id.get((t.day, p)) in avail_ts for p in get_occupied_periods(t.period, dur)))
            ]

            cand_counts.append(len(valid_rms) * len(valid_ts))

        mean_cand = sum(cand_counts) / len(cand_counts) if cand_counts else 0.0

        return {
            "duration_distribution": dur_dist,
            "lab_section_count": lab_count,
            "lecturer_load": lec_load,
            "group_load": grp_load,
            "candidate_counts": {
                "min": min(cand_counts) if cand_counts else 0,
                "max": max(cand_counts) if cand_counts else 0,
                "mean": round(mean_cand, 2),
            },
        }
